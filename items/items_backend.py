import requests
import re
from datetime import datetime
import sys
import os
import html

# Ensure root directory is in path to import database_manager
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import database_manager
except ImportError:
    print("️ Warning: Could not import database_manager. SQLite sync will be skipped.")
    database_manager = None


TALLY_URL = "http://localhost:9000"

# ----------------------------------------------------------
# HELPERS
# ----------------------------------------------------------

def _unescape_basic(text: str) -> str:
    """
    Unescape the common XML entities without converting numeric refs like &#4;.
    """
    if not text:
        return ""
    return (
        text.replace("&amp;", "&")
        .replace("&apos;", "'")
        .replace("&quot;", '"')
        .replace("&lt;", "<")
        .replace("&gt;", ">")
    )

def extract_field(xml, tag):
    if not xml:
        return ""
    m = re.search(rf"<{tag}>(.*?)</{tag}>", xml, re.DOTALL)
    return m.group(1).strip() if m else ""

def extract_number(text):
    if not text:
        return ""
    m = re.search(r"-?\d+(\.\d+)?", text)
    return float(m.group()) if m else ""

def extract_number_and_unit(text):
    if not text:
        return "", ""
    num = extract_number(text)
    unit = re.sub(r"[-\d.\s/]", "", text)
    return num, unit

def parse_date(val):
    try:
        return datetime.strptime(val, "%Y%m%d")
    except:
        return datetime.min

def pick_latest(blocks):
    latest = ""
    latest_dt = datetime.min
    for b in blocks:
        dt = parse_date(extract_field(b, "APPLICABLEFROM"))
        if dt >= latest_dt:
            latest_dt = dt
            latest = b
    return latest

def extract_supply_type(block, latest_gst, group):
    # 1. Item GST details
    val = extract_field(latest_gst, "SUPPLYTYPE")
    if val:
        return val

    # 2. Item master level
    val = extract_field(block, "GSTTYPEOFSUPPLY")
    if val:
        return val

    # 3. Group level
    return group.get("supply_type", "")

# ----------------------------------------------------------
# GST RATE LOGIC
# ----------------------------------------------------------

def calculate_gst_rate(gst_block):
    if not gst_block:
        return ""

    rates = re.findall(r"<RATEDETAILS.LIST>(.*?)</RATEDETAILS.LIST>", gst_block, re.DOTALL)

    igst = 0
    cgst = 0
    sgst = 0

    for r in rates:
        head = extract_field(r, "GSTRATEDUTYHEAD")
        rate = extract_number(extract_field(r, "GSTRATE")) or 0

        if head == "IGST":
            igst = rate
        elif head == "CGST":
            cgst = rate
        elif head == "SGST/UTGST":
            sgst = rate

    if igst:
        return igst
    if cgst or sgst:
        return cgst + sgst

    return ""


def normalize_applicability(value):
    """
    Converts Tally values like:
    '&#4; Applicable' -> 'Applicable'
    '&#4; Not Applicable' -> 'Non Applicable'
    '' -> ''
    """
    if not value:
        return ""

    val = value.replace("&#4;", "").strip().lower()

    if "applicable" in val and "not" not in val:
        return "Applicable"
    if "not applicable" in val or "non applicable" in val:
        return "Non Applicable"

    return value.strip()


# ----------------------------------------------------------
# FETCH STOCK GROUPS (HSN + GST MASTER)
# ----------------------------------------------------------

def parse_stock_groups_from_xml(xml: str):
    groups = {}

    blocks = re.findall(
        r'<STOCKGROUP NAME="([^"]*)"[^>]*>(.*?)</STOCKGROUP>',
        xml or "",
        re.DOTALL
    )

    for name, block in blocks:
        # HSN
        hsn_blocks = re.findall(r"<HSNDETAILS.LIST>(.*?)</HSNDETAILS.LIST>", block, re.DOTALL)
        latest_hsn = pick_latest(hsn_blocks)

        # GST
        gst_blocks = re.findall(r"<GSTDETAILS.LIST>(.*?)</GSTDETAILS.LIST>", block, re.DOTALL)
        latest_gst = pick_latest(gst_blocks)

        groups[_unescape_basic(name)] = {
            "hsn_source": _unescape_basic(extract_field(latest_hsn, "SRCOFHSNDETAILS")),
            "hsn": _unescape_basic(extract_field(latest_hsn, "HSNCODE")),
            "description": _unescape_basic(extract_field(latest_hsn, "HSN")),

            "gst_rate": calculate_gst_rate(latest_gst),
            "taxability": _unescape_basic(extract_field(latest_gst, "TAXABILITY")),
            "supply_type": _unescape_basic(extract_field(latest_gst, "SUPPLYTYPE")),
        }

    return groups

def fetch_stock_groups():

    xml_req = """
<ENVELOPE>
 <HEADER><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>
 <BODY>
  <EXPORTDATA>
   <REQUESTDESC>
    <REPORTNAME>List of Accounts</REPORTNAME>
    <STATICVARIABLES>
     <ACCOUNTTYPE>StockGroups</ACCOUNTTYPE>
    </STATICVARIABLES>
   </REQUESTDESC>
  </EXPORTDATA>
 </BODY>
	</ENVELOPE>
"""

    res = requests.post(TALLY_URL, data=xml_req.encode(), timeout=120)
    return parse_stock_groups_from_xml(res.text)

# ----------------------------------------------------------
# FETCH STOCK ITEMS (WITH INHERITANCE)
# ----------------------------------------------------------


def parse_stock_items_from_xml(xml: str, groups, save_to_db: bool = True):
    
    # Initialize DB if possible
    if save_to_db and database_manager:
        database_manager.init_db()

    items = []

    blocks = re.findall(
        r'<STOCKITEM NAME="([^"]*)"[^>]*>(.*?)</STOCKITEM>',
        xml or "", re.DOTALL
    )

    for name, block in blocks:

        parent = extract_field(block, "PARENT")
        parent = _unescape_basic(parent)
        group = groups.get(parent, {}) if isinstance(groups, dict) else {}

        # Extract HSN details from STOCKITEM itself (item-level)
        hsn_blocks = re.findall(r"<HSNDETAILS.LIST>(.*?)</HSNDETAILS.LIST>", block, re.DOTALL)
        latest_hsn = pick_latest(hsn_blocks)

        # Extract GST details from STOCKITEM itself (item-level)
        gst_blocks = re.findall(r"<GSTDETAILS.LIST>(.*?)</GSTDETAILS.LIST>", block, re.DOTALL)
        latest_gst = pick_latest(gst_blocks)

        # Item-level data (if available)
        item_hsn_source = extract_field(latest_hsn, "SRCOFHSNDETAILS")
        item_hsn = extract_field(latest_hsn, "HSNCODE")
        item_description = extract_field(latest_hsn, "HSN")
        item_gst_rate = calculate_gst_rate(latest_gst)
        item_taxability = extract_field(latest_gst, "TAXABILITY")
        
        # Type of Supply is at STOCKITEM level, not in GSTDETAILS
        item_supply_type = extract_supply_type(block, latest_gst, group)

        gst_rate = item_gst_rate if item_gst_rate else group.get("gst_rate", "")
        
        # Additional fields requested
        raw_app = extract_field(latest_gst, "GSTAPPLICABLE") or extract_field(block, "GSTAPPLICABLE")
        
        item_gst_rate_source = extract_field(latest_gst, "SRCOFGSTDETAILS")  # GST Rate Details source
        item_rate_of_duty = extract_field(block, "BASICRATEOFEXCISE")  # Rate of Duty
        if raw_app:
            item_gst_applicable = normalize_applicability(raw_app)
        else:
            has_item_details = bool(latest_gst.strip() or latest_hsn.strip())
            has_group_details = bool(group.get("hsn") or group.get("gst_rate") or group.get("taxability") or group.get("supply_type"))
            item_gst_applicable = "Applicable" if (has_item_details or has_group_details) else "Non Applicable"

        # Use item-level data if available, otherwise fall back to group-level
        hsn_source = item_hsn_source if item_hsn_source else group.get("hsn_source", "")
        hsn = item_hsn if item_hsn else group.get("hsn", "")
        description = item_description if item_description else group.get("description", "")
        
        taxability = item_taxability if item_taxability else group.get("taxability", "")
        if item_supply_type:
            supply_type = item_supply_type
        elif group.get("supply_type"):
            supply_type = group.get("supply_type")
        else:
            supply_type = "Goods"
        
        # New fields (no group fallback needed as they're item-specific)
        gst_applicable = item_gst_applicable
        gst_rate_source = item_gst_rate_source
        rate_of_duty = item_rate_of_duty

        opening_balance_raw = _unescape_basic(extract_field(block, "OPENINGBALANCE"))
        opening_rate_raw = _unescape_basic(extract_field(block, "OPENINGRATE"))
        opening_value_raw = _unescape_basic(extract_field(block, "OPENINGVALUE"))

        qty, qty_unit = extract_number_and_unit(opening_balance_raw)
        rate, rate_unit = extract_number_and_unit(opening_rate_raw)
        value = extract_number(opening_value_raw)
        value = abs(value) if value else ""

        # Custom field for Category if it exists, or standard CATEGORY tag
        item_category_raw = extract_field(block, "CATEGORY")
        item_category_norm = normalize_applicability(item_category_raw)
        item_category = "" if item_category_norm.lower() in ("non applicable", "not applicable") else _unescape_basic(item_category_norm)

        item_data = {
            "name": _unescape_basic(name),
            "group": parent,
            "category": item_category,
            "unit": _unescape_basic(extract_field(block, "BASEUNITS")),

            # HSN/GST details (item-level first, then group-level)
            "hsn_source": _unescape_basic(hsn_source),
            "hsn": _unescape_basic(hsn),
            "description": _unescape_basic(description),

            "gst_rate": gst_rate,
            "taxability": _unescape_basic(taxability),
            "supply_type": _unescape_basic(supply_type),
            
            # Additional fields
            "gst_applicable": _unescape_basic(gst_applicable),
            "gst_rate_source": _unescape_basic(gst_rate_source),
            "rate_of_duty": rate_of_duty,

            "qty": qty,
            "qty_unit": _unescape_basic(qty_unit),
            "rate": rate,
            "rate_unit": _unescape_basic(rate_unit),
            "value": value,
        }
        
        # ------------------------------------------------
        # SAVE TO SQLITE
        # ------------------------------------------------
        if save_to_db and database_manager:
            db_data = {
                "name": item_data["name"],
                "group_name": item_data["group"],
                "category": item_data["category"],
                "unit": item_data["unit"],
                "hsn_source": item_data.get("hsn_source", ""),
                "hsn": item_data.get("hsn", ""),
                "description": item_data.get("description", ""),
                
                "gst_applicable": item_data.get("gst_applicable", ""),
                "gst_rate_source": item_data.get("gst_rate_source", ""),
                "gst_rate": item_data.get("gst_rate", 0) or 0,
                "taxability": item_data.get("taxability", ""),
                "supply_type": item_data.get("supply_type", ""),
                "rate_of_duty": item_data.get("rate_of_duty", 0) or 0,
                
                "qty": item_data.get("qty", 0) or 0,
                "qty_unit": item_data.get("qty_unit", ""),
                "rate": item_data.get("rate", 0) or 0,
                "rate_unit": item_data.get("rate_unit", ""),
                "value": item_data.get("value", 0) or 0
            }
            database_manager.insert_or_update_item(db_data)

        items.append(item_data)

    return items


def fetch_stock_items(groups):

    xml_req = """
<ENVELOPE>
 <HEADER><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>
 <BODY>
  <EXPORTDATA>
   <REQUESTDESC>
    <REPORTNAME>List of Accounts</REPORTNAME>
    <STATICVARIABLES>
     <ACCOUNTTYPE>StockItems</ACCOUNTTYPE>
    </STATICVARIABLES>
   </REQUESTDESC>
  </EXPORTDATA>
 </BODY>
</ENVELOPE>
"""

    res = requests.post(TALLY_URL, data=xml_req.encode(), timeout=120)
    return parse_stock_items_from_xml(res.text, groups, save_to_db=True)


def import_items_from_exported_xml(xml_text: str, save_to_db: bool = True):
    """
    Import Stock Groups + Stock Items from a Tally-exported XML file.
    Returns counts and basic stats.
    """
    groups = parse_stock_groups_from_xml(xml_text or "")
    items = parse_stock_items_from_xml(xml_text or "", groups, save_to_db=save_to_db)
    return {
        "groups_count": len(groups),
        "items_count": len(items),
    }

# ----------------------------------------------------------
# SEARCH & DISPLAY
# ----------------------------------------------------------

def search_item(query, items):

    found = [i for i in items if query.lower() in i["name"].lower()]

    if not found:
        print(" No item found")
        return

    for i in found:
        print("\n" + "="*80)
        print(f" ITEM NAME            : {i['name']}")
        print(f" STOCK GROUP          : {i['group']}")
        print(f" UNIT                 : {i['unit']}\n")

        print(f" SOURCE OF HSN DETAILS: {i['hsn_source']}")
        print(f" HSN / SAC            : {i['hsn']}")
        print(f" DESCRIPTION          : {i['description']}\n")
        
        print(f" GST APPLICABILITY    : {i['gst_applicable']}")
        print(f" GST RATE SOURCE      : {i['gst_rate_source']}")
        print(f" GST RATE             : {i['gst_rate']}")
        print(f" TAXABILITY TYPE      : {i['taxability']}")
        
        print(f" TYPE OF SUPPLY       : {i['supply_type']}")
        
        print(f" RATE OF DUTY         : {i['rate_of_duty']}\n")

        print(f" OPENING QTY          : {i['qty']} {i['qty_unit']}")
        print(f" RATE / UNIT          : {i['rate']} / {i['rate_unit']}")
        print(f" TOTAL VALUE          : {i['value']}")
        print("="*80)

# ----------------------------------------------------------
# API WRAPPER
# ----------------------------------------------------------

def get_all_items_data():
    """Wrapper function for API to get all items with stats"""
    groups = fetch_stock_groups()
    items = fetch_stock_items(groups)
    
    # Calculate stats
    total_items = len(items)
    categories = list(set(i.get('group', '') for i in items if i.get('group')))
    total_categories = len(categories)
    
    return {
        "items": items,
        "stats": {
            "total_items": total_items,
            "active_items": total_items, # Logic can be improved
            "categories": total_categories,
            # "total_value": sum(float(i.get('value', 0) or 0) for i in items)  #  Disabled as requested
            "total_value": 0  # Disabled - not calculating stock value
        }
    }

# ----------------------------------------------------------
# ZOHO SYNC
# ----------------------------------------------------------

def sync_items_to_zoho(selected_items=None, *, log=None, stop_event=None):
    try:
        from modules.zoho_connector import zoho
    except ImportError:
        try:
            from zoho_connector import zoho
        except:
            return {"status": "error", "message": "Zoho Connector missing"}

    def _emit(message: str):
        try:
            if callable(log):
                log(message)
            else:
                print(message)
        except Exception:
            pass

    def _should_stop() -> bool:
        try:
            return bool(stop_event and getattr(stop_event, "is_set", None) and stop_event.is_set())
        except Exception:
            return False

    _emit("Starting Zoho Sync (Items)...")

    # ── Fetch items to sync ───────────────────────────────────────────────────
    if not selected_items:
        if database_manager:
            try:
                database_manager.init_db()
            except Exception:
                pass
            _emit("Fetching items from SQLite Database...")
            items_to_sync = database_manager.get_all_items()
            if not items_to_sync:
                _emit("No items in DB. Trying Tally fetch...")
                data = get_all_items_data()
                items_to_sync = data["items"] if data else []
        else:
            _emit("Fetching items directly from Tally...")
            data = get_all_items_data()
            if not data:
                return {"status": "error", "message": "No Tally Data"}
            items_to_sync = data["items"]
    else:
        items_to_sync = selected_items

    if not items_to_sync:
        return {"status": "error", "message": "No items to sync"}

    total = len(items_to_sync)
    _emit(f"Total items to sync: {total}")

    # ── Defaults (as requested) ──────────────────────────────────────────────
    def _env_bool(key: str, default: str = "1") -> bool:
        raw = (os.environ.get(key, default) or "").strip().lower()
        return raw not in ("0", "false", "no", "off", "")

    # Resume / delta sync behaviour
    # - ITEM_RESUME_ONLY_NEW=1 (default): skip items already synced once
    # - ITEM_UPDATE_EXISTING=1: update existing items (slower)
    resume_only_new = _env_bool("ITEM_RESUME_ONLY_NEW", "1")
    update_existing = _env_bool("ITEM_UPDATE_EXISTING", "0")

    default_gst_pct = float(os.environ.get("ITEM_DEFAULT_GST_PERCENT", "12") or 12)
    intra_tax_name = (os.environ.get("ITEM_DEFAULT_INTRA_TAX_NAME", "GST12") or "GST12").strip()
    inter_tax_name = (os.environ.get("ITEM_DEFAULT_INTER_TAX_NAME", "IGST12") or "IGST12").strip()

    purchase_account_name = (os.environ.get("ITEM_PURCHASE_ACCOUNT_NAME", "Cost of Goods Sold") or "").strip()
    inventory_account_name = (os.environ.get("ITEM_INVENTORY_ACCOUNT_NAME", "Inventory Asset") or "").strip()
    # Zoho Books v3 Items API does NOT document inventory_valuation_method.
    # Keep it behind a flag to avoid extra failures/retries.
    inventory_valuation_method = (os.environ.get("ITEM_INVENTORY_VALUATION_METHOD", "fifo") or "fifo").strip().lower()
    send_inventory_valuation_method = _env_bool("ITEM_SEND_INVENTORY_VALUATION_METHOD", "0")

    enable_purchase_info = _env_bool("ITEM_ENABLE_PURCHASE_INFO", "1")
    enable_inventory = _env_bool("ITEM_ENABLE_INVENTORY", "1")
    # Use item_tax_preferences for India (intra/inter) instead of undocumented fields.
    include_item_tax_preferences = _env_bool("ITEM_INCLUDE_ITEM_TAX_PREFERENCES", "1")

    def _norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", (s or "").lower())

    # ── Fetch taxes (Zoho org may not allow /settings/taxgroups) ─────────────
    gst_tax_by_pct = {}
    igst_tax_by_pct = {}
    gst_tax_by_name = {}
    igst_tax_by_name = {}

    try:
        t_resp = zoho.api_call("GET", "/settings/taxes")
        if t_resp.get("code") == 0:
            for t in t_resp.get("taxes", []):
                t_name = (t.get("tax_name", "") or "").strip()
                t_pct = float(t.get("tax_percentage", 0) or 0)
                t_id = t.get("tax_id") or ""
                if not t_id:
                    continue
                key_name = _norm(t_name)
                if "igst" in t_name.lower():
                    if t_pct and t_pct not in igst_tax_by_pct:
                        igst_tax_by_pct[t_pct] = t_id
                    if key_name and key_name not in igst_tax_by_name:
                        igst_tax_by_name[key_name] = t_id
                elif "gst" in t_name.lower():
                    if t_pct and t_pct not in gst_tax_by_pct:
                        gst_tax_by_pct[t_pct] = t_id
                    if key_name and key_name not in gst_tax_by_name:
                        gst_tax_by_name[key_name] = t_id
        else:
            _emit(f"Could not fetch taxes: {t_resp.get('message')}")
    except Exception as e:
        _emit(f"Error fetching taxes: {e}")

    def _pct_key(pct: float):
        try:
            p = float(pct or 0)
        except Exception:
            return None
        if not p:
            return None
        # Avoid 17.999999 type issues
        if abs(p - round(p)) < 1e-6:
            return float(int(round(p)))
        return float(p)

    def _env_tax_override(pct_key):
        """
        Optional overrides when /settings/taxgroups is not accessible.

        Supported formats:
        - ITEM_TAX_ID_MAP='{\"18\": {\"intra\": \"...\", \"inter\": \"...\"}, \"12\": {...}}'
        - ITEM_INTRA_TAX_ID_18 / ITEM_INTER_TAX_ID_18 (etc)
        - ITEM_INTRA_TAX_ID_DEFAULT / ITEM_INTER_TAX_ID_DEFAULT
        """
        # 1) JSON map
        raw_map = (os.environ.get("ITEM_TAX_ID_MAP") or "").strip()
        if raw_map:
            try:
                m = json.loads(raw_map)
                key = str(int(pct_key)) if pct_key and abs(pct_key - round(pct_key)) < 1e-6 else str(pct_key)
                obj = m.get(key) or m.get(str(pct_key)) or {}
                intra = (obj.get("intra") or "").strip()
                inter = (obj.get("inter") or "").strip()
                if intra or inter:
                    return intra, inter, "ITEM_TAX_ID_MAP"
            except Exception:
                pass

        # 2) Per-percent env vars
        if pct_key and abs(pct_key - round(pct_key)) < 1e-6:
            k = str(int(round(pct_key)))
            intra = (os.environ.get(f"ITEM_INTRA_TAX_ID_{k}") or "").strip()
            inter = (os.environ.get(f"ITEM_INTER_TAX_ID_{k}") or "").strip()
            if intra or inter:
                return intra, inter, f"ITEM_INTRA/INTER_TAX_ID_{k}"

        # 3) Default env vars
        intra = (os.environ.get("ITEM_INTRA_TAX_ID_DEFAULT") or "").strip()
        inter = (os.environ.get("ITEM_INTER_TAX_ID_DEFAULT") or "").strip()
        if intra or inter:
            return intra, inter, "ITEM_*_TAX_ID_DEFAULT"

        return "", "", ""

    def resolve_tax_ids(gst_pct: float):
        """
        For India:
        - interstate usually has an IGST tax with % equal to gst_pct
        - intrastate is often a GST tax group like GST18 [18%]

        Some orgs don't allow /settings/taxgroups, so we try:
        1) name match for GST{pct} / IGST{pct}
        2) percentage match (works for IGST, may fail for GST groups)
        3) fallback to default configured tax names/ids
        """
        intra_id = ""
        inter_id = ""

        pct_key = _pct_key(gst_pct)
        pct_int = int(round(pct_key)) if pct_key and abs(pct_key - round(pct_key)) < 1e-6 else None

        # 0) Environment overrides (best option when taxgroups endpoint is blocked)
        if pct_key:
            env_intra, env_inter, src = _env_tax_override(pct_key)
            if env_intra or env_inter:
                return env_intra, env_inter

        # 1) Prefer tax names that match the item's % (GST18/IGST18 etc.)
        if pct_int:
            intra_id = gst_tax_by_name.get(_norm(f"GST{pct_int}"), "")
            inter_id = igst_tax_by_name.get(_norm(f"IGST{pct_int}"), "")

        # 2) Percentage match (works for IGST; GST may not exist without taxgroups)
        if not intra_id and pct_key:
            intra_id = gst_tax_by_pct.get(pct_key, "")
        if not inter_id and pct_key:
            inter_id = igst_tax_by_pct.get(pct_key, "")

        # 3) If still missing, use default names as a fallback
        if not intra_id and intra_tax_name:
            intra_id = gst_tax_by_name.get(_norm(intra_tax_name), "")
        if not inter_id and inter_tax_name:
            inter_id = igst_tax_by_name.get(_norm(inter_tax_name), "")

        return intra_id, inter_id

    default_intra_tax_id, default_inter_tax_id = resolve_tax_ids(default_gst_pct)
    _emit(f"Default Taxes: intra='{intra_tax_name}' -> {default_intra_tax_id or 'MISSING'} | inter='{inter_tax_name}' -> {default_inter_tax_id or 'MISSING'}")

    # ── Fetch chart of accounts (for Purchase/Inventory accounts) ────────────
    account_name_to_id = {}
    try:
        page = 1
        while True:
            r = zoho.api_call("GET", "/chartofaccounts", params={
                "page": page,
                "per_page": 200,
                "filter_by": "AccountType.All",
            })
            if r.get("code") != 0:
                _emit(f"Could not fetch chart of accounts page {page}: {r.get('message')}")
                break
            batch = r.get("chartofaccounts", []) or []
            for acc in batch:
                nm = (acc.get("account_name") or "").strip()
                acc_id = acc.get("account_id")
                if nm and acc_id and nm.lower() not in account_name_to_id:
                    account_name_to_id[nm.lower()] = acc_id
            if not r.get("page_context", {}).get("has_more_page", False):
                break
            page += 1
    except Exception as e:
        _emit(f"Error fetching chart of accounts: {e}")

    sales_account_id = account_name_to_id.get("sales", "")
    purchase_account_id = account_name_to_id.get(purchase_account_name.lower(), "") if purchase_account_name else ""
    inventory_account_id = account_name_to_id.get(inventory_account_name.lower(), "") if inventory_account_name else ""

    _emit(f"Accounts: Sales -> {sales_account_id or 'MISSING'} | Purchase('{purchase_account_name}') -> {purchase_account_id or 'MISSING'} | Inventory('{inventory_account_name}') -> {inventory_account_id or 'MISSING'}")

    # ── Existing items map (fast resume via local DB) ────────────────────────
    existing_items = {}
    if database_manager and hasattr(database_manager, "get_zoho_item_sync_map"):
        try:
            sync_map = database_manager.get_zoho_item_sync_map() or {}
            for k, v in sync_map.items():
                item_id = (v.get("zoho_item_id") or "").strip()
                if item_id:
                    existing_items[k] = item_id
            if existing_items:
                _emit(f"Loaded {len(existing_items)} existing items from local sync map (resume).")
        except Exception as e:
            _emit(f"Could not load local item sync map: {e}")

    # Fallback: bulk pre-load existing Zoho items (slow; mainly for first run)
    if not existing_items:
        _emit("Pre-loading existing Zoho items (fallback)...")
        page = 1
        while True:
            res = zoho.api_call("GET", "/items", params={"page": page, "per_page": 200, "filter_by": "Status.All"})
            if res.get("code") != 0:
                _emit(f"Could not fetch items page {page}: {res.get('message')}")
                break
            page_items = res.get("items", [])
            for item in page_items:
                nm_raw = (item.get("name") or "").strip()
                nm = nm_raw.lower().strip()
                it_id = item.get("item_id") or ""
                if nm and it_id:
                    existing_items[nm] = it_id
                    if database_manager and hasattr(database_manager, "upsert_zoho_item_sync") and nm_raw:
                        try:
                            database_manager.upsert_zoho_item_sync(nm_raw, it_id)
                        except Exception:
                            pass
            has_more = res.get("page_context", {}).get("has_more_page", False)
            _emit(f"   Page {page} — {len(page_items)} items (has_more={has_more})")
            if not has_more:
                break
            page += 1
        _emit(f"Pre-loaded {len(existing_items)} existing Zoho items")

    # If resume-only-new is enabled, reduce the work by filtering to new items only
    skipped_by_resume = 0
    if resume_only_new and not update_existing and existing_items:
        before_count = len(items_to_sync)
        filtered = []
        for it in items_to_sync:
            nm = (it.get("name") or "").strip().lower()
            if nm and nm not in existing_items:
                filtered.append(it)
        items_to_sync = filtered
        after_count = len(items_to_sync)
        skipped_by_resume = before_count - after_count
        if skipped_by_resume:
            _emit(f"Resume mode: skipping {skipped_by_resume} already-synced items. Syncing only {after_count} new items.")
        total = len(items_to_sync)

    # ── Main sync loop ────────────────────────────────────────────────────────
    stats        = {"created": 0, "updated": 0, "skipped": skipped_by_resume, "failed": 0}
    failed_items = []

    def clean(val):
        return (val or "").replace("\r", "").replace("\n", "").strip()

    # Custom field: Category dropdown (Zoho Items custom field).
    # Set this in .env as ITEM_CATEGORY_CUSTOMFIELD_ID (numeric).
    category_cf_id = (os.environ.get("ITEM_CATEGORY_CUSTOMFIELD_ID") or "").strip().strip('"').strip("'")
    if category_cf_id and not re.fullmatch(r"\d+", category_cf_id or ""):
        _emit(f"Invalid ITEM_CATEGORY_CUSTOMFIELD_ID='{category_cf_id}' (must be numeric). Skipping category custom field.")
        category_cf_id = ""
    if not category_cf_id:
        _emit("ITEM_CATEGORY_CUSTOMFIELD_ID not set — skipping Item Category custom field mapping.")

    def _call_item_api(method: str, path: str, payload: dict):
        """
        Zoho sometimes rejects unknown/unsupported fields by DC/edition.
        Retry once without optional fields if needed.
        """
        res = zoho.api_call(method, path, payload=payload)
        if res.get("code") == 0:
            return res

        msg = (res.get("message") or "").lower()

        # Retry without optional fields (inventory valuation + tax prefs)
        optional_keys = {
            "inventory_valuation_method",
            "item_tax_preferences",
        }
        if any(k in payload for k in optional_keys) and ("invalid" in msg or "unexpected" in msg or "not allowed" in msg):
            payload2 = {k: v for k, v in payload.items() if k not in optional_keys}
            res2 = zoho.api_call(method, path, payload=payload2)
            if res2.get("code") == 0:
                return res2
        return res

    for idx, i in enumerate(items_to_sync, 1):
        if _should_stop():
            _emit("Stopped by user. Exiting sync loop.")
            return {"status": "stopped", "stats": stats, "failed_items": failed_items}

        name = clean(i.get("name", ""))
        if not name:
            continue

        name_key   = name.lower().strip()
        rate       = float(i.get("rate", 0) or 0)
        hsn        = clean(i.get("hsn", ""))
        unit       = clean(i.get("unit", ""))
        desc       = clean(i.get("description", ""))

        # Resume mode: only insert NEW items (skip already-synced names)
        if resume_only_new and (name_key in existing_items) and not update_existing:
            stats["skipped"] += 1
            continue

        # Dynamic GST from DB (fallback to default if missing/0)
        try:
            gst_pct = float(i.get("gst_rate", 0) or 0)
        except Exception:
            gst_pct = 0.0
        if not gst_pct:
            gst_pct = default_gst_pct

        tax_id_intra, tax_id_inter = resolve_tax_ids(gst_pct)
        if not tax_id_intra and default_intra_tax_id:
            tax_id_intra = default_intra_tax_id
        if not tax_id_inter and default_inter_tax_id:
            tax_id_inter = default_inter_tax_id

        if gst_pct != default_gst_pct and (tax_id_intra or tax_id_inter):
            _emit(f"   Tax for '{name}': {gst_pct}% (intra_id={tax_id_intra or '-'}, inter_id={tax_id_inter or '-'})")

        # Map Tally supply_type → Zoho product_type
        supply     = clean(i.get("supply_type", "")).lower()
        product_type = "service" if "service" in supply else "goods"

        # Progress log every 50
        if idx % 50 == 0 or idx == 1 or idx == total:
            _emit(f"   Progress: {idx}/{total} | created={stats['created']} updated={stats['updated']} skipped={stats['skipped']} failed={stats['failed']}")

        payload = {
            "name":        name,
            "rate":        rate,
            "unit":        unit,          # unit of measure from Tally
            "description": desc,          # HSN description
            "hsn_or_sac":  hsn,           # HSN/SAC code
            "product_type": product_type  # "goods" or "service"
        }

        # Sales account (helps fill Sales Information → Account)
        if sales_account_id:
            payload["account_id"] = sales_account_id

        # Tax rates per item (India: use item_tax_preferences for intra/inter)
        if include_item_tax_preferences and tax_id_intra and tax_id_inter:
            payload["item_tax_preferences"] = [
                {"tax_id": tax_id_intra, "tax_specification": "intra"},
                {"tax_id": tax_id_inter, "tax_specification": "inter"},
            ]
        elif tax_id_intra or tax_id_inter:
            # Fallback for editions that don't support item_tax_preferences
            payload["tax_id"] = tax_id_intra or tax_id_inter

        # Purchase Information defaults
        if enable_purchase_info:
            payload["purchase_rate"] = 0  # cost price default
            if purchase_account_id:
                payload["purchase_account_id"] = purchase_account_id  # Cost of Goods Sold
            payload["purchase_description"] = desc
            if tax_id_intra:
                payload["purchase_tax_id"] = tax_id_intra

        # Track Inventory defaults (only for Goods): Zoho Books uses item_type='inventory'
        if enable_inventory and product_type == "goods":
            payload["item_type"] = "inventory"
            if inventory_account_id:
                payload["inventory_account_id"] = inventory_account_id  # Inventory Asset
            # Mandatory for inventory items in Books API
            if purchase_account_id:
                payload["purchase_account_id"] = purchase_account_id  # Cost of Goods Sold
            # Optional opening stock fields (kept at 0)
            payload["initial_stock"] = 0
            payload["initial_stock_rate"] = 0
            # Optional (not in public docs): FIFO/LIFO/WAC
            if send_inventory_valuation_method and inventory_valuation_method:
                payload["inventory_valuation_method"] = inventory_valuation_method

        # Category custom field (dropdown): value should match a dropdown option label in Zoho.
        # Sent with create/update payload (avoids /item/{item_id}/customfields JSON errors).
        # category_cf_id is loaded and validated once above
        item_group = clean(i.get("group_name", "")) or clean(i.get("groups", "")) or clean(i.get("group", ""))
        should_set_category = bool(category_cf_id and item_group)
        if should_set_category:
            payload["custom_fields"] = [{"customfield_id": str(category_cf_id), "value": item_group}]

        # ── Check existence locally (no extra GET per item) ───────────────────
        if name_key in existing_items:
            # UPDATE existing item
            item_id = existing_items[name_key]
            res = _call_item_api("PUT", f"/items/{item_id}", payload)
            if res.get("code") == 0:
                stats["updated"] += 1
                _emit(f"Updated: {name}")
                # Category custom field sent via payload["custom_fields"]
                if database_manager and hasattr(database_manager, "upsert_zoho_item_sync"):
                    try:
                        database_manager.upsert_zoho_item_sync(name, item_id)
                    except Exception:
                        pass
            else:
                stats["failed"] += 1
                err = res.get("message", "Unknown error")
                failed_items.append({"name": name, "reason": err})
                _emit(f"Update Failed {name}: {err}")
        else:
            # CREATE new item
            res = _call_item_api("POST", "/items", payload)
            if res.get("code") == 0:
                stats["created"] += 1
                created_id = res.get("item", {}).get("item_id", "")
                existing_items[name_key] = created_id
                _emit(f"Created: {name}")
                # Category custom field sent via payload["custom_fields"]
                if database_manager and hasattr(database_manager, "upsert_zoho_item_sync") and created_id:
                    try:
                        database_manager.upsert_zoho_item_sync(name, created_id)
                    except Exception:
                        pass
            else:
                stats["failed"] += 1
                err = res.get("message", "Unknown error")
                failed_items.append({"name": name, "reason": err})
                _emit(f"Create Failed {name}: {err}")

    _emit(f"Items Sync Complete — Created: {stats['created']}, Updated: {stats['updated']}, Skipped: {stats['skipped']}, Failed: {stats['failed']}")
    return {"status": "success", "stats": stats, "failed_items": failed_items}


if __name__ == "__main__":
    pass


# ----------------------------------------------------------
# INVENTORY ADJUSTMENT (OPENING STOCK BY LOCATION)
# ----------------------------------------------------------

WAREHOUSES = ["ANAKAPALLI", "MAIN WAREHOUSE", "HUKUMPETA"]
ONLY_MAIN_WAREHOUSE = True  # inventory adjustment uses MAIN WAREHOUSE only (as requested)

def _norm_item(s: str) -> str:
    """
    Normalized key for matching Item Names across Excel vs Tally XML.
    Uses a "loose" normalization to improve match rate:
    - unescape XML entities
    - lowercase
    - remove non-alphanumerics (spaces, punctuation)
    """
    s = html.unescape(s or "")
    s = s.replace("\u00a0", " ")  # NBSP
    s = s.replace("’", "'").strip().lower()
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s

def _parse_float(val) -> float:
    try:
        if val is None:
            return 0.0
        s = str(val).replace(",", "").strip()
        if not s:
            return 0.0
        return float(s)
    except Exception:
        return 0.0

def _parse_qty(val) -> float:
    """
    Parse qty like: "14 NO'S" / "238 NO'S" / "0" -> 14.0
    """
    try:
        s = (val or "").replace(",", "").strip()
        if not s:
            return 0.0
        m = re.search(r"(-?\d+(?:\.\d+)?)", s)
        return float(m.group(1)) if m else 0.0
    except Exception:
        return 0.0

def _norm_warehouse(name: str) -> str:
    raw = (name or "").strip().lower()
    if not raw:
        return ""
    k = re.sub(r"[^a-z0-9]+", "", raw)
    if "anakap" in k:
        return "ANAKAPALLI"
    if "hukump" in k:
        return "HUKUMPETA"
    if "main" in k and "ware" in k:
        return "MAIN WAREHOUSE"
    # Exact matches (fallback)
    for wh in WAREHOUSES:
        if k == re.sub(r"[^a-z0-9]+", "", wh.lower()):
            return wh
    return ""

def parse_opening_excel_xlsx(file_path: str) -> dict:
    """
    Returns excel_map[norm_item] = {
        "name": original,
        "qty": float,
        "rate": float,
        "amount": float
    }
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(c or "").strip().lower() for c in rows[0]]
    def _find_col(keys):
        for i, h in enumerate(header):
            if any(k in h for k in keys):
                return i
        return None

    col_item = _find_col(["item", "name"])
    col_qty = _find_col(["qty", "quantity"])
    col_rate = _find_col(["rate", "price"])
    col_amt = _find_col(["amount", "value"])
    if col_item is None or col_qty is None:
        raise ValueError("Excel must have at least Item name and Qty columns in first row header.")

    excel_map = {}
    for r in rows[1:]:
        item = str(r[col_item] or "").strip()
        if not item:
            continue
        qty = _parse_qty(str(r[col_qty] or ""))
        rate = _parse_float(r[col_rate]) if col_rate is not None else 0.0
        amt = _parse_float(r[col_amt]) if col_amt is not None else 0.0
        key = _norm_item(item)
        if key not in excel_map:
            excel_map[key] = {"name": item, "qty": 0.0, "rate": 0.0, "amount": 0.0}
        excel_map[key]["qty"] += qty
        excel_map[key]["amount"] += amt
        # Prefer non-zero rate (keep latest non-zero)
        if rate:
            excel_map[key]["rate"] = rate

    # Fill missing rate from amount/qty when possible
    for k, v in excel_map.items():
        if (not v.get("rate")) and v.get("qty"):
            v["rate"] = (v.get("amount", 0.0) / v["qty"]) if v["qty"] else 0.0
    return excel_map

def parse_godown_xml(file_path: str) -> dict:
    """
    Best-effort parser for Tally Godown Summary export.

    Returns xml_map[norm_item] = {
        "name": original_item,
        "warehouses": {"MAIN WAREHOUSE": qty, ...},
        "rate": float,
        "amount": float
    }

    NOTE: This parser supports warehouse headers when the XML contains section headings
    with empty qty. If no warehouse header is found, items default to MAIN WAREHOUSE.
    """
    from lxml import etree

    with open(file_path, "rb") as f:
        xml_bytes = f.read()

    parser = etree.XMLParser(recover=True, huge_tree=True)
    root = etree.fromstring(xml_bytes, parser=parser)

    # The file pattern in your sample is: DSPACCNAME (name) then DSPSTKINFO (qty/rate/amt)
    children = list(root)
    pending_name = None
    current_warehouse = ""
    saw_warehouse_header = False

    xml_map = {}

    def _ensure_item(item_name: str):
        k = _norm_item(item_name)
        if k not in xml_map:
            xml_map[k] = {"name": item_name, "warehouses": {}, "rate": 0.0, "amount": 0.0}
        return k

    for ch in children:
        tag = (ch.tag or "").upper()
        if tag == "DSPACCNAME":
            pending_name = (ch.findtext("DSPDISPNAME") or "").strip()
        elif tag == "DSPSTKINFO" and pending_name is not None:
            qty_txt = (ch.findtext(".//DSPCLQTY") or "").strip()
            rate_txt = (ch.findtext(".//DSPCLRATE") or "").strip()
            amt_txt = (ch.findtext(".//DSPCLAMTA") or "").strip()

            qty = _parse_qty(qty_txt)
            rate = _parse_float(rate_txt)
            amt = _parse_float(amt_txt)

            # Treat warehouse header rows (often qty=0 with empty rate/amount)
            wh_candidate = _norm_warehouse(pending_name)
            if wh_candidate and (not rate_txt) and (not amt_txt) and (qty == 0.0):
                current_warehouse = wh_candidate
                saw_warehouse_header = True
                pending_name = None
                continue

            # If qty is empty -> treat as header/section
            if not qty_txt:
                wh = _norm_warehouse(pending_name)
                if wh:
                    current_warehouse = wh
                    saw_warehouse_header = True
                pending_name = None
                continue

            wh = current_warehouse or "MAIN WAREHOUSE"
            k = _ensure_item(pending_name)
            xml_map[k]["warehouses"][wh] = xml_map[k]["warehouses"].get(wh, 0.0) + qty
            xml_map[k]["amount"] += amt
            # keep a non-zero rate (best effort)
            if rate:
                xml_map[k]["rate"] = rate
            pending_name = None

    # Fill missing rate from amount/qty when possible
    for k, v in xml_map.items():
        total_qty = sum(v.get("warehouses", {}).values()) or 0.0
        if (not v.get("rate")) and total_qty:
            v["rate"] = (v.get("amount", 0.0) / total_qty) if total_qty else 0.0

    # If no warehouse headers were detected, assume this is a single-warehouse export
    # (often MAIN WAREHOUSE only). In that case, all parsed rows already defaulted to MAIN WAREHOUSE.
    if not saw_warehouse_header:
        print("️ No warehouse headers detected in XML — assuming single warehouse export (MAIN WAREHOUSE).")

    return xml_map

def compute_inventory_adjustment(excel_map: dict, xml_map: dict):
    """
    Returns:
      to_apply: list of {name, rate, locations: {warehouse: qty}}
      negative: list of {name, excel_qty, xml_other_qty, needed_main_qty}
      stats: summary counts
    """
    to_apply = []
    negative = []

    for k, x in (xml_map or {}).items():
        item_name = x.get("name") or k
        wh_qty = x.get("warehouses") or {}
        xml_main = wh_qty.get("MAIN WAREHOUSE", 0.0)
        xml_ana = wh_qty.get("ANAKAPALLI", 0.0) if not ONLY_MAIN_WAREHOUSE else 0.0
        xml_huku = wh_qty.get("HUKUMPETA", 0.0) if not ONLY_MAIN_WAREHOUSE else 0.0

        if k in (excel_map or {}):
            e = excel_map[k]
            excel_qty = float(e.get("qty") or 0.0)
            excel_rate = float(e.get("rate") or 0.0) or (float(e.get("amount") or 0.0) / excel_qty if excel_qty else 0.0)

            # MAIN-only adjustment:
            # - if in Excel, Opening Stock comes from Excel
            # - XML main qty is used only for items not present in Excel
            other_sum = (xml_ana or 0.0) + (xml_huku or 0.0)
            needed_main = excel_qty if ONLY_MAIN_WAREHOUSE else (excel_qty - other_sum)

            if needed_main < 0:
                negative.append({
                    "name": item_name,
                    "reason": "excel_negative",
                    "excel_qty": excel_qty,
                    "xml_main_qty": xml_main,
                    "xml_other_qty": other_sum,
                    "needed_main_qty": needed_main
                })
                continue

            locs = {}
            locs["MAIN WAREHOUSE"] = needed_main
            if not ONLY_MAIN_WAREHOUSE:
                if xml_ana:
                    locs["ANAKAPALLI"] = xml_ana
                if xml_huku:
                    locs["HUKUMPETA"] = xml_huku

            to_apply.append({"name": item_name, "rate": excel_rate, "locations": locs})

        else:
            # Not in excel:
            # - if any qty is negative, report and skip (migration team handles)
            if (xml_main or 0.0) < 0 or (xml_ana or 0.0) < 0 or (xml_huku or 0.0) < 0:
                negative.append({
                    "name": item_name,
                    "reason": "xml_negative",
                    "excel_qty": "",
                    "xml_main_qty": xml_main,
                    "xml_other_qty": (xml_ana or 0.0) + (xml_huku or 0.0),
                    "needed_main_qty": ""
                })
                continue

            # Use XML distribution as-is
            xml_rate = float(x.get("rate") or 0.0)
            locs = {}
            if xml_main:
                locs["MAIN WAREHOUSE"] = xml_main
            if not locs:
                # still create a default MAIN row if nothing parsed
                locs["MAIN WAREHOUSE"] = 0.0
            to_apply.append({"name": item_name, "rate": xml_rate, "locations": locs})

    stats = {
        "xml_items": len(xml_map or {}),
        "excel_items": len(excel_map or {}),
        "to_apply": len(to_apply),
        "negative": len(negative),
    }
    return to_apply, negative, stats

def build_matched_items_report(excel_map: dict, xml_map: dict):
    """
    Report for items that exist in BOTH Excel and XML.
    Includes both:
      - needed_main_qty (location balancing logic)
      - sum_qty (excel_qty + xml_main_qty) for MAIN-only exports / validation
    """
    rows = []
    excel_map = excel_map or {}
    xml_map = xml_map or {}

    for k, e in excel_map.items():
        if k not in xml_map:
            continue
        x = xml_map.get(k) or {}
        wh = x.get("warehouses") or {}
        xml_main = float(wh.get("MAIN WAREHOUSE", 0.0) or 0.0)
        xml_ana = float(wh.get("ANAKAPALLI", 0.0) or 0.0)
        xml_huku = float(wh.get("HUKUMPETA", 0.0) or 0.0)
        xml_other = (xml_ana or 0.0) + (xml_huku or 0.0)

        excel_qty = float(e.get("qty") or 0.0)
        excel_rate = float(e.get("rate") or 0.0)
        excel_amount = float(e.get("amount") or 0.0)
        if (not excel_rate) and excel_qty:
            excel_rate = (excel_amount / excel_qty) if excel_qty else 0.0

        needed_main = excel_qty - xml_other
        sum_qty = excel_qty + xml_main

        rows.append({
            "name": e.get("name") or x.get("name") or k,
            "excel_qty": excel_qty,
            "excel_rate": excel_rate,
            "xml_main_qty": xml_main,
            "xml_other_qty": xml_other,
            "needed_main_qty": needed_main,
            "sum_qty": sum_qty,
            "status": "negative" if needed_main < 0 else "ok"
        })

    return rows

def apply_inventory_adjustment_to_zoho(to_apply: list, dry_run: bool = True, run_id: str = "", resume: bool = True):
    """
    Applies opening stock by location using Zoho Books Items Update API.
    Uses local DB sync map (zoho_item_sync) to resolve Zoho item_id.
    """
    try:
        from modules.zoho_connector import zoho
    except Exception:
        from zoho_connector import zoho

    if database_manager:
        try:
            database_manager.init_db()
        except Exception:
            pass

    already_applied_ids = set()
    if resume and run_id and database_manager and hasattr(database_manager, "get_inv_adj_applied_ids"):
        try:
            already_applied_ids = database_manager.get_inv_adj_applied_ids(run_id) or set()
            if already_applied_ids:
                print(f"⏭️ Resume enabled — skipping {len(already_applied_ids)} items already applied for run_id={run_id}.")
        except Exception:
            already_applied_ids = set()

    # Build normalized name -> zoho_item_id map from local sync table
    norm_to_item_id = {}
    if database_manager and hasattr(database_manager, "get_zoho_item_sync_map"):
        try:
            sync_map = database_manager.get_zoho_item_sync_map() or {}
            for _k, v in sync_map.items():
                if isinstance(v, dict):
                    nm = (v.get("item_name") or "").strip()
                    zid = (v.get("zoho_item_id") or "").strip()
                else:
                    nm = str(_k or "").strip()
                    zid = str(v or "").strip()
                if nm and zid:
                    norm_to_item_id[_norm_item(nm)] = zid
        except Exception:
            norm_to_item_id = {}

    # Resolve location ids
    loc_res = zoho.api_call("GET", "/locations")
    if loc_res.get("code") != 0:
        return {"status": "error", "message": f"Failed to fetch locations: {loc_res.get('message')}"}
    locs = loc_res.get("locations", []) or loc_res.get("location", []) or []
    loc_name_to_id = {}
    for l in locs:
        nm = (l.get("location_name") or l.get("name") or "").strip()
        lid = l.get("location_id")
        if nm and lid:
            loc_name_to_id[nm.lower()] = str(lid)

    missing_locations = [wh for wh in WAREHOUSES if wh.lower() not in loc_name_to_id]
    if missing_locations:
        return {"status": "error", "message": f"Missing Zoho locations: {', '.join(missing_locations)}"}

    # Pre-fetch item details in chunks (name + rate required for update)
    item_ids = []
    name_to_id = {}
    for row in to_apply or []:
        nm = (row.get("name") or "").strip()
        if not nm:
            continue
        zid = norm_to_item_id.get(_norm_item(nm), "")
        if not zid:
            continue
        item_ids.append(str(zid))
        name_to_id[_norm_item(nm)] = str(zid)

    # /itemdetails supports comma-separated item ids
    id_to_details = {}
    chunk_size = 80
    for i in range(0, len(item_ids), chunk_size):
        chunk = item_ids[i:i+chunk_size]
        r = zoho.api_call("GET", "/itemdetails", params={"item_ids": ",".join(chunk)})
        if r.get("code") != 0:
            continue
        for it in r.get("items", []) or []:
            iid = str(it.get("item_id") or "")
            if iid:
                id_to_details[iid] = it

    results = {
        "updated": 0,
        "skipped": 0,
        "failed": 0,
        "missing_item_id": 0,
        "updated_items": [],  # [{item_id,name}]
        "skipped_items": [],  # [{item_id,name,reason}]
        "errors": [],
    }

    for row in to_apply or []:
        nm = (row.get("name") or "").strip()
        if not nm:
            continue
        key = _norm_item(nm)
        zid = name_to_id.get(key, "")
        if not zid:
            results["missing_item_id"] += 1
            results["skipped_items"].append({"item_id": "", "name": nm, "reason": "missing_item_id"})
            results["errors"].append({"name": nm, "reason": "Missing Zoho item_id (sync map not found)."})
            continue

        if resume and run_id and zid in already_applied_ids:
            results["skipped"] += 1
            results["skipped_items"].append({"item_id": zid, "name": nm, "reason": "resume_already_applied"})
            continue

        det = id_to_details.get(zid) or {}
        api_name = (det.get("name") or nm).strip()
        api_rate = det.get("rate")
        api_rate = _parse_float(api_rate)
        if not api_rate:
            # Keep existing if possible, else use provided
            api_rate = _parse_float(row.get("rate"))

        # MAIN WAREHOUSE only (as requested)
        locations_payload = []
        main_qty = _parse_float((row.get("locations") or {}).get("MAIN WAREHOUSE", 0))
        locations_payload.append({
            "location_id": loc_name_to_id["main warehouse"],
            "initial_stock": str(main_qty),
            "initial_stock_rate": str(_parse_float(row.get("rate")))
        })

        payload = {
            "name": api_name,
            "rate": api_rate,
            "locations": locations_payload
        }

        if dry_run:
            results["skipped"] += 1
            results["skipped_items"].append({"item_id": zid, "name": api_name, "reason": "dry_run"})
            continue

        upd = zoho.api_call("PUT", f"/items/{zid}", payload=payload)
        if upd.get("code") == 0:
            results["updated"] += 1
            results["updated_items"].append({"item_id": zid, "name": api_name})
            if run_id and database_manager and hasattr(database_manager, "mark_inv_adj_applied"):
                try:
                    database_manager.mark_inv_adj_applied(run_id, zid, api_name)
                    already_applied_ids.add(zid)
                except Exception:
                    pass
        else:
            results["failed"] += 1
            results["errors"].append({"name": nm, "reason": upd.get("message", "Update failed")})

    return {"status": "success", "dry_run": dry_run, "results": results}


def refresh_zoho_item_sync_map_from_zoho(max_pages: int = 1000):
    """
    Fast path to build local name->item_id map without re-syncing/creating items.
    Fetches Zoho /items pages and stores into zoho_item_sync.
    """
    try:
        from modules.zoho_connector import zoho
    except Exception:
        from zoho_connector import zoho

    if database_manager:
        try:
            database_manager.init_db()
        except Exception:
            pass

    if not database_manager or not hasattr(database_manager, "upsert_zoho_item_sync"):
        return {"status": "error", "message": "database_manager.upsert_zoho_item_sync not available"}

    total = 0
    page = 1
    while page <= max_pages:
        res = zoho.api_call("GET", "/items", params={"page": page, "per_page": 200, "filter_by": "Status.All"})
        if res.get("code") != 0:
            return {"status": "error", "message": f"Failed to fetch items page {page}: {res.get('message')}", "loaded": total}

        items = res.get("items", []) or []
        for it in items:
            nm = (it.get("name") or "").strip()
            iid = (it.get("item_id") or "").strip()
            if nm and iid:
                try:
                    database_manager.upsert_zoho_item_sync(nm, iid)
                    total += 1
                except Exception:
                    pass

        ctx = res.get("page_context", {}) or {}
        has_more = bool(ctx.get("has_more_page", False))
        if not has_more:
            break
        page += 1

    return {"status": "success", "loaded": total, "pages": page}


def build_stock_transfer_workbook_from_godown_xml(xml_map: dict):
    """
    Create an Excel workbook matching the format of "Stocks Transferred to Main warehouse.xlsx":
    - Sheet1: 4 columns, no header: Item Name | Qty | Rate | Amount
    - Negative items (qty < 0) are written to a second sheet "NEGATIVE" for review.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Stocks"

    ws_neg = wb.create_sheet("NEGATIVE")
    ws_neg.append(["Item Name", "Qty", "Rate", "Amount"])  # helpful header

    xml_map = xml_map or {}
    # stable ordering
    items = sorted(xml_map.values(), key=lambda v: (str(v.get("name") or "").lower()))

    for it in items:
        name = (it.get("name") or "").strip()
        wh = it.get("warehouses") or {}
        qty = float(wh.get("MAIN WAREHOUSE", 0.0) or 0.0)
        rate = float(it.get("rate") or 0.0)
        amount = float(it.get("amount") or 0.0)
        if not name:
            continue
        if abs(qty) < 1e-9:
            continue

        # Prefer positive values for transfer sheet
        if qty < 0:
            ws_neg.append([name, qty, rate, amount])
            continue

        # Best-effort rate from amount/qty when missing
        if (not rate) and qty:
            rate = (abs(amount) / qty) if qty else 0.0

        ws.append([name, abs(qty), abs(rate), abs(amount)])

    return wb
