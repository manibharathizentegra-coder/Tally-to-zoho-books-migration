from flask import Flask, jsonify, render_template, send_file, request, session, Response, stream_with_context
from flask_cors import CORS
import sys
import os
import json
import os
import re
import time
import io
import threading

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

import field_mapping_manager
try:
    import openpyxl
except ImportError:
    openpyxl = None

# Add modules directory to path
sys.path.append(os.path.dirname(__file__))

try:
    from modules.job_manager import jobs as job_manager, sse_format
except Exception:
    job_manager = None
    sse_format = None

# Import backend modules
try:
    from ledgers import ledgers_backend as ledgers_module
    print(" Successfully imported ledgers_text backend")
except ImportError as e:
    print(f" Error importing ledgers_backend: {e}")
    ledgers_module = None

try:
    from items import items_backend as items_module
    print(" Successfully imported items_backend")
except ImportError as e:
    print(f" Error importing items_backend: {e}")
    items_module = None

try:
    from journel import journel_backend as journel_module
    print(" Successfully imported journel_backend")
except ImportError as e:
    print(f" Error importing journel_backend: {e}")
    journel_module = None

try:
    from invoice import invoice_backend as invoice_module
    print(" Successfully imported invoice_backend")
except ImportError as e:
    print(f" Error importing invoice_backend: {e}")
    invoice_module = None

try:
    from bills import bills_backend as bills_module
    print(" Successfully imported bills_backend")
except ImportError as e:
    print(f" Error importing bills_backend: {e}")
    bills_module = None

try:
    from sales_order import sale_backend as sales_order_module
    print(" Successfully imported sales_order_backend")
except ImportError as e:
    print(f" Error importing sales_order_backend: {e}")
    sales_order_module = None

try:
    from purchase_order import purchase_order_backend as purchase_order_module
    print(" Successfully imported purchase_order_backend")
except ImportError as e:
    print(f" Error importing purchase_order_backend: {e}")
    purchase_order_module = None

try:
    from receipts import receipts_backend as receipts_module
    print(" Successfully imported receipts_backend")
except ImportError as e:
    print(f" Error importing receipts_backend: {e}")
    receipts_module = None

try:
    import importlib.util
    import os
    spec = importlib.util.spec_from_file_location("payments_backend", os.path.join(os.path.dirname(__file__), 'Payments made', 'payments_backend.py'))
    payments_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(payments_module)
    print(" Successfully imported payments_backend")
except Exception as e:
    print(f" Error importing payments_backend: {e}")
    payments_module = None

try:
    spec_contra = importlib.util.spec_from_file_location("contra_backend", os.path.join(os.path.dirname(__file__), 'contra', 'contra_backend.py'))
    contra_module = importlib.util.module_from_spec(spec_contra)
    spec_contra.loader.exec_module(contra_module)
    print(" Successfully imported contra_backend")
except Exception as e:
    print(f" Error importing contra_backend: {e}")
    contra_module = None

try:
    spec_credit_note = importlib.util.spec_from_file_location("credit_note_backend", os.path.join(os.path.dirname(__file__), 'credit_note', 'credit_note_backend.py'))
    credit_note_module = importlib.util.module_from_spec(spec_credit_note)
    spec_credit_note.loader.exec_module(credit_note_module)
    print(" Successfully imported credit_note_backend")
except Exception as e:
    print(f" Error importing credit_note_backend: {e}")
    credit_note_module = None

try:
    spec_debit_note = importlib.util.spec_from_file_location("debit_note_backend", os.path.join(os.path.dirname(__file__), 'debit_note', 'debit_note_backend.py'))
    debit_note_module = importlib.util.module_from_spec(spec_debit_note)
    spec_debit_note.loader.exec_module(debit_note_module)
    print(" Successfully imported debit_note_backend")
except Exception as e:
    print(f" Error importing debit_note_backend: {e}")
    debit_note_module = None

try:
    import database_manager
    print(" Successfully imported database_manager")
except ImportError as e:
    print(f" Error importing database_manager: {e}")
    database_manager = None

app = Flask(__name__)
CORS(app)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-key")

def _sanitize_db_filename(name: str) -> str:
    """
    Return a safe sqlite filename in the project folder.
    Allows letters/numbers/space/._- and forces .db extension.
    """
    if not name:
        return "tally_data.db"
    base = str(name).strip()
    # Normalize common inputs like "AGRITOUGH MACHINERIES - (from 1-Apr-25)"
    base = re.sub(r"\s+", " ", base)
    base = re.sub(r"[^A-Za-z0-9 ._\\-]", "", base).strip()
    base = base.replace(" ", "_")
    if not base:
        base = "company"
    if not base.lower().endswith(".db"):
        base = base + ".db"
    # Block path traversal / directories
    base = os.path.basename(base)
    return base

@app.before_request
def _set_active_company_db():
    if not database_manager:
        return
    active = session.get("active_db") or getattr(database_manager, "DEFAULT_DB_NAME", "tally_data.db")
    try:
        database_manager.set_active_db(active)
    except Exception:
        pass

@app.route('/api/db/companies', methods=['GET'])
def api_db_companies():
    """List available *.db files in the project folder."""
    try:
        files = []
        for fn in os.listdir(os.path.dirname(__file__)):
            if fn.lower().endswith(".db") and os.path.isfile(os.path.join(os.path.dirname(__file__), fn)):
                files.append(fn)
        files = sorted(files, key=lambda x: (x != "tally_data.db", x.lower()))
        return jsonify({"companies": files, "count": len(files), "active_db": session.get("active_db", "tally_data.db")})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/active', methods=['GET', 'POST'])
def api_db_active():
    if request.method == 'GET':
        return jsonify({"active_db": session.get("active_db", "tally_data.db")})
    body = request.get_json(force=True, silent=True) or {}
    db_name = _sanitize_db_filename(body.get("db_name", "tally_data.db"))
    session["active_db"] = db_name
    if database_manager:
        try:
            database_manager.set_active_db(db_name)
            database_manager.init_db(db_name=db_name)
        except Exception:
            pass
    return jsonify({"status": "ok", "active_db": db_name})

@app.route('/api/zoho/diag', methods=['GET'])
def api_zoho_diag():
    """Non-sensitive Zoho auth diagnostics (DC/domains + token refresh result)."""
    try:
        from modules.zoho_connector import zoho, ACCOUNTS_DOMAIN, API_DOMAIN
    except Exception as e:
        return jsonify({"error": f"Zoho connector not available: {e}"}), 500

    token = zoho.get_access_token()
    return jsonify({
        "accounts_domain": ACCOUNTS_DOMAIN,
        "api_domain": API_DOMAIN,
        "auth_url": zoho.auth_url,
        "token_ok": bool(token),
        "has_client_id": bool(zoho.client_id),
        "has_client_secret": bool(zoho.client_secret),
        "has_refresh_token": bool(zoho.refresh_token),
        "client_id_tail": (zoho.client_id[-6:] if zoho.client_id else ""),
        "refresh_token_tail": (zoho.refresh_token[-6:] if zoho.refresh_token else ""),
    })


@app.route('/api/zoho/contacts/<contact_id>/debug', methods=['GET'])
def api_zoho_contact_debug(contact_id):
    """
    Debug helper: fetch a single contact from Zoho Books and return only the fields
    we care about for phone/mobile formatting.
    """
    try:
        from modules.zoho_connector import zoho
    except Exception as e:
        return jsonify({"error": f"Zoho connector not available: {e}"}), 500

    res = zoho.api_call("GET", f"/contacts/{contact_id}")
    if res.get("code") != 0:
        return jsonify({"status": "error", "code": res.get("code"), "message": res.get("message"), "raw": res}), 400

    c = res.get("contact", {}) or {}
    cps = c.get("contact_persons", []) or []
    cps_slim = []
    for cp in cps:
        cps_slim.append({
            "contact_person_id": cp.get("contact_person_id"),
            "is_primary_contact": cp.get("is_primary_contact"),
            "phone": cp.get("phone"),
            "mobile": cp.get("mobile"),
            "email": cp.get("email"),
        })

    out = {
        "contact_id": c.get("contact_id"),
        "contact_name": c.get("contact_name"),
        "contact_type": c.get("contact_type"),
        "gst_treatment": c.get("gst_treatment"),
        "gst_no": c.get("gst_no"),
        "phone": c.get("phone"),
        "mobile": c.get("mobile"),
        "contact_persons": cps_slim,
    }
    return jsonify({"status": "ok", "contact": out})


@app.route('/api/zoho/items/<item_id>/debug', methods=['GET'])
def api_zoho_item_debug(item_id):
    """Debug helper: fetch one item and show tax + custom_fields payload from Zoho."""
    try:
        from modules.zoho_connector import zoho
    except Exception as e:
        return jsonify({"error": f"Zoho connector not available: {e}"}), 500

    res = zoho.api_call("GET", f"/items/{item_id}")
    if res.get("code") != 0:
        return jsonify({"status": "error", "code": res.get("code"), "message": res.get("message"), "raw": res}), 400

    it = res.get("item", {}) or {}
    out = {
        "item_id": it.get("item_id"),
        "name": it.get("name"),
        "tax_id": it.get("tax_id"),
        "tax_name": it.get("tax_name"),
        "item_tax_preferences": it.get("item_tax_preferences"),
        "custom_fields": it.get("custom_fields"),
    }
    return jsonify({"status": "ok", "item": out})



# ---------------------------------------------------------
# ROUTES
# ---------------------------------------------------------

@app.route('/api/db/ledgers', methods=['GET'])
def api_db_ledgers():
    if not database_manager: return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        ledgers = database_manager.get_all_ledgers()
        return jsonify({"ledgers": ledgers, "count": len(ledgers)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/import-masters', methods=['POST'])
def api_db_import_masters():
    """
    Import a Tally Master XML (All Masters) into SQLite.
    Supports multi-company by selecting/creating a DB file.
    """
    if not database_manager:
        return jsonify({"error": "DB Manager not loaded"}), 500

    mf = request.files.get('master_file') or request.files.get('xml_file') or request.files.get('file')
    if not mf or not mf.filename:
        return jsonify({"error": "master_file (.xml) is required"}), 400

    target = (request.form.get('target') or 'existing').strip().lower()  # existing | new
    requested_db = request.form.get('db_name') or request.form.get('company_db') or ''

    xml_bytes = mf.read()
    if not xml_bytes:
        return jsonify({"error": "Empty file"}), 400

    # Extract company name from XML (optional default db name)
    xml_text = ""
    try:
        xml_text = xml_bytes.decode("utf-16", errors="ignore")
    except Exception:
        try:
            xml_text = xml_bytes.decode("utf-8", errors="ignore")
        except Exception:
            xml_text = ""
    m = re.search(r"<SVCURRENTCOMPANY>(.*?)</SVCURRENTCOMPANY>", xml_text, flags=re.IGNORECASE | re.DOTALL)
    company_from_file = (m.group(1).strip() if m else "")

    if target == "new":
        chosen = requested_db or company_from_file or "new_company"
        db_name = _sanitize_db_filename(chosen)
    else:
        db_name = session.get("active_db", "tally_data.db")
        db_name = _sanitize_db_filename(db_name)

    # Set active db for this request + session
    session["active_db"] = db_name
    database_manager.set_active_db(db_name)
    database_manager.init_db(db_name=db_name)

    # Parse XML masters (ledgers/groups/items)
    try:
        from modules import json_to_zoho_converter as conv
    except Exception as e:
        return jsonify({"error": f"Converter not available: {e}"}), 500

    parsed = conv.parse_tally_json(xml_bytes)
    records = parsed.get("records", []) or []
    ctx = parsed.get("context", {}) or {}
    group_parent_map = ctx.get("group_parent_map", {}) if isinstance(ctx, dict) else {}

    def _safe_float(v):
        try:
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0.0

    # Resolve primary group by traversing parents
    def _primary_group(group_name: str) -> str:
        if not group_name:
            return ""
        seen = set()
        cur = group_name
        parent = group_parent_map.get(cur, "")
        while parent and parent not in seen:
            seen.add(parent)
            cur = parent
            parent = group_parent_map.get(cur, "")
        return cur

    imported = {"groups": 0, "ledgers": 0, "items": 0, "errors": []}
    seen_groups = set()
    seen_ledgers = set()
    seen_items = set()

    # 1) Groups (use parsed XML context)
    for g in (ctx.get("groups") or []):
        name = (g.get("name") or "").strip()
        if not name:
            continue
        if name in seen_groups:
            continue
        seen_groups.add(name)
        parent = (g.get("parent") or "").strip()
        data = {"name": name, "parent": parent, "primary_group": _primary_group(name)}
        try:
            database_manager.insert_or_update_group(data)
            imported["groups"] += 1
        except Exception as e:
            imported["errors"].append(f"Group '{name}': {e}")

    # 2) Ledgers (Customers/Vendors/Others)
    for r in records:
        if str(r.get("metadata_type", "")).lower() != "ledger":
            continue
        name = (r.get("ContactName") or r.get("name") or "").strip()
        if not name:
            continue
        if name in seen_ledgers:
            continue
        seen_ledgers.add(name)
        parent = (r.get("Under") or r.get("parent") or "").strip()
        ct = str(r.get("ContactType") or "").strip().lower()
        if ct == "customer":
            ltype = "customer"
        elif ct == "vendor":
            ltype = "vendor"
        else:
            ltype = "other"

        data = {
            "name": name,
            "parent": parent,
            "type": ltype,
            "address": (r.get("BillingAddress") or r.get("address") or "").strip(),
            "state": (r.get("BillingState") or r.get("state") or "").strip(),
            "country": (r.get("BillingCountry") or r.get("country") or "").strip(),
            "pincode": (r.get("BillingZip") or r.get("pincode") or "").strip(),
            "email": (r.get("EmailAddress") or r.get("email") or "").strip(),
            "phone": str(r.get("Phone") or r.get("phone") or "").strip(),
            "gstin": (r.get("GSTIN") or r.get("gstin") or "").strip(),
            "gst_reg_type": (r.get("RegistrationType") or r.get("gst_reg_type") or "").strip(),
            "pan": (r.get("PAN") or r.get("pan") or "").strip(),
            "opening_balance": _safe_float(r.get("openingbalance") or r.get("opening_balance") or r.get("OpeningBalance") or 0),
            "closing_balance": _safe_float(r.get("closingbalance") or r.get("closing_balance") or r.get("ClosingBalance") or 0),
        }
        try:
            database_manager.insert_or_update_ledger(data)
            imported["ledgers"] += 1
        except Exception as e:
            imported["errors"].append(f"Ledger '{name}': {e}")

    # 3) Items (if present in Master.xml)
    for r in records:
        if str(r.get("metadata_type", "")).lower() != "stockitem":
            continue
        name = (r.get("ItemName") or r.get("name") or "").strip()
        if not name:
            continue
        if name in seen_items:
            continue
        seen_items.add(name)
        data = {
            "name": name,
            "group_name": (r.get("parent") or r.get("group_name") or r.get("stockgroup") or "").strip(),
            "category": (r.get("category") or "").strip(),
            "unit": (r.get("BASEUNITS") or r.get("baseunits") or r.get("unit") or "").strip(),
            "hsn_source": "",
            "hsn": (r.get("hsn") or r.get("HSN") or r.get("hsncode") or "").strip(),
            "description": (r.get("description") or r.get("Description") or "").strip(),
            "gst_applicable": (r.get("gstapplicable") or r.get("GSTAPPLICABLE") or "").strip(),
            "gst_rate_source": "",
            "gst_rate": _safe_float(r.get("gstrate") or r.get("GST_RATE") or 0),
            "taxability": (r.get("taxability") or "").strip(),
            "supply_type": (r.get("supplytype") or "").strip(),
            "rate_of_duty": _safe_float(r.get("rateofduty") or 0),
            "qty": _safe_float(r.get("qty") or 0),
            "qty_unit": (r.get("qty_unit") or "").strip(),
            "rate": _safe_float(r.get("rate") or 0),
            "rate_unit": (r.get("rate_unit") or "").strip(),
            "value": _safe_float(r.get("value") or 0),
        }
        try:
            database_manager.insert_or_update_item(data)
            imported["items"] += 1
        except Exception as e:
            imported["errors"].append(f"Item '{name}': {e}")

    return jsonify({
        "status": "ok",
        "active_db": db_name,
        "company_from_file": company_from_file,
        "imported": imported,
    })


@app.route('/api/db/import-items-xml', methods=['POST'])
def api_db_import_items_xml():
    """
    Import Items XML (Stock Groups + Stock Items) into SQLite.
    Uses items/items_backend.py parsing so all item fields get populated.
    Supports multi-company by selecting/creating a DB file.
    """
    if not database_manager:
        return jsonify({"error": "DB Manager not loaded"}), 500
    if not items_module or not hasattr(items_module, "import_items_from_exported_xml"):
        return jsonify({"error": "Items backend not available"}), 500

    mf = request.files.get('items_file') or request.files.get('xml_file') or request.files.get('file')
    if not mf or not mf.filename:
        return jsonify({"error": "items_file (.xml) is required"}), 400

    target = (request.form.get('target') or 'existing').strip().lower()  # existing | new
    requested_db = request.form.get('db_name') or request.form.get('company_db') or ''

    xml_bytes = mf.read()
    if not xml_bytes:
        return jsonify({"error": "Empty file"}), 400

    # Decode XML (best-effort)
    xml_text = ""
    try:
        xml_text = xml_bytes.decode("utf-16", errors="ignore")
    except Exception:
        try:
            xml_text = xml_bytes.decode("utf-8", errors="ignore")
        except Exception:
            xml_text = ""

    # Extract company name from XML (optional default db name)
    m = re.search(r"<SVCURRENTCOMPANY>(.*?)</SVCURRENTCOMPANY>", xml_text, flags=re.IGNORECASE | re.DOTALL)
    company_from_file = (m.group(1).strip() if m else "")

    if target == "new":
        chosen = requested_db or company_from_file or "new_company"
        db_name = _sanitize_db_filename(chosen)
    else:
        db_name = session.get("active_db", "tally_data.db")
        db_name = _sanitize_db_filename(db_name)

    # Set active db for this request + session
    session["active_db"] = db_name
    database_manager.set_active_db(db_name)
    database_manager.init_db(db_name=db_name)

    try:
        imported = items_module.import_items_from_exported_xml(xml_text, save_to_db=True)
    except Exception as e:
        return jsonify({"error": f"Failed to import items: {e}"}), 500

    return jsonify({
        "status": "ok",
        "active_db": db_name,
        "company_from_file": company_from_file,
        "imported": imported,
    })

@app.route('/api/db/items', methods=['GET'])
def api_db_items():
    if not database_manager: return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        items = database_manager.get_all_items()
        return jsonify({"items": items, "count": len(items)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/groups', methods=['GET'])
def api_db_groups():
    if not database_manager: return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        groups = database_manager.get_all_groups()
        return jsonify({"groups": groups, "count": len(groups)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/cost-categories', methods=['GET'])
def api_db_cost_categories():
    if not database_manager: return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        data = database_manager.get_all_cost_categories()
        return jsonify({"categories": data, "count": len(data)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/cost-centres', methods=['GET'])
def api_db_cost_centres():
    if not database_manager: return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        data = database_manager.get_all_cost_centres()
        return jsonify({"centres": data, "count": len(data)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ---------------------------------------------------------
# OPENING BALANCE — Upload Tally file → Download Zoho Excel
# ---------------------------------------------------------

@app.route('/opening-balance')
def opening_balance_page():
    return render_template('opening_balance.html')

@app.route('/api/opening-balance/convert', methods=['POST'])
def api_convert_opening_balance():
    """
    Single-file upload: tally_file
    ?preview=1  → JSON summary + preview rows
    ?preview=0  → Excel file download
    """
    try:
        from opening_balance_converter import convert
        import io

        tally_f = request.files.get('tally_file')
        if not tally_f or not tally_f.filename:
            return jsonify({"error": "Tally file is required"}), 400

        tally_bytes    = tally_f.read()
        migration_date = request.form.get('migration_date', None)
        preview_only   = request.args.get('preview', '1') == '1'

        # Fetch existing Groups from DB to filter out Group Headers
        db_group_names = []
        if database_manager:
            try:
                # get_all_groups returns list of dicts: [{'name': 'X'}, ...]
                all_groups = database_manager.get_all_groups()
                db_group_names = [g['name'] for g in all_groups if g.get('name')]
            except Exception as e:
                print(f"Warning: Failed to fetch groups for filtering: {e}")

        output_bytes, summary, errors = convert(
            tally_bytes, tally_f.filename, migration_date, db_group_names
        )

        if output_bytes is None:
            return jsonify({"error": errors[0] if errors else "Conversion failed"}), 400

        if preview_only:
            return jsonify({"status": "ok", "summary": summary, "errors": errors})
        else:
            return send_file(
                io.BytesIO(output_bytes),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='zoho_opening_balance.xlsx'
            )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



@app.route('/api/cost-centers/fetch', methods=['GET'])
def api_fetch_cost_centers():
    try:
        from cost_centers import cost_center_backend
        data = cost_center_backend.get_all_cost_data()
        return jsonify({"status": "success", "data": data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/cost-centers/sync-reporting-tags', methods=['POST'])
def api_sync_reporting_tags():
    try:
        from cost_centers import cost_center_backend
        result = cost_center_backend.sync_reporting_tags_to_zoho()
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/')
def index():
    return render_template('ledgers.html')

@app.route('/ledgers')
def ledgers_page():
    return render_template('ledgers.html')

@app.route('/items')
def items_page():
    return render_template('items.html')

# ---------------------------------------------------------
# API ENDPOINTS
# ---------------------------------------------------------

@app.route('/api/ledgers/fetch', methods=['GET'])
def api_fetch_ledgers():
    try:
        data = ledgers_module.analyze_ledgers_and_groups()
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch data from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/items/fetch', methods=['GET'])
def api_fetch_items():
    try:
        data = items_module.get_all_items_data()
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch items from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/ledgers/sync_zoho', methods=['POST'])
def api_sync_ledgers(): 
    try:
        selected = request.json.get("ledgers") if request.is_json else None
        update_existing = request.json.get("update_existing", False) if request.is_json else False
        result = ledgers_module.sync_ledgers_to_zoho(selected, update_existing=update_existing)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/ledgers/sync_customers', methods=['POST'])
def api_sync_customers():
    """Sync ONLY customers to Zoho Books."""
    try:
        selected = request.json.get("ledgers") if request.is_json else None
        update_existing = request.json.get("update_existing", True) if request.is_json else True
        result = ledgers_module.sync_ledgers_to_zoho(selected, contact_type_filter='customer', update_existing=update_existing)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/ledgers/sync_vendors', methods=['POST'])
def api_sync_vendors():
    """Sync ONLY vendors to Zoho Books."""
    try:
        selected = request.json.get("ledgers") if request.is_json else None
        update_existing = request.json.get("update_existing", True) if request.is_json else True
        result = ledgers_module.sync_ledgers_to_zoho(selected, contact_type_filter='vendor', update_existing=update_existing)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/ledgers/save_mapping', methods=['POST'])
def api_save_group_mapping():
    try:
        mapping = request.json.get("mapping") if request.is_json else {}
        ledgers_module.save_groups_mapping(mapping)
        return jsonify({"status": "success", "message": "Mapping saved successfully"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/ledgers/get_mapping', methods=['GET'])
def api_get_group_mapping():
    try:
        mapping = ledgers_module.get_groups_mapping()
        return jsonify(mapping)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/ledgers/execute_group_sync', methods=['POST'])
def api_execute_group_sync():
    try:
        # Load mapping from file in backend
        result = ledgers_module.sync_groups_to_zoho(None)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/ledgers/create_standalone', methods=['POST'])
def api_create_standalone():
    try:
        ledger_name = request.json.get("ledger_name") if request.is_json else None
        account_type = request.json.get("account_type") if request.is_json else None
        if not ledger_name or not account_type:
            return jsonify({"status": "error", "message": "ledger_name and account_type are required"}), 400
        result = ledgers_module.create_standalone_account(ledger_name, account_type)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/items/sync_zoho', methods=['POST'])
def api_sync_items():
    try:
        selected = request.json.get("items") if request.is_json else None
        result = items_module.sync_items_to_zoho(selected)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/items/sync_zoho/start', methods=['POST'])
def api_items_sync_zoho_start():
    if not items_module:
        return jsonify({"status": "error", "message": "Items backend not available"}), 500
    if not job_manager or not sse_format:
        return jsonify({"status": "error", "message": "Job manager not available"}), 500

    body = request.get_json(force=True, silent=True) or {}
    selected = body.get("items")

    job = job_manager.create("items_sync_zoho")
    job.log("Items sync job started.")

    def _runner():
        try:
            res = items_module.sync_items_to_zoho(selected, log=job.log, stop_event=job.stop_event)
            st = (res or {}).get("status")
            if st == "success":
                job_manager.finish(job.id, "success", result=res)
            elif st == "stopped":
                job_manager.finish(job.id, "stopped", result=res, message="Stopped by user")
            else:
                job_manager.finish(job.id, "error", result=res, message=(res or {}).get("message", "Failed"))
        except Exception as e:
            job.log(f"Unhandled error: {e}")
            job_manager.finish(job.id, "error", result={"status": "error", "message": str(e)}, message=str(e))

    threading.Thread(target=_runner, daemon=True).start()
    return jsonify({"status": "success", "job_id": job.id})


@app.route('/api/jobs/<job_id>', methods=['GET'])
def api_job_status(job_id):
    if not job_manager:
        return jsonify({"status": "error", "message": "Job manager not available"}), 500
    job = job_manager.get(job_id)
    if not job:
        return jsonify({"status": "error", "message": "Job not found"}), 404
    return jsonify({"status": "success", "job": job.snapshot(), "result": job.result if job.status != "running" else None})


@app.route('/api/jobs/<job_id>/stop', methods=['POST'])
def api_job_stop(job_id):
    if not job_manager:
        return jsonify({"status": "error", "message": "Job manager not available"}), 500
    ok = job_manager.stop(job_id)
    if not ok:
        return jsonify({"status": "error", "message": "Job not found"}), 404
    return jsonify({"status": "success"})


@app.route('/api/jobs/<job_id>/stream', methods=['GET'])
def api_job_stream(job_id):
    if not job_manager or not sse_format:
        return jsonify({"status": "error", "message": "Job manager not available"}), 500
    job = job_manager.get(job_id)
    if not job:
        return jsonify({"status": "error", "message": "Job not found"}), 404

    def _gen():
        last_seq = 0
        try:
            last_seq = int((request.args.get("from") or "0").strip() or 0)
        except Exception:
            last_seq = 0

        yield sse_format("status", job.snapshot())

        while True:
            entries = job.get_logs_since(last_seq)
            for seq, ts, msg in entries:
                last_seq = seq
                yield sse_format("log", {"seq": seq, "ts": ts, "message": msg})

            if job.status != "running":
                yield sse_format("done", {"job": job.snapshot(), "result": job.result})
                break

            job.wait(timeout=1.5)
            yield ": keepalive\n\n"

    headers = {
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    }
    return Response(stream_with_context(_gen()), headers=headers, mimetype="text/event-stream")


@app.route('/api/items/inventory_adjustment/preview', methods=['POST'])
def api_items_inventory_adjustment_preview():
    """
    Upload Excel (opening stock totals) + XML (godown/warehouse summary) and return:
    - items to apply (non-negative)
    - negative items report
    """
    try:
        excel_file = request.files.get("excel_file")
        xml_file = request.files.get("xml_file")
        if not excel_file or not xml_file:
            return jsonify({"status": "error", "message": "excel_file and xml_file are required"}), 400

        upload_dir = os.path.join(os.path.dirname(__file__), "uploads", "inventory_adjustment")
        os.makedirs(upload_dir, exist_ok=True)

        ts = str(int(time.time()))
        excel_path = os.path.join(upload_dir, f"opening_{ts}.xlsx")
        xml_path = os.path.join(upload_dir, f"godown_{ts}.xml")
        excel_file.save(excel_path)
        xml_file.save(xml_path)

        excel_map = items_module.parse_opening_excel_xlsx(excel_path)
        xml_map = items_module.parse_godown_xml(xml_path)
        to_apply, negative, stats = items_module.compute_inventory_adjustment(excel_map, xml_map)

        # Store last preview file paths in session for apply
        session["inv_adj_excel_path"] = excel_path
        session["inv_adj_xml_path"] = xml_path
        session["inv_adj_run_id"] = ts

        return jsonify({
            "status": "success",
            "stats": stats,
            "to_apply_count": len(to_apply),
            "negative_count": len(negative),
            "negative_preview": negative[:200],
            "apply_preview": to_apply[:50],
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/items/inventory_adjustment/apply', methods=['POST'])
def api_items_inventory_adjustment_apply():
    """
    Apply the latest preview (stored in session) to Zoho.
    """
    try:
        dry_run = str(request.form.get("dry_run", "0")).strip() in ("1", "true", "yes", "on")
        resume = str(request.form.get("resume", "1")).strip() in ("1", "true", "yes", "on")
        limit_raw = (request.form.get("limit") or "").strip()
        limit = 0
        try:
            limit = int(limit_raw) if limit_raw else 0
        except Exception:
            limit = 0
        excel_path = session.get("inv_adj_excel_path")
        xml_path = session.get("inv_adj_xml_path")
        run_id = session.get("inv_adj_run_id") or ""
        if not excel_path or not xml_path or (not os.path.exists(excel_path)) or (not os.path.exists(xml_path)):
            return jsonify({"status": "error", "message": "No preview found. Run Preview first."}), 400

        excel_map = items_module.parse_opening_excel_xlsx(excel_path)
        xml_map = items_module.parse_godown_xml(xml_path)
        to_apply, negative, stats = items_module.compute_inventory_adjustment(excel_map, xml_map)

        # Apply only non-negative items (optional limit for testing)
        if limit and limit > 0:
            to_apply = to_apply[:limit]
        result = items_module.apply_inventory_adjustment_to_zoho(to_apply, dry_run=dry_run, run_id=run_id, resume=resume)
        result["stats"] = stats
        result["negative_count"] = len(negative)
        result["negative_preview"] = negative[:200]
        result["applied_limit"] = limit or 0
        result["resume"] = bool(resume)
        result["run_id"] = run_id

        # Build a run report (xlsx) and store path in session for download
        try:
            from openpyxl import Workbook
            wb = Workbook()

            # Summary sheet
            ws = wb.active
            ws.title = "Summary"
            ws.append(["Key", "Value"])
            ws.append(["xml_items", stats.get("xml_items")])
            ws.append(["excel_items", stats.get("excel_items")])
            ws.append(["to_apply", stats.get("to_apply")])
            ws.append(["negative", stats.get("negative")])
            ws.append(["applied_limit", limit or 0])
            ws.append(["dry_run", bool(dry_run)])

            res_core = (result.get("results") or {})
            ws.append(["updated", res_core.get("updated", 0)])
            ws.append(["failed", res_core.get("failed", 0)])
            ws.append(["missing_item_id", res_core.get("missing_item_id", 0)])
            ws.append(["skipped", res_core.get("skipped", 0)])

            # Updated sheet
            ws_u = wb.create_sheet("Updated")
            ws_u.append(["Item Name", "Zoho Item ID"])
            for it in (res_core.get("updated_items") or []):
                ws_u.append([it.get("name", ""), it.get("item_id", "")])

            # Skipped sheet (includes missing_item_id + dry_run)
            ws_s = wb.create_sheet("Skipped")
            ws_s.append(["Item Name", "Reason", "Zoho Item ID"])
            for it in (res_core.get("skipped_items") or []):
                ws_s.append([it.get("name", ""), it.get("reason", ""), it.get("item_id", "")])

            # Errors sheet
            ws_e = wb.create_sheet("Errors")
            ws_e.append(["Item Name", "Reason"])
            for er in (res_core.get("errors") or []):
                ws_e.append([er.get("name", ""), er.get("reason", "")])

            # Negative sheet
            ws_n = wb.create_sheet("Negative")
            ws_n.append(["Item Name", "Reason", "Excel Qty", "XML Main Qty", "XML Other Qty", "Needed Main Qty"])
            for r in negative:
                ws_n.append([
                    r.get("name", ""),
                    r.get("reason", ""),
                    r.get("excel_qty", ""),
                    r.get("xml_main_qty", ""),
                    r.get("xml_other_qty", ""),
                    r.get("needed_main_qty", ""),
                ])

            upload_dir = os.path.join(os.path.dirname(__file__), "uploads", "inventory_adjustment")
            os.makedirs(upload_dir, exist_ok=True)
            ts2 = str(int(time.time()))
            report_path = os.path.join(upload_dir, f"inv_adj_run_report_{ts2}.xlsx")
            wb.save(report_path)
            session["inv_adj_last_run_report"] = report_path
            result["run_report_url"] = "/api/items/inventory_adjustment/last_run_report.xlsx"
        except Exception:
            pass

        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/items/inventory_adjustment/last_run_report.xlsx', methods=['GET'])
def api_items_inventory_adjustment_last_run_report():
    """Download the last Inventory Adjustment apply run report (xlsx)."""
    try:
        p = session.get("inv_adj_last_run_report")
        if not p or not os.path.exists(p):
            return jsonify({"status": "error", "message": "No run report found. Run Apply first."}), 400
        return send_file(
            p,
            as_attachment=True,
            download_name="inventory_adjustment_run_report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/items/inventory_adjustment/negative_report.xlsx', methods=['GET'])
def api_items_inventory_adjustment_negative_report():
    """
    Download full negative report as Excel (based on last preview stored in session).
    """
    try:
        excel_path = session.get("inv_adj_excel_path")
        xml_path = session.get("inv_adj_xml_path")
        if not excel_path or not xml_path or (not os.path.exists(excel_path)) or (not os.path.exists(xml_path)):
            return jsonify({"status": "error", "message": "No preview found. Run Preview first."}), 400

        excel_map = items_module.parse_opening_excel_xlsx(excel_path)
        xml_map = items_module.parse_godown_xml(xml_path)
        _to_apply, negative, _stats = items_module.compute_inventory_adjustment(excel_map, xml_map)

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Negative Items"
        ws.append(["Item Name", "Reason", "Excel Qty", "XML Main Qty", "XML Other Qty", "Needed Main Qty"])
        for r in negative:
            ws.append([
                r.get("name", ""),
                r.get("reason", ""),
                r.get("excel_qty", ""),
                r.get("xml_main_qty", ""),
                r.get("xml_other_qty", ""),
                r.get("needed_main_qty", ""),
            ])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name="negative_stock_report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/items/inventory_adjustment/matched_report.xlsx', methods=['GET'])
def api_items_inventory_adjustment_matched_report():
    """
    Download matched (Excel+XML) report as Excel (based on last preview stored in session).
    """
    try:
        excel_path = session.get("inv_adj_excel_path")
        xml_path = session.get("inv_adj_xml_path")
        if not excel_path or not xml_path or (not os.path.exists(excel_path)) or (not os.path.exists(xml_path)):
            return jsonify({"status": "error", "message": "No preview found. Run Preview first."}), 400

        excel_map = items_module.parse_opening_excel_xlsx(excel_path)
        xml_map = items_module.parse_godown_xml(xml_path)
        matched = items_module.build_matched_items_report(excel_map, xml_map)

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Matched Items"
        ws.append([
            "Item Name",
            "Status",
            "Excel Qty",
            "Excel Rate",
            "XML Main Qty",
            "XML Other Qty",
            "Needed Main Qty",
            "ExcelQty + XMLMainQty",
        ])
        for r in matched:
            ws.append([
                r.get("name", ""),
                r.get("status", ""),
                r.get("excel_qty", 0),
                r.get("excel_rate", 0),
                r.get("xml_main_qty", 0),
                r.get("xml_other_qty", 0),
                r.get("needed_main_qty", 0),
                r.get("sum_qty", 0),
            ])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name="matched_items_report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/items/godown_xml/to_transfer_excel.xlsx', methods=['POST'])
def api_items_godown_xml_to_transfer_excel():
    """
    Upload a Godown Summary XML (MAIN warehouse export) and download an Excel file
    in the same 4-column format as 'Stocks Transferred to Main warehouse.xlsx'.
    """
    try:
        xml_file = request.files.get("xml_file")
        if not xml_file:
            return jsonify({"status": "error", "message": "xml_file is required"}), 400

        upload_dir = os.path.join(os.path.dirname(__file__), "uploads", "inventory_adjustment")
        os.makedirs(upload_dir, exist_ok=True)
        ts = str(int(time.time()))
        xml_path = os.path.join(upload_dir, f"godown_transfer_{ts}.xml")
        xml_file.save(xml_path)

        xml_map = items_module.parse_godown_xml(xml_path)
        wb = items_module.build_stock_transfer_workbook_from_godown_xml(xml_map)

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name="stocks_transferred_to_main_warehouse.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/items/sync_map/refresh_from_zoho', methods=['POST'])
def api_items_refresh_sync_map_from_zoho():
    """
    Build local Zoho item_id map by fetching Zoho items list (no create/update).
    This makes Inventory Adjustment Apply work without re-running full Items sync.
    """
    try:
        if not items_module or not hasattr(items_module, "refresh_zoho_item_sync_map_from_zoho"):
            return jsonify({"status": "error", "message": "Items module not available"}), 500
        result = items_module.refresh_zoho_item_sync_map_from_zoho()
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────
#  FIELD MAPPING  –  Items
# ─────────────────────────────────────────────────────────────
MODULE_DB_FIELDS = {
    "items": [
        "name", "group_name", "category", "unit",
        "hsn_source", "hsn", "description",
        "gst_applicable", "gst_rate_source", "gst_rate",
        "taxability", "supply_type", "rate_of_duty",
        "qty", "qty_unit", "rate", "rate_unit", "value"
    ],
    "bills": [
        "voucher_number", "date", "vendor_name", "total_amount",
        "tax_amount", "payment_status", "due_date", "reference_number", "narration"
    ],
    "invoices": [
        "voucher_number", "date", "customer_name", "total_amount",
        "tax_amount", "payment_status", "due_date", "reference_number", "narration"
    ],
    "journals": [
        "voucher_number", "date", "narration", "debit_account",
        "credit_account", "amount", "reference"
    ],
    "receipts": [
        "voucher_number", "date", "party_name", "amount",
        "payment_mode", "reference_number", "narration", "bank_account"
    ],
    "payments_made": [
        "voucher_number", "date", "party_name", "amount",
        "payment_mode", "reference_number", "narration", "bank_account"
    ],
    "sales_orders": [
        "voucher_number", "date", "customer_name", "total_amount",
        "reference_number", "status", "narration"
    ],
    "purchase_orders": [
        "voucher_number", "date", "vendor_name", "total_amount",
        "reference_number", "status", "narration"
    ],
    "contra": [
        "voucher_number", "date", "from_account", "to_account",
        "amount", "narration", "reference_number"
    ],
    "credit_note": [
        "voucher_number", "date", "customer_name", "total_amount",
        "reference_number", "narration"
    ],
    "debit_note": [
        "voucher_number", "date", "vendor_name", "total_amount",
        "reference_number", "narration"
    ],
    "opening_balance": [
        "account_name", "group_name", "opening_balance",
        "balance_type", "date"
    ],
}


@app.route('/api/field-mapping/<module>/upload-zoho-fields', methods=['POST'])
def api_upload_zoho_fields(module):
    """Upload a Zoho Books XLSX sample file; extract column headers as Zoho field names."""
    if openpyxl is None:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
    file = request.files.get('file')
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    filename = (file.filename or "").lower()
    try:
        raw = file.read()
        headers = []

        if filename.endswith('.csv'):
            import csv as _csv
            # Try UTF-8-BOM first (Excel default), fall back to latin-1
            for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
                try:
                    text = raw.decode(enc)
                    break
                except UnicodeDecodeError:
                    pass
            reader = _csv.reader(io.StringIO(text))
            first_row = next(reader, [])
            headers = [h.strip() for h in first_row if h.strip()]
        else:
            # Default: treat as XLSX
            if openpyxl is None:
                return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            for cell in next(ws.iter_rows(max_row=1)):
                val = str(cell.value).strip() if cell.value is not None else ""
                if val:
                    headers.append(val)

        if not headers:
            return jsonify({"error": "No column headers found in the uploaded file"}), 400
        field_mapping_manager.save_zoho_fields(module, headers)
        return jsonify({"status": "ok", "zoho_fields": headers, "count": len(headers)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/field-mapping/<module>/zoho-fields', methods=['GET'])
def api_get_zoho_fields(module):
    """Return saved Zoho field names for a module."""
    data = field_mapping_manager.load(module)
    return jsonify({
        "zoho_fields": data.get("zoho_fields", []),
        "has_fields": bool(data.get("zoho_fields")),
    })


@app.route('/api/field-mapping/<module>/db-fields', methods=['GET'])
def api_get_db_fields(module):
    """Return our local DB field names for a module."""
    fields = MODULE_DB_FIELDS.get(module, [])
    return jsonify({"db_fields": fields})


@app.route('/api/field-mapping/<module>/mapping', methods=['GET'])
def api_get_mapping(module):
    """Return saved mapping for a module."""
    data = field_mapping_manager.load(module)
    return jsonify({
        "mapping": data.get("mapping", {}),
        "zoho_fields": data.get("zoho_fields", []),
    })


@app.route('/api/field-mapping/<module>/mapping', methods=['POST'])
def api_save_mapping(module):
    """Save DB→Zoho field mapping."""
    body = request.get_json(force=True) or {}
    mapping = body.get("mapping", {})
    if not mapping:
        return jsonify({"error": "mapping object is required"}), 400
    field_mapping_manager.save_mapping(module, mapping)
    return jsonify({"status": "ok", "saved": len(mapping)})


@app.route('/api/field-mapping/<module>/export', methods=['GET'])
def api_export_mapped(module):
    """Export DB records as an XLSX formatted with Zoho Books column headers based on saved mapping."""
    if openpyxl is None:
        return jsonify({"error": "openpyxl not installed"}), 500
    data = field_mapping_manager.load(module)
    mapping = data.get("mapping", {})   # {zoho_field: db_field}
    if not mapping:
        return jsonify({"error": "No field mapping saved yet. Please map fields first."}), 400
    # Fetch records based on module
    records = []
    try:
        if module == "items" and database_manager:
            records = database_manager.get_all_items()
        elif module == "bills" and database_manager:
            records = database_manager.get_all_vouchers("purchase")
        elif module == "invoices" and database_manager:
            records = database_manager.get_all_vouchers("sales")
        elif module == "journals" and database_manager:
            records = database_manager.get_all_vouchers("journal")
        elif module == "receipts" and database_manager:
            records = database_manager.get_all_vouchers("receipt")
        elif module == "payments_made" and database_manager:
            records = database_manager.get_all_vouchers("payment")
        elif module == "sales_orders" and database_manager:
            records = database_manager.get_all_vouchers("sales order")
        elif module == "purchase_orders" and database_manager:
            records = database_manager.get_all_vouchers("purchase order")
        elif module == "contra" and database_manager:
            records = database_manager.get_all_vouchers("contra")
        elif module == "credit_note" and database_manager:
            records = database_manager.get_all_vouchers("credit note")
        elif module == "debit_note" and database_manager:
            records = database_manager.get_all_vouchers("debit note")
    except Exception as e:
        records = []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = module.replace("_", " ").title()

    # zoho_fields order as columns
    zoho_fields = list(mapping.keys())
    ws.append(zoho_fields)   # header row

    for rec in records:
        row = []
        if isinstance(rec, dict):
            for zf in zoho_fields:
                db_col = mapping.get(zf, "")
                row.append(rec.get(db_col, ""))
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{module}_zoho_import.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# Journal routes
@app.route('/journals')
def journals_page():
    return render_template('journals.html')

@app.route('/api/journals/fetch', methods=['POST'])
def api_fetch_journals():
    try:
        # Get date range from request
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        data = journel_module.get_all_journals_data(from_date, to_date, limit)
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch journals from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/journals/sync_zoho', methods=['POST'])
def api_sync_journals():
    try:
        selected = request.json.get("journals") if request.is_json else None
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        result = journel_module.sync_journals_to_zoho(selected, from_date, to_date, limit)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/journals/upload', methods=['POST'])
def api_upload_journals():
    """
    Upload a Tally-exported JSON file for journals (offline mode).
    Parses the file using journel_backend.parse_tally_json,
    saves to SQLite, and returns the journals list + stats.
    """
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({"error": "No selected file"}), 400

        import tempfile, os
        from datetime import datetime as _dt

        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.json') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        try:
            # Parse using existing journel_backend parser
            parsed = journel_module.parse_tally_json(tmp_path)
        finally:
            os.unlink(tmp_path)

        if not parsed:
            return jsonify({"error": "No journals found in the JSON file. Make sure it is a Tally Journal voucher export."}), 400

        # Convert parse_tally_json output (ledger_entries) to the format
        # expected by the UI (line_items with debit_or_credit field)
        journals_out = []
        for j in parsed:
            line_items = []
            for entry in j.get("ledger_entries", []):
                debit  = float(entry.get("debit",  0) or 0)
                credit = float(entry.get("credit", 0) or 0)
                if debit > 0:
                    line_items.append({
                        "ledger_name":    entry.get("ledger_name", ""),
                        "ledger_type":    "account",
                        "amount":         debit,
                        "debit_or_credit":"debit",
                        "tag_category":   "",
                        "tag_option":     ""
                    })
                if credit > 0:
                    line_items.append({
                        "ledger_name":    entry.get("ledger_name", ""),
                        "ledger_type":    "account",
                        "amount":         credit,
                        "debit_or_credit":"credit",
                        "tag_category":   "",
                        "tag_option":     ""
                    })

            journals_out.append({
                "date":           j.get("date", ""),
                "journal_number": j.get("journal_number", ""),
                "narration":      j.get("narration", ""),
                "tally_guid":     j.get("tally_guid", ""),
                "voucher_type":   j.get("voucher_type", "Journal"),
                "line_items":     line_items,
                "cost_center_allocations": j.get("cost_center_allocations", [])
            })

        # ── Save to SQLite so the DB tab also shows them ──────────────────
        if database_manager and journals_out:
            now = _dt.now().isoformat()
            db_data_list = []
            for jrnl in journals_out:
                td = sum(i["amount"] for i in jrnl["line_items"] if i["debit_or_credit"] == "debit")
                tc = sum(i["amount"] for i in jrnl["line_items"] if i["debit_or_credit"] == "credit")
                db_data_list.append({
                    "journal_number": jrnl["journal_number"],
                    "date":           jrnl["date"],
                    "narration":      jrnl["narration"],
                    "total_debit":    round(td, 2),
                    "total_credit":   round(tc, 2),
                    "line_items":     json.dumps(jrnl["line_items"]),
                    "from_date":      "",
                    "to_date":        "",
                    "created_at":     now,
                    "updated_at":     now,
                })
            try:
                database_manager.bulk_save_journals(db_data_list)
            except AttributeError:
                pass  # bulk_save_journals may not exist on older DB manager

        total_debit  = sum(
            i["amount"] for j in journals_out for i in j["line_items"]
            if i["debit_or_credit"] == "debit"
        )
        total_credit = sum(
            i["amount"] for j in journals_out for i in j["line_items"]
            if i["debit_or_credit"] == "credit"
        )

        return jsonify({
            "journals": journals_out,
            "stats": {
                "total_journals": len(journals_out),
                "total_debit":    round(total_debit,  2),
                "total_credit":   round(total_credit, 2),
            }
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



# Invoice routes
@app.route('/invoices')
def invoices_page():
    return render_template('invoices.html')

@app.route('/api/invoices/fetch', methods=['POST'])
def api_fetch_invoices():
    try:
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        data = invoice_module.get_all_invoices_data(from_date, to_date, limit)
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch invoices from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/invoices/sync_zoho', methods=['POST'])
def api_sync_invoices():
    try:
        selected = request.json.get("invoices") if request.is_json else None
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        result = invoice_module.sync_invoices_to_zoho(selected, from_date, to_date, limit)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# Bills routes
@app.route('/bills')
def bills_page():
    return render_template('bills.html')

@app.route('/api/bills/fetch', methods=['POST'])
def api_fetch_bills():
    try:
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        data = bills_module.get_all_bills_data(from_date, to_date, limit)
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch bills from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/bills/sync_zoho', methods=['POST'])
def api_sync_bills():
    try:
        selected = request.json.get("bills") if request.is_json else None
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        result = bills_module.sync_bills_to_zoho(selected, from_date, to_date, limit)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# Sales Order routes
@app.route('/sales_orders')
def sales_orders_page():
    return render_template('sales_orders.html')

@app.route('/api/sales_orders/fetch', methods=['POST'])
def api_fetch_sales_orders():
    try:
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        data = sales_order_module.get_all_sales_orders_data(from_date, to_date, limit)
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch sales orders from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/sales_orders/sync_zoho', methods=['POST'])
def api_sync_sales_orders():
    try:
        selected = request.json.get("sales_orders") if request.is_json else None
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        result = sales_order_module.sync_sales_orders_to_zoho(selected, from_date, to_date, limit)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# Purchase Order routes
@app.route('/purchase_orders')
def purchase_orders_page():
    return render_template('purchase_orders.html')

@app.route('/api/purchase_orders/fetch', methods=['POST'])
def api_fetch_purchase_orders():
    try:
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        data = purchase_order_module.get_all_purchase_orders_data(from_date, to_date, limit)
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch purchase orders from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/purchase_orders/sync_zoho', methods=['POST'])
def api_sync_purchase_orders():
    try:
        selected = request.json.get("purchase_orders") if request.is_json else None
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        result = purchase_order_module.sync_purchase_orders_to_zoho(selected, from_date, to_date, limit)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# Payments Made routes
@app.route('/payments_made')
def payments_made_page():
    return render_template('payments_made.html')

@app.route('/api/payments_made/fetch', methods=['POST'])
def api_fetch_payments_made():
    try:
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        company_name = request.json.get("company_name") if request.is_json else None
        
        data = payments_module.get_all_payments_data(from_date, to_date, limit, company_name)
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch payments made from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/payments_made/upload', methods=['POST'])
def api_upload_payments_made():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400
        
        import tempfile
        import os
        from datetime import datetime
        with tempfile.NamedTemporaryFile(delete=False, suffix='.json') as temp:
            file.save(temp.name)
            temp_path = temp.name
            
        parsed_payments = payments_module.parse_tally_json(temp_path)
        
        # Save to SQLite
        if database_manager and parsed_payments:
            database_manager.init_db()
            db_data_list = []
            for payment in parsed_payments:
                db_data = {
                    "payment_number": payment.get("payment_number", ""),
                    "voucher_type": payment.get("voucher_type", ""),
                    "date": payment.get("date", ""),
                    "vendor_name": payment.get("vendor_name", ""),
                    "vendor_ledger_amount": payment.get("vendor_ledger_amount", 0) or 0,
                    "payment_mode": payment.get("payment_mode", ""),
                    "bank_account": payment.get("bank_account", ""),
                    "account_current_balance": payment.get("account_current_balance", 0) or 0,
                    "amount": payment.get("amount", 0) or 0,
                    "reference_number": payment.get("reference_number", ""),
                    "against_reference": payment.get("against_reference", ""),
                    "narration": payment.get("narration", ""),
                    "bill_allocations": json.dumps(payment.get("bill_allocations", [])),
                    "ledger_entries": json.dumps(payment.get("ledger_entries", [])),
                    "cost_center_allocations": json.dumps(payment.get("cost_center_allocations", [])),
                    "rounding_amount": payment.get("rounding_amount", 0) or 0,
                    "rounding_ledger": payment.get("rounding_ledger", ""),
                    "tally_guid": payment.get("tally_guid", ""),
                    "company_name": "",
                    "created_at": datetime.now().isoformat(),
                    "updated_at": datetime.now().isoformat()
                }
                db_data_list.append(db_data)
                
            try:
                database_manager.bulk_save_payments_made(db_data_list)
            except AttributeError:
                pass
                
        os.unlink(temp_path)
        
        total_amount = sum(float(p.get("amount", 0)) for p in parsed_payments)
        return jsonify({"payments": parsed_payments, "total_amount": total_amount})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/payments_made/sync_zoho', methods=['POST'])
def api_sync_payments_made():
    try:
        selected = request.json.get("payments_made") if request.is_json else None
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        company_name = request.json.get("company_name") if request.is_json else None
        
        result = payments_module.sync_payments_to_zoho(selected, from_date, to_date, limit, company_name)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
# Receipts (Payment Received) routes
@app.route('/receipts')
def receipts_page():
    return render_template('receipts.html')

@app.route('/api/receipts/fetch', methods=['POST'])
def api_fetch_receipts():
    try:
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        company_name = request.json.get("company_name") if request.is_json else None
        
        data = receipts_module.get_all_receipts_data(from_date, to_date, limit, company_name)
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch receipts from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/payments_made/fetch/start', methods=['POST'])
def api_payments_made_fetch_start():
    if not payments_module:
        return jsonify({"status": "error", "message": "Payments backend not available"}), 500
    if not job_manager:
        return jsonify({"status": "error", "message": "Job manager not available"}), 500

    body = request.get_json(force=True, silent=True) or {}
    from_date = body.get("from_date", "20250401")
    to_date = body.get("to_date", "20250430")
    limit = body.get("limit")
    company_name = body.get("company_name")

    job = job_manager.create("payments_made_fetch")
    job.log(f"Payments Made fetch job started: {from_date} -> {to_date}")

    def _runner():
        try:
            fn = getattr(payments_module, "get_all_payments_data_day_by_day", None) or getattr(payments_module, "get_all_payments_data", None)
            if not callable(fn):
                raise RuntimeError("Payments fetch function not available")
            if getattr(fn, "__name__", "") == "get_all_payments_data_day_by_day":
                res = fn(from_date, to_date, limit, company_name, log=job.log, stop_event=job.stop_event)
            else:
                res = fn(from_date, to_date, limit, company_name)
            status = (res or {}).get("status") or "success"
            if job.stop_event.is_set() and status == "success":
                status = "stopped"
                res["status"] = "stopped"
            job_manager.finish(job.id, "success" if status == "success" else status, result=res)
        except Exception as e:
            job.log(f"Unhandled error: {e}")
            job_manager.finish(job.id, "error", result={"status": "error", "message": str(e)}, message=str(e))

    threading.Thread(target=_runner, daemon=True).start()
    return jsonify({"status": "success", "job_id": job.id})


@app.route('/api/receipts/fetch/start', methods=['POST'])
def api_receipts_fetch_start():
    if not receipts_module:
        return jsonify({"status": "error", "message": "Receipts backend not available"}), 500
    if not job_manager:
        return jsonify({"status": "error", "message": "Job manager not available"}), 500

    body = request.get_json(force=True, silent=True) or {}
    from_date = body.get("from_date", "20250401")
    to_date = body.get("to_date", "20250430")
    limit = body.get("limit")
    company_name = body.get("company_name")

    job = job_manager.create("receipts_fetch")
    job.log(f"Receipts fetch job started: {from_date} -> {to_date}")

    def _runner():
        try:
            fn = getattr(receipts_module, "get_all_receipts_data_day_by_day", None) or getattr(receipts_module, "get_all_receipts_data", None)
            if not callable(fn):
                raise RuntimeError("Receipts fetch function not available")
            res = fn(from_date, to_date, limit, company_name, log=job.log, stop_event=job.stop_event) if "day_by_day" in getattr(fn, "__name__", "") else fn(from_date, to_date, limit, company_name)
            status = (res or {}).get("status") or "success"
            if job.stop_event.is_set() and status == "success":
                status = "stopped"
                res["status"] = "stopped"
            job_manager.finish(job.id, "success" if status == "success" else status, result=res)
        except Exception as e:
            job.log(f"Unhandled error: {e}")
            job_manager.finish(job.id, "error", result={"status": "error", "message": str(e)}, message=str(e))

    threading.Thread(target=_runner, daemon=True).start()
    return jsonify({"status": "success", "job_id": job.id})

@app.route('/api/receipts/sync_zoho', methods=['POST'])
def api_sync_receipts():
    try:
        selected = request.json.get("receipts") if request.is_json else None
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        company_name = request.json.get("company_name") if request.is_json else None
        
        result = receipts_module.sync_receipts_to_zoho(selected, from_date, to_date, limit, company_name)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/receipts/zoho/sync/start', methods=['POST'])
def api_receipts_zoho_sync_start():
    if not receipts_module:
        return jsonify({"status": "error", "message": "Receipts backend not available"}), 500
    if not job_manager:
        return jsonify({"status": "error", "message": "Job manager not available"}), 500

    body = request.get_json(force=True, silent=True) or {}
    from_date = body.get("from_date", "20250401")
    to_date = body.get("to_date", "20250430")
    limit = body.get("limit")
    company_name = body.get("company_name")
    cutoff_date = (body.get("cutoff_date") or os.environ.get("MIGRATION_CUTOFF_DATE") or "2025-03-31").strip()
    opening_invoice_id = (body.get("opening_invoice_id") or os.environ.get("OPENING_BALANCE_ZOHO_INVOICE_ID") or "").strip()

    job = job_manager.create("receipts_zoho_sync")
    job.log(f"Receipts Zoho sync started: {from_date} -> {to_date} cutoff={cutoff_date}")

    def _runner():
        try:
            fn = getattr(receipts_module, "sync_receipts_to_zoho_job", None)
            if not callable(fn):
                raise RuntimeError("Receipts Zoho sync function not available")
            res = fn(from_date, to_date, limit, company_name, cutoff_date=cutoff_date, opening_invoice_id=opening_invoice_id, log=job.log, stop_event=job.stop_event)
            st = (res or {}).get("status") or "success"
            if st == "success":
                job_manager.finish(job.id, "success", result=res)
            elif st == "stopped":
                job_manager.finish(job.id, "stopped", result=res, message="Stopped by user")
            else:
                job_manager.finish(job.id, "error", result=res, message=(res or {}).get("message", "Failed"))
        except Exception as e:
            job.log(f"Unhandled error: {e}")
            job_manager.finish(job.id, "error", result={"status": "error", "message": str(e)}, message=str(e))

    threading.Thread(target=_runner, daemon=True).start()
    return jsonify({"status": "success", "job_id": job.id})

@app.route('/api/db/sales_orders', methods=['GET'])
def api_db_sales_orders():
    """Fetch Sales Order vouchers stored in SQLite database (tally_data.db)"""
    if not database_manager:
        return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        orders = database_manager.get_all_sales_orders()

        json_fields = ('customer_address', 'line_items', 'taxes')
        for so in orders:
            for field in json_fields:
                if so.get(field):
                    try:
                        if isinstance(so[field], str):
                            so[field] = json.loads(so[field])
                    except Exception:
                        so[field] = []
                else:
                    so[field] = []

        total_amount = sum(float(so.get('total_amount', 0) or 0) for so in orders)
        return jsonify({
            "sales_orders": orders,
            "count":        len(orders),
            "total_amount": round(total_amount, 2)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/purchase_orders', methods=['GET'])
def api_db_purchase_orders():
    """Fetch Purchase Order vouchers stored in SQLite database (tally_data.db)"""
    if not database_manager:
        return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        orders = database_manager.get_all_purchase_orders()

        json_fields = ('vendor_address', 'line_items', 'taxes')
        for po in orders:
            for field in json_fields:
                if po.get(field):
                    try:
                        if isinstance(po[field], str):
                            po[field] = json.loads(po[field])
                    except Exception:
                        po[field] = []
                else:
                    po[field] = []

        total_amount = sum(float(po.get('total_amount', 0) or 0) for po in orders)
        return jsonify({
            "purchase_orders": orders,
            "count":           len(orders),
            "total_amount":    round(total_amount, 2)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/bills', methods=['GET'])
def api_db_bills():
    """Fetch bill vouchers stored in SQLite database (tally_data.db)"""
    if not database_manager:
        return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        bills = database_manager.get_all_bills()

        # Parse JSON columns back to lists — same pattern as receipts
        json_fields = ('vendor_address', 'line_items', 'taxes')
        for bill in bills:
            for field in json_fields:
                if bill.get(field):
                    try:
                        if isinstance(bill[field], str):
                            bill[field] = json.loads(bill[field])
                    except Exception:
                        bill[field] = []
                else:
                    bill[field] = []

        total_amount = sum(float(bill.get('total_amount', 0) or 0) for bill in bills)

        return jsonify({
            "bills":        bills,
            "count":        len(bills),
            "total_amount": round(total_amount, 2)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/invoices', methods=['GET'])
def api_db_invoices():
    """Fetch invoice vouchers stored in SQLite database (tally_data.db)"""
    if not database_manager:
        return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        invoices = database_manager.get_all_invoices()

        # Parse JSON columns back to lists — same pattern as receipts
        json_fields = ('buyer_address', 'line_items', 'taxes')
        for inv in invoices:
            for field in json_fields:
                if inv.get(field):
                    try:
                        if isinstance(inv[field], str):
                            inv[field] = json.loads(inv[field])
                    except Exception:
                        inv[field] = []
                else:
                    inv[field] = []

        total_amount = sum(float(inv.get('total_amount', 0) or 0) for inv in invoices)

        return jsonify({
            "invoices":     invoices,
            "count":        len(invoices),
            "total_amount": round(total_amount, 2)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/journals', methods=['GET'])
def api_db_journals():
    """Fetch journal vouchers stored in SQLite database (tally_data.db)"""
    if not database_manager:
        return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        journals = database_manager.get_all_journals()

        # Parse line_items JSON back to list — same as receipts does for its JSON columns
        for j in journals:
            if j.get('line_items'):
                try:
                    if isinstance(j['line_items'], str):
                        j['line_items'] = json.loads(j['line_items'])
                except Exception:
                    j['line_items'] = []
            else:
                j['line_items'] = []

        total_debit  = sum(float(j.get('total_debit',  0) or 0) for j in journals)
        total_credit = sum(float(j.get('total_credit', 0) or 0) for j in journals)

        return jsonify({
            "journals":     journals,
            "count":        len(journals),
            "total_debit":  round(total_debit,  2),
            "total_credit": round(total_credit, 2)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/payments_made', methods=['GET'])
def api_db_payments_made():
    """Fetch payments made from SQLite database"""
    try:
        payments = database_manager.get_all_payments_made()
        
        for payment in payments:
            if payment.get('bill_allocations'):
                try:
                    if isinstance(payment['bill_allocations'], str):
                        payment['bill_allocations'] = json.loads(payment['bill_allocations'])
                except Exception as e:
                    print(f"️ Error parsing bill_allocations for payment {payment.get('payment_number')}: {e}")
                    payment['bill_allocations'] = []
            else:
                payment['bill_allocations'] = []
            
            if payment.get('ledger_entries'):
                try:
                    if isinstance(payment['ledger_entries'], str):
                        payment['ledger_entries'] = json.loads(payment['ledger_entries'])
                except Exception as e:
                    print(f"️ Error parsing ledger_entries for payment {payment.get('payment_number')}: {e}")
                    payment['ledger_entries'] = []
            else:
                payment['ledger_entries'] = []
            
            if payment.get('cost_center_allocations'):
                try:
                    if isinstance(payment['cost_center_allocations'], str):
                        payment['cost_center_allocations'] = json.loads(payment['cost_center_allocations'])
                except Exception as e:
                    print(f"️ Error parsing cost_center_allocations for payment {payment.get('payment_number')}: {e}")
                    payment['cost_center_allocations'] = []
            else:
                payment['cost_center_allocations'] = []
        
        total_amount = sum(float(p.get('amount', 0) or 0) for p in payments)
        
        return jsonify({
            "payments": payments,
            "count": len(payments),
            "total_amount": total_amount
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/db/receipts', methods=['GET'])
def api_db_receipts():
    """Fetch receipts from SQLite database"""
    try:
        receipts = database_manager.get_all_receipts()
        
        # Parse JSON fields back to lists/dicts
        for receipt in receipts:
            # Parse invoice_allocations
            if receipt.get('invoice_allocations'):
                try:
                    if isinstance(receipt['invoice_allocations'], str):
                        receipt['invoice_allocations'] = json.loads(receipt['invoice_allocations'])
                except Exception as e:
                    print(f"️ Error parsing invoice_allocations for receipt {receipt.get('receipt_number')}: {e}")
                    receipt['invoice_allocations'] = []
            else:
                receipt['invoice_allocations'] = []
            
            # Parse ledger_entries
            if receipt.get('ledger_entries'):
                try:
                    if isinstance(receipt['ledger_entries'], str):
                        receipt['ledger_entries'] = json.loads(receipt['ledger_entries'])
                except Exception as e:
                    print(f"️ Error parsing ledger_entries for receipt {receipt.get('receipt_number')}: {e}")
                    receipt['ledger_entries'] = []
            else:
                receipt['ledger_entries'] = []
            
            # Parse cost_center_allocations
            if receipt.get('cost_center_allocations'):
                try:
                    if isinstance(receipt['cost_center_allocations'], str):
                        receipt['cost_center_allocations'] = json.loads(receipt['cost_center_allocations'])
                except Exception as e:
                    print(f"️ Error parsing cost_center_allocations for receipt {receipt.get('receipt_number')}: {e}")
                    receipt['cost_center_allocations'] = []
            else:
                receipt['cost_center_allocations'] = []
        
        # Calculate stats
        total_amount = sum(float(r.get('amount', 0) or 0) for r in receipts)
        
        return jsonify({
            "receipts": receipts,
            "count": len(receipts),
            "total_amount": total_amount
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/journals/refresh_cache', methods=['POST'])
def api_refresh_cache():
    try:
        refresh_type = request.json.get("type", "all") if request.is_json else "all"
        
        stats = {}
        
        # Refresh Tally data
        if refresh_type in ["all", "tally"]:
            ledger_map = journel_module.get_ledger_map_from_tally(use_cache=False, force_refresh=True)
            if ledger_map:
                vendors = sum(1 for t in ledger_map.values() if t == "vendor")
                customers = sum(1 for t in ledger_map.values() if t == "customer")
                accounts = sum(1 for t in ledger_map.values() if t == "account")
                stats["tally"] = {
                    "ledgers": len(ledger_map),
                    "vendors": vendors,
                    "customers": customers,
                    "others": accounts
                }
        
        # Refresh Zoho data
        if refresh_type in ["all", "zoho"]:
            token = journel_module.get_access_token()
            if token:
                # Refresh accounts (Chart of Accounts)
                account_map = journel_module.get_zoho_accounts(token, use_cache=False, force_refresh=True)
                # Refresh contacts
                contact_map = journel_module.get_zoho_contacts(token, use_cache=False, force_refresh=True)
                
                # Count contact types
                zoho_vendors = sum(1 for c in contact_map.values() if c["contact_type"] == "vendor")
                zoho_customers = sum(1 for c in contact_map.values() if c["contact_type"] == "customer")
                
                stats["zoho"] = {
                    "total_contacts": len(contact_map) if contact_map else 0,
                    "vendors": zoho_vendors,
                    "customers": zoho_customers,
                    "chart_of_accounts": len(account_map) if account_map else 0
                }
        
        return jsonify({
            "status": "success",
            "stats": stats
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# ---------------------------------------------------------
# CONTRA ROUTES
# ---------------------------------------------------------
@app.route('/contra')
def contra_page():
    return render_template('contra.html')

@app.route('/api/contra/fetch', methods=['POST'])
def api_fetch_contra():
    try:
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        data = contra_module.get_all_contra_data(from_date, to_date, limit)
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch contra from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/contra/upload', methods=['POST'])
def api_upload_contra():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400
        
        import tempfile, os
        from datetime import datetime
        with tempfile.NamedTemporaryFile(delete=False, suffix='.json') as temp:
            file.save(temp.name)
            temp_path = temp.name
            
        parsed_contras = contra_module.parse_tally_json(temp_path)
        
        # Save to SQLite
        if database_manager and parsed_contras:
            database_manager.init_db()
            db_data_list = []
            for contra in parsed_contras:
                db_data = {
                    "contra_number": contra.get("contra_number", ""),
                    "voucher_type": contra.get("voucher_type", "Contra"),
                    "date": contra.get("date", ""),
                    "from_account": contra.get("from_account", ""),
                    "to_account": contra.get("to_account", ""),
                    "amount": contra.get("amount", 0) or 0,
                    "narration": contra.get("narration", ""),
                    "ledger_entries": json.dumps(contra.get("ledger_entries", [])),
                    "cost_center_allocations": json.dumps(contra.get("cost_center_allocations", [])),
                    "tally_guid": contra.get("tally_guid", ""),
                    "company_name": "",
                    "created_at": datetime.now().isoformat(),
                    "updated_at": datetime.now().isoformat()
                }
                db_data_list.append(db_data)
                
            try:
                database_manager.bulk_save_contra(db_data_list)
            except AttributeError:
                pass
                
        os.unlink(temp_path)
        
        total_amount = sum(float(c.get("amount", 0)) for c in parsed_contras)
        return jsonify({"contra_vouchers": parsed_contras, "total_amount": total_amount})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/contra/sync_zoho', methods=['POST'])
def api_sync_contra():
    try:
        selected = request.json.get("contras") if request.is_json else None
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        result = contra_module.sync_contra_to_zoho(selected, from_date, to_date, limit)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/db/contra', methods=['GET'])
def api_db_contra():
    if not database_manager: return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        c_rows = database_manager.get_all_contra()
        contras = []
        for r in c_rows:
            d = dict(r)
            if isinstance(d.get('ledger_entries'), str):
                try: d['ledger_entries'] = json.loads(d['ledger_entries'])
                except: d['ledger_entries'] = []
            if isinstance(d.get('cost_center_allocations'), str):
                try: d['cost_center_allocations'] = json.loads(d['cost_center_allocations'])
                except: d['cost_center_allocations'] = []
            contras.append(d)
                
        total_amount = sum(float(r.get('amount', 0) or 0) for r in contras)
        return jsonify({
            "contra_vouchers": contras,
            "count": len(contras),
            "total_amount": total_amount
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ---------------------------------------------------------
# CREDIT NOTE ROUTES
# ---------------------------------------------------------
@app.route('/credit_note')
def credit_note_page():
    return render_template('credit_note.html')

@app.route('/api/credit_note/fetch', methods=['POST'])
def api_fetch_credit_note():
    try:
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        data = credit_note_module.get_all_credit_note_data(from_date, to_date, limit)
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch credit notes from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/credit_note/upload', methods=['POST'])
def api_upload_credit_note():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400
        
        import tempfile, os
        from datetime import datetime
        with tempfile.NamedTemporaryFile(delete=False, suffix='.json') as temp:
            file.save(temp.name)
            temp_path = temp.name
            
        parsed_credit_notes = credit_note_module.parse_tally_json(temp_path)
        
        # Save to SQLite
        if database_manager and parsed_credit_notes:
            database_manager.init_db()
            db_data_list = []
            for credit_note in parsed_credit_notes:
                db_data = {
                    "credit_note_number": credit_note.get("credit_note_number", ""),
                    "voucher_type": credit_note.get("voucher_type", "Credit Note"),
                    "date": credit_note.get("date", ""),
                    "from_account": credit_note.get("from_account", ""),
                    "to_account": credit_note.get("to_account", ""),
                    "amount": credit_note.get("amount", 0) or 0,
                    "narration": credit_note.get("narration", ""),
                    "ledger_entries": json.dumps(credit_note.get("ledger_entries", [])),
                    "line_items": json.dumps(credit_note.get("line_items", [])),
                    "cost_center_allocations": json.dumps(credit_note.get("cost_center_allocations", [])),
                    "tally_guid": credit_note.get("tally_guid", ""),
                    "company_name": "",
                    "created_at": datetime.now().isoformat(),
                    "updated_at": datetime.now().isoformat()
                }
                db_data_list.append(db_data)
                
            try:
                database_manager.bulk_save_credit_notes(db_data_list)
            except AttributeError:
                pass
                
        os.unlink(temp_path)
        
        total_amount = sum(float(c.get("amount", 0)) for c in parsed_credit_notes)
        return jsonify({"credit_notes": parsed_credit_notes, "total_amount": total_amount})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/credit_note/sync_zoho', methods=['POST'])
def api_sync_credit_note():
    try:
        selected = request.json.get("credit_notes") if request.is_json else None
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        result = credit_note_module.sync_credit_note_to_zoho(selected, from_date, to_date, limit)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/db/credit_notes', methods=['GET'])
def api_db_credit_notes():
    if not database_manager: return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        c_rows = database_manager.get_all_credit_notes()
        credit_notes = []
        for r in c_rows:
            d = dict(r)
            if isinstance(d.get('ledger_entries'), str):
                try: d['ledger_entries'] = json.loads(d['ledger_entries'])
                except: d['ledger_entries'] = []
            if isinstance(d.get('line_items'), str):
                try: d['line_items'] = json.loads(d['line_items'])
                except: d['line_items'] = []
            if isinstance(d.get('cost_center_allocations'), str):
                try: d['cost_center_allocations'] = json.loads(d['cost_center_allocations'])
                except: d['cost_center_allocations'] = []
            credit_notes.append(d)
                
        total_amount = sum(float(r.get('amount', 0) or 0) for r in credit_notes)
        return jsonify({
            "credit_notes": credit_notes,
            "count": len(credit_notes),
            "total_amount": total_amount
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ---------------------------------------------------------
# DEBIT NOTE ROUTES
# ---------------------------------------------------------
@app.route('/debit_note')
def debit_note_page():
    return render_template('debit_note.html')

@app.route('/api/debit_note/fetch', methods=['POST'])
def api_fetch_debit_note():
    try:
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        data = debit_note_module.get_all_debit_note_data(from_date, to_date, limit)
        if data:
            return jsonify(data)
        return jsonify({"error": "Failed to fetch debit notes from Tally"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/debit_note/upload', methods=['POST'])
def api_upload_debit_note():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400
        
        import tempfile, os
        from datetime import datetime
        with tempfile.NamedTemporaryFile(delete=False, suffix='.json') as temp:
            file.save(temp.name)
            temp_path = temp.name
            
        parsed_debit_notes = debit_note_module.parse_tally_json(temp_path)
        
        # Save to SQLite
        if database_manager and parsed_debit_notes:
            database_manager.init_db()
            db_data_list = []
            for debit_note in parsed_debit_notes:
                db_data = {
                    "debit_note_number": debit_note.get("debit_note_number", ""),
                    "voucher_type": debit_note.get("voucher_type", "Debit Note"),
                    "date": debit_note.get("date", ""),
                    "from_account": debit_note.get("from_account", ""),
                    "to_account": debit_note.get("to_account", ""),
                    "amount": debit_note.get("amount", 0) or 0,
                    "narration": debit_note.get("narration", ""),
                    "ledger_entries": json.dumps(debit_note.get("ledger_entries", [])),
                    "line_items": json.dumps(debit_note.get("line_items", [])),
                    "cost_center_allocations": json.dumps(debit_note.get("cost_center_allocations", [])),
                    "tally_guid": debit_note.get("tally_guid", ""),
                    "company_name": "",
                    "created_at": datetime.now().isoformat(),
                    "updated_at": datetime.now().isoformat()
                }
                db_data_list.append(db_data)
                
            try:
                database_manager.bulk_save_debit_notes(db_data_list)
            except AttributeError:
                pass
                
        os.unlink(temp_path)
        
        total_amount = sum(float(c.get("amount", 0)) for c in parsed_debit_notes)
        return jsonify({"debit_notes": parsed_debit_notes, "total_amount": total_amount})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/debit_note/sync_zoho', methods=['POST'])
def api_sync_debit_note():
    try:
        selected = request.json.get("debit_notes") if request.is_json else None
        from_date = request.json.get("from_date", "20250401") if request.is_json else "20250401"
        to_date = request.json.get("to_date", "20250430") if request.is_json else "20250430"
        limit = request.json.get("limit") if request.is_json else None
        
        result = debit_note_module.sync_debit_note_to_zoho(selected, from_date, to_date, limit)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/db/debit_notes', methods=['GET'])
def api_db_debit_notes():
    if not database_manager: return jsonify({"error": "DB Manager not loaded"}), 500
    try:
        c_rows = database_manager.get_all_debit_notes()
        debit_notes = []
        for r in c_rows:
            d = dict(r)
            if isinstance(d.get('ledger_entries'), str):
                try: d['ledger_entries'] = json.loads(d['ledger_entries'])
                except: d['ledger_entries'] = []
            if isinstance(d.get('line_items'), str):
                try: d['line_items'] = json.loads(d['line_items'])
                except: d['line_items'] = []
            if isinstance(d.get('cost_center_allocations'), str):
                try: d['cost_center_allocations'] = json.loads(d['cost_center_allocations'])
                except: d['cost_center_allocations'] = []
            debit_notes.append(d)
                
        total_amount = sum(float(r.get('amount', 0) or 0) for r in debit_notes)
        return jsonify({
            "debit_notes": debit_notes,
            "count": len(debit_notes),
            "total_amount": total_amount
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/payments/split_expenses', methods=['POST'])
def split_payments_expenses():
    if not database_manager:
        from flask import jsonify
        return jsonify({"status": "error", "error": "Database manager not available"}), 500
        
    try:
        from flask import request, jsonify
        import json as _json

        data = request.json or {}
        payments = data.get('payments', [])
        
        conn = database_manager.get_db_connection()
        cursor = conn.cursor()
        
        # Use SQL directly with COLLATE NOCASE to avoid Python string comparison issues
        # Mark as Vendor Payment: vendor_name exists in ledgers with type='vendor'
        cursor.execute("""
            UPDATE payments_made 
            SET voucher_type = 'Payment'
            WHERE vendor_name IN (
                SELECT pm.vendor_name FROM payments_made pm
                INNER JOIN ledgers l ON LOWER(TRIM(pm.vendor_name)) = LOWER(TRIM(l.name)) COLLATE NOCASE
                WHERE LOWER(l.type) = 'vendor'
            )
        """)
        vendor_updated = cursor.rowcount

        # Mark as Expense: vendor_name NOT in ledgers with type='vendor'
        cursor.execute("""
            UPDATE payments_made 
            SET voucher_type = 'Expense'
            WHERE vendor_name NOT IN (
                SELECT pm.vendor_name FROM payments_made pm
                INNER JOIN ledgers l ON LOWER(TRIM(pm.vendor_name)) = LOWER(TRIM(l.name)) COLLATE NOCASE
                WHERE LOWER(l.type) = 'vendor'
            )
        """)
        expense_updated = cursor.rowcount
        conn.commit()
        print(f" DB Update: {vendor_updated} vendor payments, {expense_updated} expenses")

        # Re-read directly from DB after update - DB is the truth source
        all_updated = cursor.execute(
            "SELECT * FROM payments_made ORDER BY date DESC"
        ).fetchall()
        
        vendor_payments = []
        expenses = []
        for row in all_updated:
            r = dict(row)
            # SQLite stores JSON arrays as strings, parse back to lists for frontend
            for field in ['bill_allocations', 'ledger_entries', 'cost_center_allocations']:
                if r.get(field):
                    try:
                        r[field] = _json.loads(r[field])
                    except:
                        r[field] = []
                else:
                    r[field] = []
                    
            if r.get('voucher_type', '').lower() == 'payment':
                vendor_payments.append(r)
            else:
                expenses.append(r)
                
        conn.close()
        return jsonify({
            "status": "success",
            "vendor_payments": vendor_payments,
            "expenses": expenses,
            "message": f"Split complete: {len(vendor_payments)} vendor payments, {len(expenses)} expenses"
        })
    except Exception as e:
        import traceback
        from flask import jsonify
        traceback.print_exc()
        return jsonify({"status": "error", "error": str(e)}), 500



# ─────────────────────────────────────────────────────────────────────────────
# OFFLINE MIGRATION TOOL — JSON Converter
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/json-converter')
def json_converter_page():
    return render_template('json_converter.html')


@app.route('/api/converter/parse-json', methods=['POST'])
def api_converter_parse_json():
    """
    Accepts:
      - json_file  (multipart) — Tally exported JSON
      - excel_file (multipart, optional) — Zoho Books sample Excel/CSV

    Returns JSON with:
      detected_type, records, raw_fields, count, errors, sample_cols
    """
    try:
        import importlib.util as _ilu
        _conv_path = os.path.join(os.path.dirname(__file__), 'modules', 'json_to_zoho_converter.py')
        _conv_spec = _ilu.spec_from_file_location('json_to_zoho_converter', _conv_path)
        conv = _ilu.module_from_spec(_conv_spec)
        _conv_spec.loader.exec_module(conv)

        # ── JSON file ──────────────────────────────────────────────────
        jf = request.files.get('json_file')
        if not jf or not jf.filename:
            return jsonify({"error": "json_file is required"}), 400

        json_bytes = jf.read()

        # ── Parse Tally JSON ───────────────────────────────────────────
        result = conv.parse_tally_json(json_bytes)

        # ── Excel sample (optional) ────────────────────────────────────
        sample_cols = []
        ef = request.files.get('excel_file')
        if ef and ef.filename:
            ef_bytes = ef.read()
            ex_result = conv.parse_sample_excel(ef_bytes)
            sample_cols = ex_result.get('columns', [])
            if ex_result.get('error'):
                result['errors'] = result.get('errors', []) + [
                    f"Sample Excel warning: {ex_result['error']}"
                ]

        result['sample_cols'] = sample_cols

        # Limit records returned (preview only — mapping uses all)
        # We send all records (export needs them), but cap at 2000 to avoid huge JSON
        MAX_RECORDS = 2000
        if len(result['records']) > MAX_RECORDS:
            result['errors'] = result.get('errors', []) + [
                f"Showing first {MAX_RECORDS} of {len(result['records'])} records"
            ]
            result['records'] = result['records'][:MAX_RECORDS]

        return jsonify(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/converter/export', methods=['POST'])
def api_converter_export():
    """
    Accepts JSON body:
      {
        "records":       [...],          # list of flat dicts (from parse step)
        "field_mapping": {...},          # { zoho_col: tally_field }
        "export_format": "xlsx" | "csv"
      }

    Returns: binary file download
    """
    try:
        import io as _io
        import importlib.util as _ilu
        _conv_path = os.path.join(os.path.dirname(__file__), 'modules', 'json_to_zoho_converter.py')
        _conv_spec = _ilu.spec_from_file_location('json_to_zoho_converter', _conv_path)
        conv = _ilu.module_from_spec(_conv_spec)
        _conv_spec.loader.exec_module(conv)
        from flask import send_file

        body = request.get_json(force=True)
        if not body:
            return jsonify({"error": "Request body is required"}), 400

        records       = body.get('records', [])
        field_mapping = body.get('field_mapping', {})
        export_format = body.get('export_format', 'xlsx')

        if not records:
            return jsonify({"error": "No records to export"}), 400
        if not field_mapping:
            return jsonify({"error": "field_mapping is required"}), 400

        file_bytes, mimetype, filename = conv.build_export(
            records, field_mapping, export_format
        )

        return send_file(
            _io.BytesIO(file_bytes),
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    print(" Starting Tally Software Frontend...")
    print(" URL: http://localhost:5000")
    app.run(debug=True, port=5000)

