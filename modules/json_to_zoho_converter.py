"""
json_to_zoho_converter.py
─────────────────────────
Offline Migration Tool Backend
- Parses any Tally-exported JSON file
- Dynamically detects voucher type & fields
- Applies user-defined field mapping
- Exports Zoho Books-ready Excel (xlsx) or CSV

No live Tally / Zoho connection needed — works fully offline.
"""

import json
import io
import csv
import traceback
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


# ──────────────────────────────────────────────────────────────────────────────
# ZOHO BOOKS FIELD TEMPLATES  (canonical column names per module)
# ──────────────────────────────────────────────────────────────────────────────

ZOHO_FIELDS = {
    "invoices": [
        "InvoiceDate", "InvoiceNumber", "CustomerName", "PlaceOfSupply",
        "PaymentTerms", "DueDate", "ItemName", "ItemDescription",
        "Quantity", "Unit", "Price", "DiscountAmount", "TaxName", "TaxType",
        "TaxPercent", "SubTotal", "TotalTaxAmount", "Total",
        "RoundOff", "Notes", "Terms", "Reference",
        "ShippingCharge", "ShippingTax", "SalesOrderNumber",
        "Currency", "ExchangeRate",
    ],
    "bills": [
        "BillDate", "BillNumber", "VendorName", "PlaceOfSupply",
        "PaymentTerms", "DueDate", "ItemName", "ItemDescription",
        "Quantity", "Unit", "Price", "DiscountAmount", "TaxName", "TaxType",
        "TaxPercent", "SubTotal", "TotalTaxAmount", "Total",
        "RoundOff", "Notes", "Reference",
        "Currency", "ExchangeRate",
    ],
    "journals": [
        "JournalDate", "JournalNumber", "ReferenceNumber",
        "Notes", "AccountName", "DebitAmount", "CreditAmount",
        "TaxName", "TaxType", "TaxPercent",
    ],
    "payments_made": [
        "Date", "PaymentNumber", "VendorName", "PaymentMode",
        "Amount", "ReferenceNumber", "Notes", "Currency",
        "BankCharges",
    ],
    "receipts": [
        "Date", "ReceiptNumber", "CustomerName", "PaymentMode",
        "Amount", "ReferenceNumber", "Notes", "Currency",
        "BankCharges",
    ],
    "sales_orders": [
        "SalesOrderDate", "SalesOrderNumber", "CustomerName",
        "ShipmentDate", "PaymentTerms", "DeliveryMethod",
        "ItemName", "Quantity", "Unit", "Price", "DiscountAmount",
        "TaxName", "TaxPercent", "SubTotal", "Total", "Notes",
    ],
    "purchase_orders": [
        "PurchaseOrderDate", "PurchaseOrderNumber", "VendorName",
        "DeliveryDate", "PaymentTerms",
        "ItemName", "Quantity", "Unit", "Price", "DiscountAmount",
        "TaxName", "TaxPercent", "SubTotal", "Total", "Notes",
    ],
    "credit_notes": [
        "CreditNoteDate", "CreditNoteNumber", "CustomerName",
        "InvoiceNumber", "ItemName", "Quantity", "Unit",
        "Price", "TaxName", "TaxPercent", "Total", "Notes",
    ],
    "debit_notes": [
        "DebitNoteDate", "DebitNoteNumber", "VendorName",
        "BillNumber", "ItemName", "Quantity", "Unit",
        "Price", "TaxName", "TaxPercent", "Total", "Notes",
    ],
    "contacts": [
        "ContactName", "CompanyName", "ContactType",
        "EmailAddress", "Phone", "Website",
        "BillingAddress", "BillingCity", "BillingState",
        "BillingCountry", "BillingZip",
        "ShippingAddress", "ShippingCity", "ShippingState",
        "ShippingCountry", "ShippingZip",
        # Zoho Books (India) commonly expects these exact headers for GST fields:
        "GST Treatment", "GSTIN / UIN", "Business Legal Name", "Business Trade Name", "Place of Supply",
        "PAN", "Currency", "PaymentTerms",
        # Extra helpful master fields (not a standard Zoho template, but export supports it)
        "Under", "UnderParent", "RegistrationType",
    ],
    "items": [
        "ItemName", "SKU", "Unit", "SalesPrice", "SalesAccount",
        "PurchasePrice", "PurchaseAccount", "Tax", "Description",
    ],
    "generic": [
        "Date", "Number", "Party", "Amount",
        "Description", "Reference", "Notes",
    ],
}


# ──────────────────────────────────────────────────────────────────────────────
# TALLY JSON PARSER
# ──────────────────────────────────────────────────────────────────────────────

def _safe_float(v):
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return 0.0


def _tally_date(v):
    """Convert Tally date string (YYYYMMDD or DD-MM-YYYY) → DD/MM/YYYY"""
    if not v:
        return ""
    v = str(v).strip()
    if len(v) == 8 and v.isdigit():
        return f"{v[6:8]}/{v[4:6]}/{v[0:4]}"
    return v


def _flatten_dict(d, parent_key="", sep="_"):
    """Recursively flatten a nested dict."""
    items = {}
    if not isinstance(d, dict):
        return {parent_key: d}
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.update(_flatten_dict(v, new_key, sep))
        elif isinstance(v, list):
            items[new_key] = json.dumps(v)
        else:
            items[new_key] = v
    return items


def parse_tally_json(file_bytes_or_str):
    """
    Parse Tally-exported JSON OR XML (any structure).
    Returns:
        {
          "detected_type": str,   # e.g. "invoices", "journals", ...
          "records": [...],        # list of flat dicts (one per row)
          "raw_fields": [...],     # all field names found in records
          "count": int,
          "errors": [...],
        }
    """
    errors = []
    records = []
    detected_type = "generic"
    voucher_list = []
    context = {}
    raw = ""

    try:
        if isinstance(file_bytes_or_str, (bytes, bytearray)):
            decoded = False
            # Try multiple encodings for Tally exports (especially UTF-16 for Masters)
            for enc in ('utf-16', 'utf-8-sig', 'utf-8', 'latin-1'):
                try:
                    raw = file_bytes_or_str.decode(enc)
                    decoded = True
                    break
                except Exception:
                    continue
            if not decoded:
                raise ValueError("Could not decode file with any encoding.")
        else:
            raw = str(file_bytes_or_str)

        # ── Detect XML vs JSON ──
        if raw.lstrip().startswith('<'): # It's an XML file
            voucher_list, context = _parse_tally_xml(raw, errors)
        else:
            data = json.loads(raw)
            voucher_list = _extract_voucher_list(data)

    except Exception as e:
        return {
            "detected_type": "generic",
            "records": [],
            "raw_fields": [],
            "count": 0,
            "errors": [f"File parse error: {e}"],
        }

    if not voucher_list:
        errors.append("Could not find any vouchers/ledgers in file.")

    # ── Detect type from voucher_type / first keys ──
    detected_type = _detect_type(voucher_list)

    # ── Flatten every voucher into a plain row ──
    for v in voucher_list:
        try:
            row = _flatten_voucher(v, detected_type, context=context)
            records.append(row)
        except Exception as e:
            errors.append(f"Row parse error: {e}")

    # ── Collect all field names ──
    raw_fields = []
    seen = set()
    for r in records:
        for k in r.keys():
            if k not in seen:
                raw_fields.append(k)
                seen.add(k)

    return {
        "detected_type": detected_type,
        "records": records,
        "raw_fields": raw_fields,
        "count": len(records),
        "errors": errors,
        "context": context,
    }


def _extract_voucher_list(data):
    """Try common Tally JSON structures to find the list of vouchers."""
    if isinstance(data, list):
        return data

    # Common Tally export wrappers
    for key in [
        "VOUCHER", "LEDGER", "STOCKITEM", "STOCKGROUP",
        "vouchers", "ledgers", "items", "data", "records",
        "TALLYMESSAGE", "ENVELOPE",
    ]:
        if isinstance(data, dict) and key in data:
            v = data[key]
            if isinstance(v, list):
                return v
            if isinstance(v, dict):
                # Recurse one level
                result = _extract_voucher_list(v)
                if result:
                    return result

    # Deep search for first list with >0 dict items
    if isinstance(data, dict):
        for v in data.values():
            if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                return v
            if isinstance(v, dict):
                result = _extract_voucher_list(v)
                if result:
                    return result

    return []


def _parse_tally_xml(raw_xml_string, errors):
    """Safely parse Tally XML file into flat dictionary items + context."""
    import xml.etree.ElementTree as ET
    import re
    
    # Clean invalid XML control chars that Tally exports (e.g. &#x4;)
    clean = re.sub(r'&#[xX]?[0-9a-fA-F]+;', '', raw_xml_string)
    clean = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', clean)
    
    try:
        root = ET.fromstring(clean)
    except Exception as e:
        errors.append(f"XML parse failed: {e}")
        return []

    def flatten_elem(elem):
        d = {}
        if elem.attrib:
            for k, v in elem.attrib.items():
                d[k.lower()] = v
        for child in elem:
            # Normalize namespaces: "{uri}TAG" -> "tag"
            k = child.tag
            if isinstance(k, str) and '}' in k:
                k = k.split('}', 1)[1]
            k = str(k).lower()
            if len(child) > 0:
                sub = flatten_elem(child)
                for sk, sv in sub.items():
                    d[f"{k}_{sk}"] = sv
            else:
                val = child.text.strip() if child.text else ''
                if k in d:
                    if isinstance(d[k], list): d[k].append(val)
                    else: d[k] = [d[k], val]
                else:
                    d[k] = val
        return d

    def norm_tag(tag):
        if not isinstance(tag, str):
            return ""
        if '}' in tag:
            tag = tag.split('}', 1)[1]
        return tag.upper()

    # Build GROUP hierarchy map (needed to resolve: Ledger -> Under -> UnderParent)
    group_parent_map = {}
    group_rows = []
    for elem in root.iter():
        if norm_tag(elem.tag) != "GROUP":
            continue
        flat = flatten_elem(elem)
        group_name = (flat.get("name") or "").strip()
        parent_name = (flat.get("parent") or "").strip()
        if group_name:
            group_parent_map[group_name] = parent_name
            group_rows.append({"name": group_name, "parent": parent_name})

    vouchers = []
    # Search for common element names, tolerant to namespaces and case.
    # Some Tally exports wrap nodes in different parents, and may include namespaces (esp. UDF fields).
    wanted = {"VOUCHER": None, "LEDGER": "Ledger", "STOCKITEM": "StockItem"}

    for elem in root.iter():
        tag_upper = norm_tag(elem.tag)
        if tag_upper not in wanted:
            continue

        flat = flatten_elem(elem)
        meta = wanted[tag_upper]
        if meta:
            flat['metadata_type'] = meta
        vouchers.append(flat)

    return vouchers, {"group_parent_map": group_parent_map, "groups": group_rows}


def _detect_type(voucher_list):
    """Heuristically detect voucher type from the data."""
    if not voucher_list:
        return "generic"

    # Try first non-empty item
    sample = {}
    for item in voucher_list[:5]:
        if isinstance(item, dict):
            sample = item
            break

    keys_lower = {str(k).lower() for k in sample.keys()}
    vals_lower = {str(v).lower() for v in sample.values() if isinstance(v, str)}
    all_text = keys_lower | vals_lower

    def has(words):
        return any(w in all_text for w in words)

    # Check VOUCHERTYPENAME value
    vtype = str(sample.get("VOUCHERTYPENAME", sample.get("voucher_type", ""))).lower()

    if any(x in vtype for x in ["sales", "invoice"]):  return "invoices"
    if any(x in vtype for x in ["purchase", "bill"]):  return "bills"
    if any(x in vtype for x in ["journal", "jrnl"]):   return "journals"
    if any(x in vtype for x in ["payment"]):           return "payments_made"
    if any(x in vtype for x in ["receipt"]):           return "receipts"
    if any(x in vtype for x in ["credit"]):            return "credit_notes"
    if any(x in vtype for x in ["debit"]):             return "debit_notes"
    if any(x in vtype for x in ["contra"]):            return "journals"
    if any(x in vtype for x in ["sales order"]):       return "sales_orders"
    if any(x in vtype for x in ["purchase order"]):    return "purchase_orders"

    # Check for Master data (metadata > type: Ledger)
    meta = sample.get("metadata", {})
    if isinstance(meta, dict) and str(meta.get("type", "")).lower() == "ledger":
        return "contacts"
    if isinstance(meta, dict) and str(meta.get("type", "")).lower() == "stockitem":
        return "items"

    # Also check if it's generic Tally master
    if "metadata_type" in all_text and "ledger" in all_text:
        return "contacts"

    # Key-based heuristics
    if has(["invoice_number", "invoicenumber", "customer_name"]):  return "invoices"
    if has(["bill_number", "billnumber", "vendor_name"]):          return "bills"
    if has(["debit", "credit", "ledger_entry"]):                   return "journals"
    if has(["payment_number", "payment_mode"]):                    return "payments_made"
    if has(["receipt_number"]):                                    return "receipts"
    if has(["itemname", "item_name", "stockitem", "hsn", "stockgroup"]): return "items"
    if has(["ledger_name", "ledgername", "openingbalance", "parent"]):   return "contacts"

    return "generic"


def _flatten_voucher(v, detected_type, context=None):
    """Convert a single Tally voucher dict into a flat row dict."""
    if not isinstance(v, dict):
        return {"value": str(v)}

    row = {}

    # ── Common date / number fields ──
    for date_key in ["DATE", "date", "Date", "VOUCHERDATE"]:
        if date_key in v:
            row["Date"] = _tally_date(v[date_key])
            break

    for num_key in ["VOUCHERNUMBER", "number", "Number", "voucher_number"]:
        if num_key in v:
            row["Number"] = str(v[num_key])
            break

    for party_key in ["PARTYLEDGERNAME", "party_name", "PartyName", "customer_name", "vendor_name"]:
        if party_key in v:
            row["Party"] = str(v[party_key])
            break

    # ── Type-specific mappings ──
    if detected_type == "invoices":
        row.update(_map_invoice(v))
    elif detected_type == "bills":
        row.update(_map_bill(v))
    elif detected_type == "journals":
        row.update(_map_journal(v))
    elif detected_type in ("payments_made", "receipts"):
        row.update(_map_payment_receipt(v))
    elif detected_type == "credit_notes":
        row.update(_map_credit_note(v))
    elif detected_type == "debit_notes":
        row.update(_map_debit_note(v))
    elif detected_type == "items":
        row.update(_map_item(v))
    elif detected_type == "contacts":
        row.update(_map_contact(v, context=context))

    # ── Always include ALL original flat fields ──
    for k, val in v.items():
        flat_key = str(k)
        if flat_key not in row:
            if isinstance(val, (dict, list)):
                row[f"_raw_{flat_key}"] = json.dumps(val, ensure_ascii=False)
            else:
                row[flat_key] = val

    return row


# ── Type-specific mappers ──

def _map_invoice(v):
    row = {}
    row["InvoiceNumber"]  = v.get("VOUCHERNUMBER") or v.get("invoice_number", "")
    row["InvoiceDate"]    = _tally_date(v.get("DATE") or v.get("date", ""))
    row["CustomerName"]   = v.get("PARTYLEDGERNAME") or v.get("customer_name", "")
    row["Reference"]      = v.get("REFERENCE") or v.get("reference", "")
    row["Notes"]          = v.get("NARRATION") or v.get("narration", "")
    row["PlaceOfSupply"]  = v.get("PLACEOFSUPPLY") or v.get("place_of_supply", "")

    # Try to get total from AMOUNT or nested
    amt = v.get("AMOUNT") or v.get("total_amount") or v.get("amount", 0)
    row["Total"] = abs(_safe_float(amt))

    tax = v.get("TAX") or v.get("tax_total") or 0
    row["TotalTaxAmount"] = abs(_safe_float(tax))

    # Line items
    items = v.get("ALLINVENTORYENTRIES") or v.get("INVENTORYENTRIES") or v.get("line_items", [])
    if isinstance(items, list) and items:
        first = items[0] if isinstance(items[0], dict) else {}
        row["ItemName"]   = first.get("STOCKITEMNAME") or first.get("item_name", "")
        row["Quantity"]   = first.get("ACTUALQTY") or first.get("quantity", "")
        row["Unit"]       = first.get("UNIT") or first.get("unit", "")
        row["Price"]      = abs(_safe_float(first.get("RATE") or first.get("rate", 0)))
    return row


def _map_bill(v):
    row = {}
    row["BillNumber"]  = v.get("VOUCHERNUMBER") or v.get("bill_number", "")
    row["BillDate"]    = _tally_date(v.get("DATE") or v.get("date", ""))
    row["VendorName"]  = v.get("PARTYLEDGERNAME") or v.get("vendor_name", "")
    row["Reference"]   = v.get("REFERENCE") or v.get("reference", "")
    row["Notes"]       = v.get("NARRATION") or v.get("narration", "")
    amt = v.get("AMOUNT") or v.get("total_amount") or 0
    row["Total"] = abs(_safe_float(amt))
    items = v.get("ALLINVENTORYENTRIES") or v.get("INVENTORYENTRIES") or v.get("line_items", [])
    if isinstance(items, list) and items:
        first = items[0] if isinstance(items[0], dict) else {}
        row["ItemName"] = first.get("STOCKITEMNAME") or first.get("item_name", "")
        row["Quantity"] = first.get("ACTUALQTY") or first.get("quantity", "")
        row["Unit"]     = first.get("UNIT") or first.get("unit", "")
        row["Price"]    = abs(_safe_float(first.get("RATE") or first.get("rate", 0)))
    return row


def _map_journal(v):
    row = {}
    row["JournalDate"]   = _tally_date(v.get("DATE") or v.get("date", ""))
    row["JournalNumber"] = v.get("VOUCHERNUMBER") or v.get("voucher_number", "")
    row["Notes"]         = v.get("NARRATION") or v.get("narration", "")
    row["ReferenceNumber"] = v.get("REFERENCE") or v.get("reference", "")
    entries = v.get("ALLLEDGERENTRIES") or v.get("ledger_entries", [])
    if isinstance(entries, list) and entries:
        e0 = entries[0] if isinstance(entries[0], dict) else {}
        row["AccountName"]   = e0.get("LEDGERNAME") or e0.get("ledger_name", "")
        amt = _safe_float(e0.get("AMOUNT") or e0.get("amount", 0))
        row["DebitAmount"]  = abs(amt) if amt < 0 else 0
        row["CreditAmount"] = amt if amt > 0 else 0
    return row


def _map_payment_receipt(v):
    row = {}
    row["Date"]            = _tally_date(v.get("DATE") or v.get("date", ""))
    row["PaymentNumber"]   = v.get("VOUCHERNUMBER") or v.get("payment_number") or v.get("receipt_number", "")
    row["Party"]           = v.get("PARTYLEDGERNAME") or v.get("vendor_name") or v.get("customer_name", "")
    amt = v.get("AMOUNT") or v.get("amount", 0)
    row["Amount"]          = abs(_safe_float(amt))
    row["Notes"]           = v.get("NARRATION") or v.get("narration", "")
    row["ReferenceNumber"] = v.get("REFERENCE") or v.get("reference_number", "")
    row["PaymentMode"]     = v.get("payment_mode", "")
    return row


def _map_credit_note(v):
    row = {}
    row["CreditNoteDate"]   = _tally_date(v.get("DATE") or v.get("date", ""))
    row["CreditNoteNumber"] = v.get("VOUCHERNUMBER") or v.get("credit_note_number", "")
    row["CustomerName"]     = v.get("PARTYLEDGERNAME") or v.get("customer_name", "")
    amt = v.get("AMOUNT") or v.get("total_amount", 0)
    row["Total"] = abs(_safe_float(amt))
    row["Notes"] = v.get("NARRATION") or v.get("narration", "")
    return row


def _map_debit_note(v):
    row = {}
    row["DebitNoteDate"]   = _tally_date(v.get("DATE") or v.get("date", ""))
    row["DebitNoteNumber"] = v.get("VOUCHERNUMBER") or v.get("debit_note_number", "")
    row["VendorName"]      = v.get("PARTYLEDGERNAME") or v.get("vendor_name", "")
    amt = v.get("AMOUNT") or v.get("total_amount", 0)
    row["Total"] = abs(_safe_float(amt))
    row["Notes"] = v.get("NARRATION") or v.get("narration", "")
    return row


def _map_item(v):
    row = {}
    row["ItemName"]  = v.get("NAME") or v.get("item_name", "")
    row["Unit"]      = v.get("BASEUNITS") or v.get("unit", "")
    row["SalesPrice"] = abs(_safe_float(v.get("BATCHALLOCATIONS_RATE") or v.get("rate", 0)))
    row["Description"] = v.get("DESCRIPTION") or v.get("description", "")
    return row


def _map_contact(v, context=None):
    row = {}
    row["ContactName"] = v.get("NAME") or v.get("name") or v.get("ledger_name", "")
    row["GSTIN"]       = (
        v.get("PARTYGSTIN") or v.get("partygstin") or
        v.get("LEDGSTREGDETAILS.LIST_GSTIN") or v.get("ledgstregdetails.list_gstin") or
        v.get("GSTIN") or v.get("gstin") or ""
    )
    row["PAN"]         = v.get("INCOMETAXNUMBER") or v.get("incometaxnumber") or v.get("pan", "")
    row["CompanyName"] = row["ContactName"]

    def _first_non_empty(*keys):
        for k in keys:
            val = v.get(k)
            if val is None:
                continue
            if isinstance(val, list):
                val = [str(x).strip() for x in val if str(x).strip()]
                if val:
                    return val
            else:
                s = str(val).strip()
                if s:
                    return s
        return ""

    # Address: Tally masters commonly store it under LEDMAILINGDETAILS.LIST -> ADDRESS.LIST -> ADDRESS
    address = _first_non_empty(
        "LEDMAILINGDETAILS.LIST_ADDRESS.LIST_ADDRESS",
        "ledmailingdetails.list_address.list_address",
        "ADDRESS.LIST_ADDRESS",
        "address.list_address",
        "OLDADDRESS.LIST_OLDADDRESS",
        "oldaddress.list_oldaddress",
        "address",
    )
    if isinstance(address, list):
        address_text = ", ".join(address)
    else:
        address_text = address
    row["BillingAddress"] = address_text

    # State/Country/Pincode (Billing*)
    row["BillingState"] = _first_non_empty(
        "LEDMAILINGDETAILS.LIST_STATE",
        "ledmailingdetails.list_state",
        "PRIORSTATENAME",
        "priorstatename",
        "OLDLEDSTATENAME",
        "oldledstatename",
        "STATE",
        "state",
    )
    row["BillingCountry"] = _first_non_empty(
        "LEDMAILINGDETAILS.LIST_COUNTRY",
        "ledmailingdetails.list_country",
        "OLDCOUNTRYNAME",
        "oldcountryname",
        "COUNTRY",
        "country",
    )
    zip_code = _first_non_empty(
        "PINCODE",
        "pincode",
        "LEDMAILINGDETAILS.LIST_PINCODE",
        "ledmailingdetails.list_pincode",
    )
    if not zip_code and address_text:
        import re as _re
        m = _re.search(r'(\d{6})\s*$', str(address_text).strip())
        if m:
            zip_code = m.group(1)
    row["BillingZip"] = zip_code

    # Ledger "Under" (Group) + parent group
    under = _first_non_empty("PARENT", "parent")
    row["Under"] = under
    group_parent_map = (context or {}).get("group_parent_map", {}) if isinstance(context, dict) else {}
    row["UnderParent"] = group_parent_map.get(under, "") if under else ""

    # GST Registration type (helps audits/migration)
    row["RegistrationType"] = _first_non_empty(
        "LEDGSTREGDETAILS.LIST_GSTREGISTRATIONTYPE",
        "ledgstregdetails.list_gstregistrationtype",
        "GSTREGISTRATIONTYPE",
        "gstregistrationtype",
    )

    # Zoho Books (India) helps: GST Treatment + Business names
    # If GSTIN exists -> Registered Business - Regular; else keep blank.
    # (If Zoho expects different labels, upload Zoho sample Excel so the UI uses exact column names.)
    gstin_val = str(row.get("GSTIN") or "").strip()
    row["GSTTreatment"] = "Registered Business - Regular" if gstin_val else ""

    mailing_name = _first_non_empty(
        "LEDMAILINGDETAILS.LIST_MAILINGNAME",
        "ledmailingdetails.list_mailingname",
        "OLDMAILINGNAME",
        "oldmailingname.list_oldmailingname",
        "MAILINGNAME",
        "mailingname",
    )
    biz_name = mailing_name or row["ContactName"]
    row["BusinessLegalName"] = biz_name
    row["BusinessTradeName"] = biz_name

    # Place of Supply (Zoho Books India) – use GST reg details place/state if present, else mailing state.
    row["PlaceOfSupply"] = _first_non_empty(
        "LEDGSTREGDETAILS.LIST_PLACEOFSUPPLY",
        "ledgstregdetails.list_placeofsupply",
        "LEDGSTREGDETAILS.LIST_STATE",
        "ledgstregdetails.list_state",
        "LEDMAILINGDETAILS.LIST_STATE",
        "ledmailingdetails.list_state",
        "STATE",
        "state",
    )
    
    phone = v.get("LEDGERMOBILE") or v.get("ledgermobile") or v.get("phonenumber") or v.get("phone", "")
    if isinstance(phone, list): phone = phone[0]
    row["Phone"] = phone
    
    email = v.get("EMAIL") or v.get("email") or v.get("emailid", "")
    row["EmailAddress"] = email

    # Usually in Tally, Debtors are Customers, Creditors are Vendors
    parent = str(v.get("PARENT") or v.get("parent", "")).lower()
    if "debtor" in parent:
        row["ContactType"] = "Customer"
    elif "creditor" in parent:
        row["ContactType"] = "Vendor"
    else:
        row["ContactType"] = "Other Account"
    return row


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL SAMPLE PARSER  (read Zoho Books sample export to detect columns)
# ──────────────────────────────────────────────────────────────────────────────

def parse_sample_excel(file_bytes):
    """
    Read a Zoho Books sample Excel/CSV export.
    Returns list of column names found in the first header row.
    """
    if not HAS_PANDAS:
        return {"columns": [], "error": "pandas not installed"}
    try:
        buf = io.BytesIO(file_bytes)
        try:
            df = pd.read_excel(buf, nrows=3)
        except Exception:
            buf.seek(0)
            df = pd.read_csv(buf, nrows=3)
        columns = [str(c).strip() for c in df.columns.tolist() if str(c).strip()]
        sample_rows = df.head(2).fillna("").to_dict(orient="records")
        return {"columns": columns, "sample_rows": sample_rows, "error": None}
    except Exception as e:
        return {"columns": [], "error": str(e)}


# ──────────────────────────────────────────────────────────────────────────────
# EXPORT BUILDER
# ──────────────────────────────────────────────────────────────────────────────

def build_export(records, field_mapping, export_format="xlsx"):
    """
    Apply field_mapping to records and export.

    field_mapping: dict  { zoho_column_name: tally_field_name_or_literal }
       e.g. { "InvoiceNumber": "VOUCHERNUMBER", "CustomerName": "PARTYLEDGERNAME", ... }

    Returns: (bytes, mimetype, filename)
    """
    rows = []
    zoho_columns = list(field_mapping.keys())

    for rec in records:
        row = {}
        for zoho_col, tally_field in field_mapping.items():
            if not tally_field:
                row[zoho_col] = ""
            elif tally_field.startswith("__literal__:"):
                row[zoho_col] = tally_field[len("__literal__:"):]
            else:
                row[zoho_col] = rec.get(tally_field, "")
        rows.append(row)

    if export_format == "csv":
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=zoho_columns)
        writer.writeheader()
        writer.writerows(rows)
        content = buf.getvalue().encode("utf-8-sig")
        return content, "text/csv", "zoho_import.csv"

    # xlsx
    if not HAS_OPENPYXL:
        # Fall back to CSV
        return build_export(records, field_mapping, "csv")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zoho Import"

    # Header style
    header_fill  = PatternFill("solid", fgColor="1A73E8")
    header_font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 28

    for ci, col_name in enumerate(zoho_columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = max(16, len(col_name) + 4)

    # Data rows
    alt_fill = PatternFill("solid", fgColor="F8FAFE")
    data_font = Font(name="Calibri", size=10)
    for ri, row in enumerate(rows, start=2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, col_name in enumerate(zoho_columns, start=1):
            cell = ws.cell(row=ri, column=ci, value=str(row.get(col_name, "")))
            cell.font   = data_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
        ws.row_dimensions[ri].height = 18

    # Freeze header
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return (
        buf.read(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "zoho_import.xlsx",
    )


# ──────────────────────────────────────────────────────────────────────────────
# AUTO-MAPPING SUGGESTIONS
# ──────────────────────────────────────────────────────────────────────────────

_COMMON_ALIASES = {
    # Zoho col      → possible Tally keys (lowercase)
    "invoicenumber":   ["vouchernumber", "invoice_number", "number"],
    "invoicedate":     ["date", "voucherdate", "invoice_date"],
    "billnumber":      ["vouchernumber", "bill_number", "number"],
    "billdate":        ["date", "voucherdate", "bill_date"],
    "journaldate":     ["date", "voucherdate", "journal_date"],
    "journalnumber":   ["vouchernumber", "journal_number", "number"],
    "customername":    ["partyledgername", "customer_name", "party", "party_name"],
    "vendorname":      ["partyledgername", "vendor_name", "party", "party_name"],
    "total":           ["amount", "total_amount", "grand_total"],
    "totaltaxamount":  ["tax_total", "tax_amount", "totaltaxamount"],
    "notes":           ["narration", "narrations", "notes"],
    "reference":       ["reference", "reference_number", "referencenumber"],
    "itemname":        ["name", "stockitemname", "item_name", "itemname"],
    "quantity":        ["actualqty", "quantity", "qty"],
    "unit":            ["unit", "baseunits"],
    "price":           ["rate", "price", "salesrate"],
    "amount":          ["amount", "total_amount", "grand_total"],
    "paymentmode":     ["payment_mode", "paymentmode"],
    "paymentnumber":   ["vouchernumber", "payment_number"],
    "date":            ["date", "voucherdate"],
    "accountname":     ["ledgername", "ledger_name", "account_name"],
    "debitamount":     ["debit_amount", "debit"],
    "creditamount":    ["credit_amount", "credit"],
    "placeofsupply":   ["placeofsupply", "place_of_supply"],
    "gstin":           ["partygstin", "gstin"],
    "contactname":     ["name", "ledger_name", "contact_name"],
    "openingbalance":  ["openingbalance", "opening_balance", "balance"],
    "parent":          ["parent", "group", "under"],
}


def suggest_mapping(raw_fields, zoho_fields_list):
    """
    Returns a dict { zoho_field: best_tally_field_or_empty }
    using simple alias matching.
    """
    raw_lower = {f.lower(): f for f in raw_fields}
    mapping = {}
    for zf in zoho_fields_list:
        aliases = _COMMON_ALIASES.get(zf.lower(), [zf.lower()])
        found = ""
        for alias in aliases:
            if alias in raw_lower:
                found = raw_lower[alias]
                break
        mapping[zf] = found
    return mapping
