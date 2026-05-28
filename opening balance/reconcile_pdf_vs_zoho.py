"""
PDF Trial Balance (Tally) vs Zoho Books Excel Report - Reconciliation Engine
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
"""
Generates a rich Excel reconciliation report with:
  - Sheet 1: Summary Overview
  - Sheet 2: Matched Accounts 
  - Sheet 3: Amount Mismatches
  - Sheet 4: Missing in Zoho (Tally-only)
  - Sheet 5: Extra in Zoho (Zoho-only)
  - Sheet 6: Raw Tally PDF Data
  - Sheet 7: Raw Zoho Data
"""

import json
import re
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import pdfplumber
from datetime import datetime

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
PDF_PATH   = 'new_TrialBal.pdf'
EXCEL_PATH = 'zoho_report.xlsx'
OUT_PATH   = f'Reconciliation_Report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

MATCH_THRESHOLD = 0.60   # fuzzy match ratio

# ─────────────────────────────────────────────
# COLORS
# ─────────────────────────────────────────────
C_DARK_BG    = '0B0F1A'
C_HEADER_BG  = '111827'
C_GREEN      = '10B981'
C_GREEN_LITE = 'D1FAE5'
C_RED        = 'EF4444'
C_RED_LITE   = 'FEE2E2'
C_ORANGE     = 'F97316'
C_ORANGE_LITE= 'FFEDD5'
C_BLUE       = '3B82F6'
C_BLUE_LITE  = 'DBEAFE'
C_PURPLE     = '8B5CF6'
C_PURPLE_LITE= 'EDE9FE'
C_YELLOW     = 'EAB308'
C_YELLOW_LITE= 'FEF9C3'
C_WHITE      = 'FFFFFF'
C_LIGHT_BG   = 'F8FAFC'
C_MUTED      = '64748B'
C_BORDER     = 'E2E8F0'

# ─────────────────────────────────────────────
# PDF PARSING (same logic as parse_pdf_trial_balance.py)
# ─────────────────────────────────────────────
SKIP_TEXTS = {
    'Opening Balance', 'Debit', 'Credit',
    'Carried Over', 'Brought Forward', 'continued ...',
    'Trial Balance', 'Page', 'For 1-Apr-25',
}
DEBIT_X_MIN  = 390
DEBIT_X_MAX  = 480
CREDIT_X_MIN = 480
CREDIT_X_MAX = 600

def is_skip(text):
    t = text.strip()
    for s in SKIP_TEXTS:
        if s in t:
            return True
    if re.match(r'^Trial Balance.*Page \d+$', t): return True
    if re.match(r'^For \d', t): return True
    return False

def x0_to_level(x0):
    if x0 < 50:   return 0
    if x0 < 70:   return 1
    if x0 < 90:   return 2
    return 3

def parse_amount(text):
    t = text.strip().replace(',', '')
    try:    return float(t)
    except: return None

def parse_tally_pdf(pdf_path):
    rows = []
    company = ''
    period  = ''
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words(keep_blank_chars=True, x_tolerance=3, y_tolerance=3)
            if page_num == 0:
                for w in words[:5]:
                    if any(kw in w['text'] for kw in ['PVT LTD', 'LTD', 'INDIA', 'PRIVATE']):
                        company = w['text']
                    elif 'For ' in w['text']:
                        period = w['text']

            line_groups = {}
            for w in words:
                top = round(w['top'] / 3) * 3
                line_groups.setdefault(top, []).append(w)

            for top in sorted(line_groups):
                line_words = line_groups[top]
                name_words, debit_words, credit_words = [], [], []
                for w in line_words:
                    x0, x1 = w['x0'], w['x1']
                    if x0 >= DEBIT_X_MIN and x1 <= DEBIT_X_MAX + 5:
                        debit_words.append(w)
                    elif x0 >= CREDIT_X_MIN:
                        credit_words.append(w)
                    else:
                        name_words.append(w)

                if not name_words: continue
                name_text = ' '.join(w['text'] for w in name_words).strip()
                if is_skip(name_text) or not name_text or name_text == company: continue

                level  = x0_to_level(name_words[0]['x0'])
                debit  = parse_amount(' '.join(w['text'] for w in debit_words)) if debit_words else None
                credit = parse_amount(' '.join(w['text'] for w in credit_words)) if credit_words else None
                rows.append({'name': name_text, 'level': level, 'debit': debit or 0.0, 'credit': credit or 0.0})

    return company, period, rows

# ─────────────────────────────────────────────
# ZOHO EXCEL PARSING
# ─────────────────────────────────────────────
def parse_zoho_excel(excel_path):
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    rows = []
    company = ''
    period  = ''

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            if row[0]:
                # Extract company and period from header cell
                txt = str(row[0])
                lines = [l.strip() for l in txt.split('\n') if l.strip()]
                if lines: company = lines[0]
                for l in lines:
                    if 'As of' in l or 'as of' in l:
                        period = l
            continue
        if i == 1:  # header row
            continue
        if row[0] is None and row[1] is None:
            continue

        name = str(row[0]).strip() if row[0] else ''
        if not name or name in ('Account ', 'Total for Trial Balance', ''):
            continue
        # Strip leading spaces (Zoho uses spaces for indent)
        indent = len(name) - len(name.lstrip())
        level  = indent // 10  # approx level
        clean  = name.strip()

        # Skip "Total for X" rows
        if clean.lower().startswith('total for '):
            continue

        try:
            debit  = float(row[1]) if row[1] not in (None, '') else 0.0
            credit = float(row[2]) if row[2] not in (None, '') else 0.0
        except (ValueError, TypeError):
            debit, credit = 0.0, 0.0

        rows.append({'name': clean, 'level': level, 'debit': debit, 'credit': credit})

    return company, period, rows

# ─────────────────────────────────────────────
# NORMALISE NAME for matching
# ─────────────────────────────────────────────
def normalise(name):
    n = name.lower()
    n = re.sub(r'[^a-z0-9 ]', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    # Common abbreviation expansions
    n = n.replace(' a/c', ' account')
    n = n.replace(' &', ' and')
    return n

def simple_similarity(a, b):
    """Simple token overlap similarity"""
    ta = set(normalise(a).split())
    tb = set(normalise(b).split())
    if not ta or not tb: return 0.0
    intersection = ta & tb
    return len(intersection) / max(len(ta), len(tb))

# ─────────────────────────────────────────────
# RECONCILIATION ENGINE
# ─────────────────────────────────────────────
def reconcile(tally_rows, zoho_rows):
    """Match Tally rows to Zoho rows by name, return categorised results."""
    # Build Zoho lookup dict (normalised name -> row)
    zoho_lookup = {}
    for r in zoho_rows:
        key = normalise(r['name'])
        zoho_lookup[key] = r

    matched    = []
    mismatched = []
    only_tally = []
    zoho_used  = set()

    for tr in tally_rows:
        tkey = normalise(tr['name'])

        # Exact match first
        if tkey in zoho_lookup:
            zr = zoho_lookup[tkey]
            zoho_used.add(tkey)
        else:
            # Fuzzy match
            best_score = 0.0
            best_key   = None
            for zkey in zoho_lookup:
                if zkey in zoho_used: continue
                score = simple_similarity(tr['name'], zoho_lookup[zkey]['name'])
                if score > best_score:
                    best_score = score
                    best_key   = zkey
            if best_score >= MATCH_THRESHOLD and best_key:
                zr = zoho_lookup[best_key]
                zoho_used.add(best_key)
            else:
                only_tally.append(tr)
                continue

        # Compare amounts
        t_dr, t_cr = tr['debit'], tr['credit']
        z_dr, z_cr = zr['debit'], zr['credit']

        # Compute net difference
        t_net = t_dr - t_cr
        z_net = z_dr - z_cr
        diff  = abs(t_net - z_net)

        result = {
            'tally_name':  tr['name'],
            'zoho_name':   zr['name'],
            'tally_level': tr['level'],
            'tally_dr':    t_dr,
            'tally_cr':    t_cr,
            'zoho_dr':     z_dr,
            'zoho_cr':     z_cr,
            't_net':       t_net,
            'z_net':       z_net,
            'diff':        diff,
        }

        if diff < 1.0:
            matched.append(result)
        else:
            # Classify root cause
            if abs(abs(t_net) - abs(z_net)) < 1.0:
                result['root_cause'] = 'DR/CR Side Reversed'
            elif z_dr == 0 and z_cr == 0:
                result['root_cause'] = 'Zero in Zoho – Opening balance not set'
            elif t_dr == 0 and t_cr == 0:
                result['root_cause'] = 'Zero in Tally – Entry may be missing'
            else:
                result['root_cause'] = 'Partial amount mismatch'
            mismatched.append(result)

    # Zoho-only rows
    only_zoho = [r for k, r in zoho_lookup.items() if k not in zoho_used
                 and (r['debit'] != 0 or r['credit'] != 0)]

    return matched, mismatched, only_tally, only_zoho

# ─────────────────────────────────────────────
# EXCEL STYLE HELPERS
# ─────────────────────────────────────────────
def make_fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def make_font(bold=False, color='000000', size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name='Calibri')

def make_border(style='thin', color=C_BORDER):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def make_align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def write_header_row(ws, row_num, headers, bg_color=C_HEADER_BG, fg_color=C_WHITE, height=22):
    ws.row_dimensions[row_num].height = height
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=hdr)
        cell.fill = make_fill(bg_color)
        cell.font = make_font(bold=True, color=fg_color, size=9)
        cell.alignment = make_align('center')
        cell.border = make_border('medium', color='334155')

INR_FMT  = '#,##0.00'
DIFF_FMT = '#,##0.00'

def write_data_row(ws, row_num, values, bg_color=None, height=16):
    ws.row_dimensions[row_num].height = height
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        if bg_color:
            cell.fill = make_fill(bg_color)
        cell.border = make_border(color=C_BORDER)
        cell.alignment = make_align('left')
        if isinstance(val, float) or (isinstance(val, (int,)) and col > 1):
            cell.alignment = make_align('right')
            if isinstance(val, float):
                cell.number_format = INR_FMT
    return ws

def amt_cell(ws, row, col, val, color=None):
    c = ws.cell(row=row, column=col, value=val if val else 0.0)
    c.number_format = INR_FMT
    c.alignment = make_align('right')
    c.border = make_border(color=C_BORDER)
    if color:
        c.font = make_font(color=color)
    return c

# ─────────────────────────────────────────────
# SHEET BUILDERS
# ─────────────────────────────────────────────
def build_summary_sheet(wb, matched, mismatched, only_tally, only_zoho,
                        t_company, t_period, z_company, z_period):
    ws = wb.active
    ws.title = ' Summary'
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells('A1:G1')
    ws['A1'] = '  TALLY ↔ ZOHO BOOKS — TRIAL BALANCE RECONCILIATION REPORT'
    ws['A1'].font = make_font(bold=True, color=C_WHITE, size=16)
    ws['A1'].fill = make_fill(C_DARK_BG)
    ws['A1'].alignment = make_align('center', 'center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:D2')
    ws['A2'] = f'Tally Source: {t_company}  |  Period: {t_period}'
    ws['A2'].font = make_font(color='94A3B8', size=9)
    ws['A2'].fill = make_fill('0D1B35')
    ws['A2'].alignment = make_align('center', 'center')
    ws.row_dimensions[2].height = 18

    ws.merge_cells('E2:G2')
    ws['E2'] = f'Zoho Source: {z_company}  |  {z_period}'
    ws['E2'].font = make_font(color='94A3B8', size=9)
    ws['E2'].fill = make_fill('0D1B35')
    ws['E2'].alignment = make_align('center', 'center')

    ws.merge_cells('A3:G3')
    ws['A3'] = f'Generated: {datetime.now().strftime("%d %b %Y  %I:%M %p")}'
    ws['A3'].font = make_font(color=C_MUTED, size=8, italic=True)
    ws['A3'].fill = make_fill('111827')
    ws['A3'].alignment = make_align('right', 'center')
    ws.row_dimensions[3].height = 16

    # Spacer
    ws.row_dimensions[4].height = 10

    # KPI Cards (row 5–8)
    kpis = [
        ('A', ' Matched',       len(matched),    C_GREEN,  C_GREEN_LITE,  ' Amounts match in both systems'),
        ('B', ' Mismatched',    len(mismatched), C_RED,    C_RED_LITE,    ' Different amounts detected'),
        ('C', ' Missing in Zoho', len(only_tally), C_ORANGE, C_ORANGE_LITE, '→ In Tally but not in Zoho'),
        ('D', ' Extra in Zoho',  len(only_zoho),  C_BLUE,   C_BLUE_LITE,   '→ In Zoho but not in Tally'),
    ]
    for col_letter, title, count, accent, lite, desc in kpis:
        col_idx = ord(col_letter) - ord('A') + 1
        for r in range(5, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = make_fill(lite)
            ws.row_dimensions[r].height = 20

        ws.cell(row=5, column=col_idx, value=title).font = make_font(bold=True, color=accent, size=9)
        ws.cell(row=5, column=col_idx).fill = make_fill(lite)
        ws.cell(row=5, column=col_idx).alignment = make_align('center')

        count_cell = ws.cell(row=6, column=col_idx, value=count)
        count_cell.font = make_font(bold=True, color=accent, size=22)
        count_cell.fill = make_fill(lite)
        count_cell.alignment = make_align('center')
        ws.row_dimensions[6].height = 36

        ws.cell(row=7, column=col_idx, value='accounts').font = make_font(color=C_MUTED, size=8)
        ws.cell(row=7, column=col_idx).fill = make_fill(lite)
        ws.cell(row=7, column=col_idx).alignment = make_align('center')

        ws.cell(row=8, column=col_idx, value=desc).font = make_font(color=C_MUTED, size=7, italic=True)
        ws.cell(row=8, column=col_idx).fill = make_fill(lite)
        ws.cell(row=8, column=col_idx).alignment = make_align('center', wrap=True)

    # Totals section (row 10+)
    ws.row_dimensions[9].height = 12

    total = len(matched) + len(mismatched) + len(only_tally) + len(only_zoho)
    match_pct = (len(matched) / total * 100) if total else 0

    # Tally totals
    t_total_dr = sum(r['tally_dr'] for r in matched + mismatched) + \
                 sum(r['debit'] for r in only_tally)
    t_total_cr = sum(r['tally_cr'] for r in matched + mismatched) + \
                 sum(r['credit'] for r in only_tally)

    z_total_dr = sum(r['zoho_dr'] for r in matched + mismatched) + \
                 sum(r['debit'] for r in only_zoho)
    z_total_cr = sum(r['zoho_cr'] for r in matched + mismatched) + \
                 sum(r['credit'] for r in only_zoho)

    ws.row_dimensions[10].height = 20
    sum_headers = ['Metric', 'Tally (PDF)', 'Zoho Books (Excel)', 'Difference', 'Status']
    for col, hdr in enumerate(sum_headers, 1):
        c = ws.cell(row=10, column=col, value=hdr)
        c.fill = make_fill('1E293B')
        c.font = make_font(bold=True, color=C_WHITE, size=9)
        c.alignment = make_align('center')
        c.border = make_border()

    summary_data = [
        ('Total Debit',  t_total_dr, z_total_dr),
        ('Total Credit', t_total_cr, z_total_cr),
        ('Total Accounts', total, len(matched)+len(mismatched)+len(only_zoho)),
        ('Match Rate', None, None),
    ]
    for i, (label, t_val, z_val) in enumerate(summary_data, 11):
        ws.row_dimensions[i].height = 18
        ws.cell(row=i, column=1, value=label).font = make_font(bold=True, size=9)
        ws.cell(row=i, column=1).border = make_border()
        if t_val is not None:
            c = ws.cell(row=i, column=2, value=t_val)
            c.number_format = INR_FMT if isinstance(t_val, float) else '0'
            c.alignment = make_align('right')
            c.border = make_border()
            c = ws.cell(row=i, column=3, value=z_val)
            c.number_format = INR_FMT if isinstance(z_val, float) else '0'
            c.alignment = make_align('right')
            c.border = make_border()
            diff_val = abs(t_val - z_val) if isinstance(t_val, float) else abs(t_val - z_val)
            c = ws.cell(row=i, column=4, value=diff_val)
            c.number_format = INR_FMT if isinstance(diff_val, float) else '0'
            c.alignment = make_align('right')
            c.border = make_border()
            ok = diff_val < 1
            c = ws.cell(row=i, column=5, value=' OK' if ok else ' Differs')
            c.font = make_font(color=C_GREEN if ok else C_RED, bold=True, size=9)
            c.border = make_border()
        else:
            # Match rate row
            c = ws.cell(row=i, column=2, value=f'{match_pct:.1f}%')
            c.font = make_font(bold=True, color=C_GREEN if match_pct > 80 else C_ORANGE, size=12)
            c.alignment = make_align('center')
            c.border = make_border()
            ws.merge_cells(f'B{i}:E{i}')

    # Root cause analysis (row 16+)
    ws.row_dimensions[16].height = 10
    ws.row_dimensions[17].height = 20
    ws.cell(row=17, column=1, value=' ROOT CAUSE ANALYSIS').font = make_font(bold=True, color=C_WHITE, size=11)
    ws.cell(row=17, column=1).fill = make_fill('1E293B')
    ws.merge_cells('A17:G17')

    reversed_count = sum(1 for r in mismatched if r['root_cause'] == 'DR/CR Side Reversed')
    zero_zoho      = sum(1 for r in mismatched if 'Zero in Zoho' in r['root_cause'])
    partial        = sum(1 for r in mismatched if r['root_cause'] == 'Partial amount mismatch')

    causes = [
        (' DR/CR Side Reversed',        reversed_count, C_RED,    'Opening balance sign convention inverted during migration. Fix: Reverse DR/CR in Zoho → Accountant → Opening Balances.'),
        (' Partial Amount Mismatch',     partial,        C_ORANGE, 'Amounts differ but not reversed. May be due to TDS, rounding, or partial entries. Verify individually.'),
        ('⭕ Zero Balance in Zoho',        zero_zoho,      C_BLUE,   'Account exists in Zoho but opening balance was not set. Go to Zoho → Chart of Accounts → Opening Balances.'),
        (' Missing in Zoho',             len(only_tally),C_PURPLE, 'Ledgers from Tally not created in Zoho. Likely Sundry Debtors/Creditors rolled up into AR/AP.'),
    ]
    for i, (cause, count, color, fix) in enumerate(causes, 18):
        ws.row_dimensions[i].height = 36
        ws.cell(row=i, column=1, value=cause).font = make_font(bold=True, color=color, size=9)
        ws.cell(row=i, column=1).fill = make_fill('0F172A')
        ws.cell(row=i, column=1).border = make_border()
        c = ws.cell(row=i, column=2, value=count)
        c.font = make_font(bold=True, color=color, size=14)
        c.alignment = make_align('center')
        c.fill = make_fill('0F172A')
        c.border = make_border()
        c = ws.cell(row=i, column=3, value=fix)
        c.font = make_font(color='94A3B8', size=8)
        c.alignment = make_align('left', wrap=True)
        c.fill = make_fill('0F172A')
        c.border = make_border()
        ws.merge_cells(f'C{i}:G{i}')

    set_col_widths(ws, [28, 22, 22, 18, 16, 16, 16])
    return ws


def build_mismatch_sheet(wb, mismatched):
    ws = wb.create_sheet(' Mismatches')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    hdrs = ['#', 'Tally Account Name', 'Zoho Account Name',
            'Tally DR (₹)', 'Tally CR (₹)',
            'Zoho DR (₹)',  'Zoho CR (₹)',
            'Net Diff (₹)', 'Root Cause', 'Recommended Fix']
    write_header_row(ws, 1, hdrs, bg_color='7F1D1D', fg_color=C_WHITE)

    fixes = {
        'DR/CR Side Reversed':
            'Go to Zoho → Accountant → Opening Balances → Edit → Swap DR/CR amount',
        'Zero in Zoho – Opening balance not set':
            'Go to Zoho → Chart of Accounts → Opening Balance → Enter the Tally balance',
        'Zero in Tally – Entry may be missing':
            'Verify in Tally if the account had a balance on this date',
        'Partial amount mismatch':
            'Check for TDS deductions, rounding differences, or partial entries',
    }

    for i, r in enumerate(sorted(mismatched, key=lambda x: x['diff'], reverse=True), 2):
        bg = C_RED_LITE if i % 2 == 0 else C_WHITE
        diff_val = r['diff']
        severity_bg = 'FEE2E2' if diff_val >= 100000 else ('FFEDD5' if diff_val >= 10000 else 'FEF9C3')

        ws.cell(row=i, column=1, value=i-1).alignment = make_align('center')
        ws.cell(row=i, column=1).border = make_border()

        name_cell = ws.cell(row=i, column=2, value=r['tally_name'])
        name_cell.font = make_font(bold=True, size=9)
        name_cell.border = make_border()
        name_cell.alignment = make_align('left', wrap=True)

        ws.cell(row=i, column=3, value=r['zoho_name']).border = make_border()
        ws.cell(row=i, column=3).alignment = make_align('left', wrap=True)
        ws.cell(row=i, column=3).font = make_font(size=9, color=C_MUTED)

        for col, val, color in [
            (4, r['tally_dr'], '1D4ED8'),
            (5, r['tally_cr'], '065F46'),
            (6, r['zoho_dr'],  '2563EB'),
            (7, r['zoho_cr'],  '047857'),
        ]:
            c = ws.cell(row=i, column=col, value=val if val else 0.0)
            c.number_format = INR_FMT
            c.alignment = make_align('right')
            c.border = make_border()
            c.font = make_font(color=color if val and val > 0 else C_MUTED)

        diff_c = ws.cell(row=i, column=8, value=diff_val)
        diff_c.number_format = INR_FMT
        diff_c.alignment = make_align('right')
        diff_c.fill = make_fill(severity_bg)
        diff_c.font = make_font(bold=True, color=C_RED if diff_val >= 100000 else (C_ORANGE if diff_val >= 10000 else 'B45309'))
        diff_c.border = make_border()

        rc_cell = ws.cell(row=i, column=9, value=r['root_cause'])
        rc_cell.font = make_font(size=8, color=C_ORANGE)
        rc_cell.border = make_border()
        rc_cell.alignment = make_align('left', wrap=True)

        fix = fixes.get(r['root_cause'], 'Verify manually')
        fix_cell = ws.cell(row=i, column=10, value=fix)
        fix_cell.font = make_font(size=8, color=C_MUTED)
        fix_cell.border = make_border()
        fix_cell.alignment = make_align('left', wrap=True)

        ws.row_dimensions[i].height = 28

    set_col_widths(ws, [5, 30, 26, 16, 16, 16, 16, 16, 22, 35])
    ws.auto_filter.ref = f'A1:{get_column_letter(len(hdrs))}1'
    return ws


def build_matched_sheet(wb, matched):
    ws = wb.create_sheet(' Matched')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    hdrs = ['#', 'Tally Account Name', 'Zoho Account Name',
            'Tally DR (₹)', 'Tally CR (₹)', 'Zoho DR (₹)', 'Zoho CR (₹)', 'Status']
    write_header_row(ws, 1, hdrs, bg_color='14532D', fg_color=C_WHITE)

    for i, r in enumerate(matched, 2):
        ws.cell(row=i, column=1, value=i-1).alignment = make_align('center')
        ws.cell(row=i, column=1).border = make_border()
        ws.cell(row=i, column=2, value=r['tally_name']).border = make_border()
        ws.cell(row=i, column=2).font = make_font(bold=True, size=9)
        ws.cell(row=i, column=3, value=r['zoho_name']).border = make_border()
        ws.cell(row=i, column=3).font = make_font(size=9, color=C_MUTED)
        for col, val in [(4, r['tally_dr']), (5, r['tally_cr']),
                          (6, r['zoho_dr']),  (7, r['zoho_cr'])]:
            c = ws.cell(row=i, column=col, value=val if val else 0.0)
            c.number_format = INR_FMT
            c.alignment = make_align('right')
            c.border = make_border()
        ok_c = ws.cell(row=i, column=8, value=' MATCHED')
        ok_c.font = make_font(bold=True, color=C_GREEN, size=9)
        ok_c.fill = make_fill(C_GREEN_LITE)
        ok_c.alignment = make_align('center')
        ok_c.border = make_border()
        ws.row_dimensions[i].height = 16

    set_col_widths(ws, [5, 32, 28, 16, 16, 16, 16, 14])
    ws.auto_filter.ref = f'A1:{get_column_letter(len(hdrs))}1'
    return ws


def build_tally_only_sheet(wb, only_tally):
    ws = wb.create_sheet(' Tally Only')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    hdrs = ['#', 'Account Name (Tally)', 'Level', 'Tally DR (₹)', 'Tally CR (₹)', 'Net (₹)', 'Action Required']
    write_header_row(ws, 1, hdrs, bg_color='7C2D12', fg_color=C_WHITE)

    for i, r in enumerate(only_tally, 2):
        ws.cell(row=i, column=1, value=i-1).alignment = make_align('center')
        ws.cell(row=i, column=1).border = make_border()
        ws.cell(row=i, column=2, value=r['name']).font = make_font(bold=True, size=9)
        ws.cell(row=i, column=2).border = make_border()
        ws.cell(row=i, column=3, value=r['level']).alignment = make_align('center')
        ws.cell(row=i, column=3).border = make_border()
        ws.cell(row=i, column=4, value=r['debit'] or 0).number_format = INR_FMT
        ws.cell(row=i, column=4).alignment = make_align('right')
        ws.cell(row=i, column=4).border = make_border()
        ws.cell(row=i, column=5, value=r['credit'] or 0).number_format = INR_FMT
        ws.cell(row=i, column=5).alignment = make_align('right')
        ws.cell(row=i, column=5).border = make_border()
        net = (r['debit'] or 0) - (r['credit'] or 0)
        ws.cell(row=i, column=6, value=net).number_format = INR_FMT
        ws.cell(row=i, column=6).alignment = make_align('right')
        ws.cell(row=i, column=6).border = make_border()
        ws.cell(row=i, column=7, value='Create account in Zoho Books with this opening balance, OR verify it is rolled into Accounts Payable/Receivable')
        ws.cell(row=i, column=7).font = make_font(size=8, color=C_MUTED)
        ws.cell(row=i, column=7).border = make_border()
        ws.cell(row=i, column=7).alignment = make_align('left', wrap=True)
        ws.row_dimensions[i].height = 22

    set_col_widths(ws, [5, 36, 8, 16, 16, 16, 45])
    ws.auto_filter.ref = f'A1:{get_column_letter(len(hdrs))}1'
    return ws


def build_zoho_only_sheet(wb, only_zoho):
    ws = wb.create_sheet(' Zoho Only')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    hdrs = ['#', 'Account Name (Zoho)', 'Zoho DR (₹)', 'Zoho CR (₹)', 'Net (₹)', 'Explanation']
    write_header_row(ws, 1, hdrs, bg_color='1E3A5F', fg_color=C_WHITE)

    explanations = {
        'accounts receivable': 'Aggregate of all Tally Sundry Debtors',
        'accounts payable':    'Aggregate of all Tally Sundry Creditors',
        'inventory asset':     'Maps to Opening Stock / Closing Stock in Tally',
        'opening balance':     'Zoho auto-entry to balance books on import',
        'retained earnings':   'Zoho system account for accumulated profit/loss',
    }

    for i, r in enumerate(only_zoho, 2):
        ws.cell(row=i, column=1, value=i-1).alignment = make_align('center')
        ws.cell(row=i, column=1).border = make_border()
        ws.cell(row=i, column=2, value=r['name']).font = make_font(bold=True, size=9)
        ws.cell(row=i, column=2).border = make_border()
        ws.cell(row=i, column=3, value=r['debit'] or 0).number_format = INR_FMT
        ws.cell(row=i, column=3).alignment = make_align('right')
        ws.cell(row=i, column=3).border = make_border()
        ws.cell(row=i, column=4, value=r['credit'] or 0).number_format = INR_FMT
        ws.cell(row=i, column=4).alignment = make_align('right')
        ws.cell(row=i, column=4).border = make_border()
        net = (r['debit'] or 0) - (r['credit'] or 0)
        ws.cell(row=i, column=5, value=net).number_format = INR_FMT
        ws.cell(row=i, column=5).alignment = make_align('right')
        ws.cell(row=i, column=5).border = make_border()
        name_lower = r['name'].lower()
        expl = next((v for k, v in explanations.items() if k in name_lower),
                    'Zoho-specific account — verify if it exists in Tally under a different name')
        ws.cell(row=i, column=6, value=expl).font = make_font(size=8, color=C_MUTED)
        ws.cell(row=i, column=6).border = make_border()
        ws.cell(row=i, column=6).alignment = make_align('left', wrap=True)
        ws.row_dimensions[i].height = 20

    set_col_widths(ws, [5, 38, 18, 18, 18, 45])
    ws.auto_filter.ref = f'A1:{get_column_letter(len(hdrs))}1'
    return ws


def build_raw_tally_sheet(wb, rows, company, period):
    ws = wb.create_sheet(' Raw Tally PDF')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    ws.merge_cells('A1:E1')
    ws['A1'] = f'{company}  |  Trial Balance  |  {period}'
    ws['A1'].font = make_font(bold=True, color=C_WHITE, size=11)
    ws['A1'].fill = make_fill(C_DARK_BG)
    ws['A1'].alignment = make_align('center')
    ws.row_dimensions[1].height = 26

    hdrs = ['#', 'Account Name', 'Level', 'Debit (₹)', 'Credit (₹)']
    write_header_row(ws, 2, hdrs, bg_color='1E293B')

    for i, r in enumerate(rows, 3):
        indent = '  ' * r['level']
        ws.cell(row=i, column=1, value=i-2).alignment = make_align('center')
        ws.cell(row=i, column=1).border = make_border()
        name_c = ws.cell(row=i, column=2, value=indent + r['name'])
        name_c.font = make_font(bold=(r['level'] == 0), size=9,
                                color='FFFFFF' if r['level'] == 0 else ('CBD5E1' if r['level'] == 1 else C_MUTED))
        name_c.fill = make_fill('0F172A' if r['level'] == 0 else ('111827' if r['level'] == 1 else 'FFFFFF'))
        name_c.border = make_border()
        ws.cell(row=i, column=3, value=r['level']).alignment = make_align('center')
        ws.cell(row=i, column=3).border = make_border()
        ws.cell(row=i, column=4, value=r['debit'] or 0).number_format = INR_FMT
        ws.cell(row=i, column=4).alignment = make_align('right')
        ws.cell(row=i, column=4).border = make_border()
        if r['debit'] and r['debit'] > 0:
            ws.cell(row=i, column=4).font = make_font(color='1D4ED8')
        ws.cell(row=i, column=5, value=r['credit'] or 0).number_format = INR_FMT
        ws.cell(row=i, column=5).alignment = make_align('right')
        ws.cell(row=i, column=5).border = make_border()
        if r['credit'] and r['credit'] > 0:
            ws.cell(row=i, column=5).font = make_font(color='065F46')
        ws.row_dimensions[i].height = 15

    set_col_widths(ws, [5, 42, 8, 18, 18])
    ws.auto_filter.ref = 'A2:E2'
    return ws


def build_raw_zoho_sheet(wb, rows, company, period):
    ws = wb.create_sheet(' Raw Zoho Excel')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    ws.merge_cells('A1:E1')
    ws['A1'] = f'{company}  |  Trial Balance  |  {period}'
    ws['A1'].font = make_font(bold=True, color=C_WHITE, size=11)
    ws['A1'].fill = make_fill('0D2A1F')
    ws['A1'].alignment = make_align('center')
    ws.row_dimensions[1].height = 26

    hdrs = ['#', 'Account Name', 'Indent Level', 'Net Debit (₹)', 'Net Credit (₹)']
    write_header_row(ws, 2, hdrs, bg_color='14532D')

    for i, r in enumerate(rows, 3):
        indent = '  ' * min(r['level'], 6)
        ws.cell(row=i, column=1, value=i-2).alignment = make_align('center')
        ws.cell(row=i, column=1).border = make_border()
        name_c = ws.cell(row=i, column=2, value=indent + r['name'])
        name_c.font = make_font(size=9)
        name_c.border = make_border()
        ws.cell(row=i, column=3, value=r['level']).alignment = make_align('center')
        ws.cell(row=i, column=3).border = make_border()
        ws.cell(row=i, column=4, value=r['debit'] or 0).number_format = INR_FMT
        ws.cell(row=i, column=4).alignment = make_align('right')
        ws.cell(row=i, column=4).border = make_border()
        if r['debit'] and r['debit'] > 0:
            ws.cell(row=i, column=4).font = make_font(color='1D4ED8')
        ws.cell(row=i, column=5, value=r['credit'] or 0).number_format = INR_FMT
        ws.cell(row=i, column=5).alignment = make_align('right')
        ws.cell(row=i, column=5).border = make_border()
        if r['credit'] and r['credit'] > 0:
            ws.cell(row=i, column=5).font = make_font(color='065F46')
        ws.row_dimensions[i].height = 15

    set_col_widths(ws, [5, 42, 10, 18, 18])
    ws.auto_filter.ref = 'A2:E2'
    return ws

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print('=' * 60)
    print('  Tally PDF ↔ Zoho Excel — Reconciliation Engine')
    print('=' * 60)

    print(f'\n[1/4] Parsing Tally PDF: {PDF_PATH}')
    t_company, t_period, tally_rows = parse_tally_pdf(PDF_PATH)
    print(f'       {len(tally_rows)} rows  |  Company: {t_company}  |  Period: {t_period}')

    print(f'\n[2/4] Parsing Zoho Excel: {EXCEL_PATH}')
    z_company, z_period, zoho_rows = parse_zoho_excel(EXCEL_PATH)
    print(f'       {len(zoho_rows)} rows  |  Company: {z_company}  |  Period: {z_period}')

    print('\n[3/4] Running reconciliation...')
    matched, mismatched, only_tally, only_zoho = reconcile(tally_rows, zoho_rows)
    total = len(matched) + len(mismatched) + len(only_tally) + len(only_zoho)
    match_pct = len(matched) / total * 100 if total else 0

    print(f'       Matched:      {len(matched):4d}  ({match_pct:.1f}%)')
    print(f'       Mismatched:   {len(mismatched):4d}')
    print(f'       Tally only:   {len(only_tally):4d}')
    print(f'       Zoho only:    {len(only_zoho):4d}')

    print(f'\n[4/4] Generating Excel report: {OUT_PATH}')
    wb = openpyxl.Workbook()

    build_summary_sheet(wb, matched, mismatched, only_tally, only_zoho,
                        t_company, t_period, z_company, z_period)
    build_mismatch_sheet(wb, mismatched)
    build_matched_sheet(wb, matched)
    build_tally_only_sheet(wb, only_tally)
    build_zoho_only_sheet(wb, only_zoho)
    build_raw_tally_sheet(wb, tally_rows, t_company, t_period)
    build_raw_zoho_sheet(wb, zoho_rows, z_company, z_period)

    wb.save(OUT_PATH)
    print(f'\n Report saved: {OUT_PATH}')
    print(f'   Sheets:')
    for s in wb.sheetnames:
        print(f'     • {s}')
    print()

if __name__ == '__main__':
    main()
