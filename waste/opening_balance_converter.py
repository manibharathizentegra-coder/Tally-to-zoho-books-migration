"""
Opening Balance Converter
=========================
Supports TWO input types:
  1. Excel / CSV  (Tally exported spreadsheet)
  2. PDF          (Tally Trial Balance printed PDF)

Both produce the same Zoho Books opening balance import format:
  Migration Date | Account Name | Debit or Credit | Currency Code | Amount | Exchange Rate | Contact Name
"""

import re
import io
import json
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS & CONFIG
# ─────────────────────────────────────────────────────────────────────────────

ZOHO_HEADERS = [
    "Migration Date",
    "Account Name",
    "Debit or Credit",
    "Currency Code",
    "Amount",
    "Exchange Rate",
    "Contact Name",
]

# Rows to skip (case-insensitive startswith check)
SKIP_PREFIXES = {
    "carried over", "brought forward", "grand total", "total",
    "sub total", "sub-total", "closing balance", "opening balance",
    "trial balance", "particulars", "debit", "credit", "page",
    "for ", "date", "balance sheet", "profit & loss", "profit and loss",
    "continued", "amount", "curr.", "nett"
}

# Skip if entire cell is just numbers / dashes / empty
_PURE_NUM = re.compile(r"^[\d,.\-\s]*$")


def _should_skip(account: str) -> bool:
    """Check if an account name line should be skipped."""
    low = account.strip().lower()
    if not low or len(low) < 2:
        return True
    # If it's just a number (like page number or a stray table value)
    if _PURE_NUM.match(low):
        return True
    # Check prefixes
    for prefix in SKIP_PREFIXES:
        if low.startswith(prefix):
            return True
    return False


def _parse_indian_amount(text: str) -> float:
    """Parse Indian number format: '5,72,33,448.28' → 57233448.28"""
    if not text:
        return 0.0
    text = str(text).strip()
    # Remove currency symbols and internal spaces
    text = re.sub(r"[₹$\s]", "", text)
    # Remove commas
    text = text.replace(",", "")
    # Remove Dr/Cr suffix if present
    text = re.sub(r"(?i)(dr|cr)$", "", text).strip()
    
    if not text or text == "-":
        return 0.0
        
    try:
        return abs(float(text))
    except ValueError:
        return 0.0


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL / CSV PARSER
# ─────────────────────────────────────────────────────────────────────────────

def _load_spreadsheet_rows(data: bytes, filename: str):
    ext = filename.rsplit(".", 1)[-1].lower()
    rows = []
    if ext in ("xlsx", "xls"):
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(tuple("" if c is None else str(c).strip() for c in row))
    elif ext == "csv":
        import csv
        text = data.decode("utf-8-sig", errors="replace")
        for row in csv.reader(text.splitlines()):
            rows.append(tuple(c.strip() for c in row))
    else:
        raise ValueError(f"Unsupported file type: .{ext}")
    return rows


def parse_excel(data: bytes, filename: str):
    """Parse Tally exported Excel/CSV → list of {account, debit, credit}"""
    rows = _load_spreadsheet_rows(data, filename)

    # Auto-detect header row (find row with Debit + Credit in same row)
    particulars_col = 0
    debit_col = 1
    credit_col = 2
    data_start = 0

    # Scan first 30 rows for headers
    for i, row in enumerate(rows[:30]):
        cells = [str(c).lower() for c in row]
        has_dr = any("debit" in c for c in cells)
        has_cr = any("credit" in c for c in cells)
        
        if has_dr and has_cr:
            # We found the header row! Auto-map columns
            for j, c in enumerate(cells):
                if any(k in c for k in ["particular", "account", "ledger", "name"]):
                    particulars_col = j
                if "debit" in c:
                    debit_col = j
                if "credit" in c:
                    credit_col = j
            data_start = i + 1
            break

    entries = []
    for row in rows[data_start:]:
        # Safety check: row must be long enough
        if len(row) <= max(particulars_col, debit_col, credit_col):
            continue
            
        acct = str(row[particulars_col]).strip()
        if _should_skip(acct):
            continue
            
        d_val = row[debit_col]  if debit_col  < len(row) else ""
        c_val = row[credit_col] if credit_col < len(row) else ""
        
        dr = _parse_indian_amount(d_val)
        cr = _parse_indian_amount(c_val)
        
        if dr == 0 and cr == 0:
            continue
            
        entries.append({"account": acct, "debit": round(dr, 2), "credit": round(cr, 2)})

    return entries


# ─────────────────────────────────────────────────────────────────────────────
# PDF PARSER  (Tally Trial Balance printed PDF)
# ─────────────────────────────────────────────────────────────────────────────

def parse_pdf(data: bytes):
    """
    Parse Tally Trial Balance PDF using pdfplumber.
    """
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("pdfplumber is required for PDF parsing. Please install it.")

    entries = []
    _number_re = re.compile(r"^[\d,]+\.?\d*$")

    with pdfplumber.open(io.BytesIO(data)) as pdf:
        
        # We need to detect Column X-Positions. 
        # Tally Header usually: "Particulars ... Debit ... Credit"
        debit_x  = None
        credit_x = None

        # Pass 1: Scan pages to find Header Positions
        for page in pdf.pages:
            words = page.extract_words(keep_blank_chars=False)
            for w in words:
                txt = w["text"].strip().lower()
                if txt == "debit" and not debit_x:
                    debit_x = w["x0"]
                if txt == "credit" and not credit_x:
                    credit_x = w["x0"]
            if debit_x and credit_x:
                break
        
        # Fallback if headers not found (e.g. only on page 1 and we missed it?)
        # Tally standard layout logic: 
        # - Credit is usually near right margin
        # - Debit is to the left of Credit
        if not credit_x: credit_x = 500  # A4 width is ~595. 500 is right-ish
        if not debit_x:  debit_x  = 400

        # Pass 2: Extract Data
        for page in pdf.pages:
            # Extract words for Y-clustering and column headers
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words: continue

            # Group words by Line (Y-axis)
            lines = {}
            for w in words:
                y = round(w["top"])
                if y not in lines: lines[y] = []
                lines[y].append(w)

            # Sort lines top-to-bottom
            for y in sorted(lines.keys()):
                line_words = sorted(lines[y], key=lambda w: w["x0"])
                
                # Determine vertical bounds of this line for cropping
                min_top = min(w["top"] for w in line_words)
                max_bot = max(w["bottom"] for w in line_words)
                
                # Dynamic Thresholds
                # Name <--- thresh ---> Debit <--- thresh ---> Credit
                dr_thresh = debit_x - 20   # tight buffer to maximize name area
                cr_thresh = credit_x - 20

                debit_parts  = []
                credit_parts = []
                
                # Identify Amounts using words (safer to isolate numbers)
                for w in line_words:
                    if w["x0"] >= cr_thresh:
                        credit_parts.append(w["text"])
                    elif w["x0"] >= dr_thresh:
                        debit_parts.append(w["text"])
                
                # Extract Name using CROP to preserve EXACT spacing (e.g. "M K  SILK")
                # Crop area: Left=0, Top=min_top, Right=dr_thresh, Bottom=max_bot
                try:
                    # Crop requires (x0, top, x1, bottom)
                    # Ensure dimensions are valid
                    if dr_thresh > 0 and max_bot > min_top:
                        name_crop = page.crop((0, min_top, dr_thresh, max_bot))
                        raw_name = name_crop.extract_text(x_tolerance=2) or ""
                        # extract_text might return newlines if fonts vary slightly, flatten it
                        account = raw_name.replace("\n", " ").strip()
                    else:
                        account = ""
                except Exception:
                    # Fallback to word join if crop fails
                    name_parts = [w["text"] for w in line_words if w["x0"] < dr_thresh]
                    account = " ".join(name_parts).strip()

                dr_str = " ".join(debit_parts).strip()
                cr_str = " ".join(credit_parts).strip()

                if _should_skip(account):
                    continue

                dr = _parse_indian_amount(dr_str)
                cr = _parse_indian_amount(cr_str)

                if dr == 0 and cr == 0:
                    continue

                entries.append({
                    "account": account,
                    "debit":   round(dr, 2),
                    "credit":  round(cr, 2),
                })

    return entries


# ─────────────────────────────────────────────────────────────────────────────
# BUILD ZOHO EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def build_zoho_excel(entries, migration_date=None):
    if not migration_date:
        migration_date = date.today().strftime("%d/%m/%Y")
    else:
        # Ensure DD/MM/YYYY format if input is YYYY-MM-DD
        if "-" in migration_date:
            try:
                y, m, d = migration_date.split("-")
                migration_date = f"{d}/{m}/{y}"
            except: pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opening Balances"

    # Column widths
    col_widths = [16, 45, 16, 16, 18, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Styled Header
    hdr_fill  = PatternFill("solid", fgColor="5E35B1")
    hdr_font  = Font(color="FFFFFF", bold=True, size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_i, h in enumerate(ZOHO_HEADERS, 1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
    
    ws.row_dimensions[1].height = 24

    # Data Rows
    alt_fill  = PatternFill("solid", fgColor="F3F0FA")
    num_fmt   = '#,##0.00'
    dr_font   = Font(color="1B5E20") # Green for Debit
    cr_font   = Font(color="B71C1C") # Red for Credit

    current_row = 2
    for entry in entries:
        # Determine if we write 1 or 2 rows
        rows_to_write = []
        if entry["debit"]  > 0: rows_to_write.append(("Debit",  entry["debit"]))
        if entry["credit"] > 0: rows_to_write.append(("Credit", entry["credit"]))

        for dc_text, amount in rows_to_write:
            fill = alt_fill if current_row % 2 == 0 else None
            
            # Row Values: matches Zoho import format exactly
            row_data = [
                migration_date,     # Migration Date
                entry["account"],   # Account Name
                dc_text,            # Debit or Credit
                "INR",              # Currency
                amount,             # Amount
                1,                  # Exchange Rate
                ""                  # Contact Name
            ]

            for col_i, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_i, value=val)
                if fill: cell.fill = fill
                
                # Styling
                cell.alignment = Alignment(vertical="center")
                if col_i == 3: # Dr/Cr text
                    cell.font = dr_font if dc_text == "Debit" else cr_font
                if col_i == 5: # Amount
                    cell.number_format = num_fmt
            
            current_row += 1

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()




def _get_signed_amount(val):
    """
    Tally JSON amounts are floats/ints (not strings) and can be negative.
    Negative means the balance is on the OPPOSITE side:
       dspcldramta = -5000  → means Credit of 5000
       dspclcramta = -5000  → means Debit of 5000
    Returns (debit_component, credit_component)
    """
    if val is None:
        return 0.0, 0.0
    try:
        amount = float(val)
    except (TypeError, ValueError):
        # Maybe it's a string like "5,72,33,448.28"
        amount = _parse_indian_amount(str(val))
        # String amounts are always positive
        return amount, 0.0

    return (abs(amount), 0.0) if amount >= 0 else (0.0, abs(amount))


def _extract_amounts_from_info(infos):
    """
    Extract debit/credit from a dspaccinfo list.
    Handles BOTH Tally export formats:
      Format A – Closing Balance:  dspcldramt / dspcldramta  &  dspclcramt / dspclcramta
      Format B – Opening Balance:  dspopdramt / dspopdramta  &  dspopcramt / dspopcramta

    Negative value in the Debit column → treat as Credit (Tally sign convention).
    Negative value in the Credit column → treat as Debit.
    """
    debit_total  = 0.0
    credit_total = 0.0

    if not isinstance(infos, list):
        infos = [infos] if infos else []

    for info in infos:
        if not isinstance(info, dict):
            continue

        # ── Format A: Closing Balance keys ──────────────────────────────────
        dr_obj_cl = info.get("dspcldramt", None)
        cr_obj_cl = info.get("dspclcramt", None)

        # ── Format B: Opening Balance keys ──────────────────────────────────
        dr_obj_op = info.get("dspopdramt", None)
        cr_obj_op = info.get("dspopcramt", None)

        # Pick whichever format is present in this info block
        # (some files mix both — handle independently)
        for dr_obj, dr_key, cr_obj, cr_key in [
            (dr_obj_cl, "dspcldramta", cr_obj_cl, "dspclcramta"),
            (dr_obj_op, "dspopdramta", cr_obj_op, "dspopcramta"),
        ]:
            if isinstance(dr_obj, dict):
                dr_raw = dr_obj.get(dr_key, None)
                dr_pos, dr_neg = _get_signed_amount(dr_raw)
                debit_total  += dr_pos
                credit_total += dr_neg

            if isinstance(cr_obj, dict):
                cr_raw = cr_obj.get(cr_key, None)
                cr_pos, cr_neg = _get_signed_amount(cr_raw)
                credit_total += cr_pos
                debit_total  += cr_neg

    return round(debit_total, 2), round(credit_total, 2)


def parse_json(file_bytes):
    """
    Parse Tally JSON export. Supports TWO export formats:

    FORMAT A – Closing Balance (e.g. Tally_TrialBal.json):
    {
      "dspaccbody": { "dspaccline": [
          { "dspaccname": {"dspdispname": "Name"},
            "dspaccinfo": [{"dspcldramt": {"dspcldramta": <float>},
                            "dspclcramt": {"dspclcramta": <float>}}]
          }, ...
      ]}
    }
    → FLAT list of ~499 ledgers

    FORMAT B – Opening Balance (e.g. TrialBal.json):
    {
      "dspaccbody": { "dspaccline": [
          { "dspaccname": {"dspdispname": "Capital Account"},
            "dspaccinfo": [{"dspopdramt": {}, "dspopcramt": {"dspopcramta": <float>}}],
            "grpexplosion": { "dspaccline": [
                { "dspaccname": {"dspdispname": "Reserve & Surplus"}, ...
                  "grpexplosion": { "dspaccline": [ ... ] }   ← further nesting
                }
            ]}
          }, ...
      ]}
    }
    → NESTED tree – only leaf nodes (no grpexplosion) are actual Ledgers.

    KEY RULE: We skip any entry that has a "grpexplosion" key (those are Group
    Headers with subtotals), and only collect LEAF nodes as real ledger entries.
    """
    entries = []
    try:
        # ── Decode file (Tally exports UTF-16 with BOM) ──────────────────────
        text = ""
        for enc in ('utf-8-sig', 'utf-16', 'utf-16-le', 'utf-16-be', 'latin-1'):
            try:
                text = file_bytes.decode(enc)
                break
            except (UnicodeDecodeError, Exception):
                continue

        if not text:
            print("JSON parse warning: could not decode file")
            return []

        data = json.loads(text)

        # ── Recursive walker: collect ALL dspaccline items at every depth ────
        def walk_lines(node, result):
            """Recursively walk dspaccline lists at every nesting level."""
            if isinstance(node, list):
                for item in node:
                    walk_lines(item, result)
            elif isinstance(node, dict):
                if "dspaccname" in node:
                    result.append(node)
                # Continue into grpexplosion or any nested dspaccline
                for key in ("grpexplosion",):
                    if key in node:
                        sub = node[key]
                        if isinstance(sub, dict) and "dspaccline" in sub:
                            walk_lines(sub["dspaccline"], result)
                        else:
                            walk_lines(sub, result)
                # Also recurse directly into dspaccline if present here
                if "dspaccline" in node and "dspaccname" not in node:
                    walk_lines(node["dspaccline"], result)

        # Start from the known root path
        all_items = []
        try:
            root_lines = data['dspaccbody']['dspaccline']
            walk_lines(root_lines, all_items)
        except (KeyError, TypeError):
            # Fallback: blind recursive collect
            def collect_blind(obj, result):
                if isinstance(obj, dict):
                    if "dspaccname" in obj:
                        result.append(obj)
                    for v in obj.values():
                        collect_blind(v, result)
                elif isinstance(obj, list):
                    for item in obj:
                        collect_blind(item, result)
            collect_blind(data, all_items)

        if not all_items:
            print("JSON parse warning: No account entries found in entire tree")
            return []

        print(f"JSON parse: found {len(all_items)} raw nodes (incl. group headers)")

        for item in all_items:
            # 1. Skip Group Headers that have sub-items (grpexplosion)
            #    Group headers carry subtotals — we want only leaf ledgers.
            #    However: if a group has NO sub-ledgers (leaf group), keep it.
            has_explosion = (
                "grpexplosion" in item and
                isinstance(item.get("grpexplosion"), dict) and
                item["grpexplosion"].get("dspaccline")
            )
            if has_explosion:
                continue   # skip — it's a group summary row

            # 2. Extract Name
            name_obj = item.get("dspaccname", {})
            if not isinstance(name_obj, dict):
                continue
            raw_name = name_obj.get("dspdispname", "")
            if not raw_name:
                continue
            account = str(raw_name).strip()
            if _should_skip(account):
                continue

            # 3. Extract Amounts (handles both Format A and B keys)
            infos = item.get("dspaccinfo", [])
            debit_total, credit_total = _extract_amounts_from_info(infos)

            if debit_total == 0 and credit_total == 0:
                continue

            entries.append({
                "account": account,
                "debit":   debit_total,
                "credit":  credit_total,
            })

    except Exception as e:
        print(f"JSON Parsing Error: {e}")
        import traceback
        traceback.print_exc()
        return []

    print(f"JSON parse result: {len(entries)} valid ledger entries")
    return entries



# ─────────────────────────────────────────────────────────────────────────────
# FILTERS & DEDUPLICATION
# ─────────────────────────────────────────────────────────────────────────────

TALLY_DEFAULT_GROUPS = {
    "branch / divisions", "capital account", "reserves & surplus", "current assets",
    "bank accounts", "cash-in-hand", "deposits (asset)", "loans & advances (asset)",
    "stock-in-hand", "sundry debtors", "current liabilities", "duties & taxes",
    "provisions", "sundry creditors", "fixed assets", "investments", "loans (liability)",
    "bank od a/c", "secured loans", "unsecured loans", "suspense a/c", "misc. expenses (asset)",
    "sales accounts", "purchase accounts", "direct incomes", "indirect incomes",
    "direct expenses", "indirect expenses", "profit & loss a/c", "difference in opening balances"
}

def filter_entries(entries, db_group_names=None):
    """
    Remove duplicates and Group Headers.
    1. Filter out known Tally Groups (e.g. Current Assets)
    2. Filter out Groups from DB (if provided)
    3. Deduplicate exact (Account, Amount, Type) rows
    """
    filtered = []
    seen = set()
    
    # Normalize DB groups
    groups_to_skip = set(TALLY_DEFAULT_GROUPS)
    if db_group_names:
        groups_to_skip.update(n.lower().strip() for n in db_group_names if n)

    for e in entries:
        name = e["account"]
        norm_name = name.lower().strip()
        
        # Rule 1: Skip if it is a Group Name
        if norm_name in groups_to_skip:
            continue
            
        # Rule 2: Skip exact duplicates
        # Tuple: (LowName, Debit, Credit)
        key = (norm_name, e["debit"], e["credit"])
        if key in seen:
            continue
        seen.add(key)
        
        filtered.append(e)
        
    return filtered


# ─────────────────────────────────────────────────────────────────────────────
# MAIN API
# ─────────────────────────────────────────────────────────────────────────────

def convert(file_bytes, filename, migration_date=None, db_group_names=None):
    """
    Main entry point. Auto-detects Excel vs PDF vs JSON.
    Returns (output_bytes, summary_dict, errors_list)
    """
    errors = []
    
    # Simple extension check
    filename_lower = filename.lower()
    
    try:
        if filename_lower.endswith(".pdf"):
            entries = parse_pdf(file_bytes)
            file_type = "PDF"
        elif filename_lower.endswith(".json"):
            entries = parse_json(file_bytes)
            file_type = "JSON"
        else:
            entries = parse_excel(file_bytes, filename)
            file_type = "EXCEL"
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, {}, [f"Parse Error: {str(e)}"]

    if not entries:
        return None, {}, [
            "No valid data found. Please check:\n"
            "1. File contains valid Account Names and Balances\n"
            "2. Balances are not all zero"
        ]

    # Apply Filters (Groups & Dupes)
    original_count = len(entries)
    filtered_entries = filter_entries(entries, db_group_names)
    
    if not filtered_entries:
        # Diagnostic: Why did we lose everything?
        sample_names = [e['account'] for e in entries[:5]]
        msg = f"Found {original_count} entries, but ALL were filtered out as Groups or Duplicates.\n"
        msg += "Likely cause: File contains only Group Summaries (e.g. 'Capital Account') instead of distinct Ledgers.\n"
        msg += "Please export 'Detailed' view or 'Ledgers' only.\n"
        msg += f"Samples filtered: {', '.join(sample_names)}"
        return None, {}, [msg]
        
    entries = filtered_entries

    # Generate Output
    output_bytes = build_zoho_excel(entries, migration_date)

    # Summary Stats
    zoho_rows = sum((1 if e["debit"]>0 else 0) + (1 if e["credit"]>0 else 0) for e in entries)
    total_dr  = round(sum(e["debit"] for e in entries), 2)
    total_cr  = round(sum(e["credit"] for e in entries), 2)

    # Preview (First 10)
    preview = []
    for e in entries[:10]:
        val = e["debit"] if e["debit"] > 0 else e["credit"]
        typ = "Debit" if e["debit"] > 0 else "Credit"
        preview.append((e["account"], typ, val))
        
    summary = {
        "accounts_found": len(entries),
        "original_found": original_count,
        "zoho_rows":      zoho_rows,
        "total_debit":    total_dr,
        "total_credit":   total_cr,
        "file_type":      file_type,
        "preview":        preview,
        "col_detected":   {"particulars":0, "debit":1, "credit":2} 
    }

    return output_bytes, summary, errors
